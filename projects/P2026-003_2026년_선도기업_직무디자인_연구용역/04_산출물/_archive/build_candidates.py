# -*- coding: utf-8 -*-
"""해양환경공단·KF 적합 직무 후보군 도출 결과(6탭 Excel + 보고서) 생성."""
import os, sys, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SKILL = "/sessions/sweet-intelligent-newton/mnt/P2026-003_2026년_선도기업_직무디자인_연구용역/.claude/skills/candidate-derivation/scripts"
sys.path.insert(0, SKILL)
from report_sheets import (build_guide_sheet, build_summary_sheet, build_selection_form,
                           set_print, VTNAME, LFILL)

GR_KO = {"A": "직무 후보", "B": "과업 요소", "C": "개선 제안", "D": "보류"}
GR_FILL = {"직무 후보": "D9EAD3", "과업 요소": "D9EAD3", "개선 제안": "FFF2CC", "보류": "FCE4D6"}
WHITE = Font(bold=True, color="FFFFFF", size=10); SECT = "305496"
WRAP = Alignment(wrap_text=True, vertical="top"); CTR = Alignment(horizontal="center", vertical="center")
_th = Side(style="thin", color="BFBFBF"); BD = Border(_th, _th, _th, _th)
def fill(h): return PatternFill("solid", fgColor=h)
def header(ws, row, cols):
    for i, t in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=t); c.font = WHITE; c.fill = fill(SECT); c.alignment = CTR; c.border = BD

def read_pool(path):
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb["아이디어Pool"]
    rows = list(ws.iter_rows(values_only=True)); hi = 0
    for i, r in enumerate(rows[:6]):
        j = " ".join(str(c) for c in r if c)
        if "연번" in j: hi = i; break
    items = {}; order = []
    for r in rows[hi + 1:]:
        if not r or r[0] is None: continue
        try: num = int(r[0])
        except: continue
        idea = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        desc = str(r[2]).strip().replace("\n", " ") if len(r) > 2 and r[2] else ""
        if not idea: continue
        items[num] = {"idea": idea, "desc": desc}; order.append(num)
    wb.close()
    return items, order


def build(company, pool_path, candidates, grade_map, cd_detail, outdir, ws_info):
    os.makedirs(outdir, exist_ok=True)
    items, order = read_pool(pool_path)
    n = len(order)
    today = datetime.date.today().strftime("%Y%m%d")

    idea2cand = {}
    for c in candidates:
        for num in c["nums"]:
            idea2cand[num] = c
    dist = {g: 0 for g in "ABCD"}
    for num in order: dist[grade_map[num]] += 1
    vt_counts = {k: 0 for k in "①②③④⑤"}
    for c in candidates: vt_counts[c["vt"]] += 1

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    build_guide_sheet(wb, company, n_input=n, index=0)
    build_summary_sheet(wb, company, n, dist, vt_counts=vt_counts,
                        recs=[(c["code"], c["name"], c["vt"], c["mark"]) for c in candidates if c["mark"]], index=1)

    s1 = wb.create_sheet("3. 전체 아이디어 분류")
    header(s1, 1, ["연번", "워크숍 아이디어", "설명", "직무화 판정", "5대 가치 유형", "통합 직무코드", "통합 직무명", "비고"])
    for num in order:
        g = grade_map[num]; ko = GR_KO[g]
        if g in ("A", "B"):
            c = idea2cand[num]; vt = VTNAME[c["vt"]]; code = c["code"]; cname = c["name"]; note = ""
        else:
            vt = "—"; code = ""; cname = ""; note = cd_detail[num][1]
        it = items[num]
        s1.append([num, it["idea"], it["desc"], ko, vt, code, cname, note])
    for col, w in zip("ABCDEFGH", [5, 30, 40, 11, 20, 11, 30, 26]): s1.column_dimensions[col].width = w
    for row in s1.iter_rows(min_row=2):
        for c in row: c.alignment = WRAP; c.border = BD
        row[0].alignment = CTR; row[3].alignment = CTR; row[5].alignment = CTR
        row[3].fill = fill(GR_FILL[row[3].value])
        if row[4].value != "—": row[4].fill = fill(LFILL[row[4].value[0]])
    s1.freeze_panes = "A2"; set_print(s1)

    s2 = wb.create_sheet("4. 직무 후보군"); s2.sheet_properties.tabColor = "FFC000"
    header(s2, 1, ["직무코드", "5대 가치 유형", "직무명(안)", "직무 정의(안)", "통합 아이디어",
                   "직무 특성·장애 적합성", "컨설턴트 추천", "추천 근거 / 검토"])
    for c in candidates:
        merged = "\n".join(f"{num} {items[num]['idea']}" for num in c["nums"])
        basis = c.get("basis", "")
        if c.get("flag"): basis = (basis + (" / " if basis else "") + "검토: " + c["flag"])
        if not basis: basis = "-"
        s2.append([c["code"], VTNAME[c["vt"]], c["name"], c["defn"], merged, c["fit"], c["mark"] or "", basis])
    for col, w in zip("ABCDEFGH", [8, 18, 24, 44, 30, 30, 11, 28]): s2.column_dimensions[col].width = w
    for row in s2.iter_rows(min_row=2):
        for c in row: c.alignment = WRAP; c.border = BD
        row[0].alignment = CTR; row[6].alignment = CTR
        row[1].fill = fill(LFILL[row[1].value[0]])
        if row[6].value == "◎": row[6].font = Font(bold=True, color="C00000")
        elif row[6].value == "○": row[6].font = Font(bold=True, color="2E75B6")
    s2.freeze_panes = "A2"; set_print(s2)

    s3 = wb.create_sheet("5. 개선 제안·보류")
    header(s3, 1, ["연번", "워크숍 아이디어", "구분", "직무 후보 제외 사유", "활용 방안"])
    cd_nums = [num for num in order if grade_map[num] in ("C", "D")]
    for num in cd_nums:
        g = grade_map[num]; label = GR_KO[g]; _, reason, plan = cd_detail[num]
        s3.append([num, items[num]["idea"], label, reason, plan])
    if not cd_nums:
        s3.append(["", "해당 없음 — 워크숍 아이디어가 모두 직무 후보/과업 요소로 정제됨", "", "", ""])
    for col, w in zip("ABCDE", [5, 28, 11, 36, 34]): s3.column_dimensions[col].width = w
    for row in s3.iter_rows(min_row=2):
        for c in row: c.alignment = WRAP; c.border = BD
        row[0].alignment = CTR; row[2].alignment = CTR
        if row[2].value: row[2].fill = fill("FFF2CC" if row[2].value == "개선 제안" else "FCE4D6")
    s3.freeze_panes = "A2"; set_print(s3)

    sel_rows = [(c["code"], c["vt"], c["name"], c["mark"] or "", c.get("basis", "")) for c in candidates]
    build_selection_form(wb, company, sel_rows)

    xlsx = os.path.join(outdir, f"{company}_적합직무_후보군_도출결과_{today}.xlsx")
    wb.save(xlsx)
    md = os.path.join(outdir, f"{company}_적합직무_후보군_도출보고서_{today}.md")
    write_report(md, company, items, order, candidates, grade_map, cd_detail, dist, vt_counts, ws_info, today)
    return xlsx, md, dist, vt_counts, len(candidates), len(cd_nums)


def write_report(md, company, items, order, candidates, grade_map, cd_detail, dist, vt_counts, ws_info, today):
    rec_o = [c for c in candidates if c["mark"] == "◎"]
    rec_c = [c for c in candidates if c["mark"] == "○"]
    by_vt = {k: [c for c in candidates if c["vt"] == k] for k in "①②③④⑤"}
    L = []
    L.append(f"# {company} 적합 직무 후보군 도출 결과\n")
    L.append("> M2. 적합 직무 발굴 / S4. 직무 후보군 도출")
    L.append(f"> 작성일: {today[:4]}-{today[4:6]}-{today[6:]} · 분류 체계: 장애인 적합 직무 5대 가치 유형")
    L.append(f"> (Excel 결과본 `{company}_적합직무_후보군_도출결과_{today}.xlsx`와 동기화)\n")
    L.append("---\n")
    L.append("## 1. 개요\n")
    L.append(f"직무발굴 워크숍(M2/S3)에서 도출된 **아이디어 {len(order)}건**을, 장애인 적합 직무 **5대 가치 유형** "
             f"분류 체계와 직무화 적합성 판정을 거쳐 **직무 후보 {len(candidates)}개**로 체계화하였다. "
             "본 후보군은 기업의 후보군 선정을 거쳐 다음 단계 직무 선정 평가(S5)로 연계된다.\n")
    L.append(f"- {ws_info}")
    L.append(f"- 입력 아이디어: 총 {len(order)}건")
    L.append(f"- 도출 직무 후보: **{len(candidates)}개**")
    L.append(f"- 컨설턴트 추천: ◎ 적극 추천 {len(rec_o)}개 · ○ 추천 {len(rec_c)}개 = {len(rec_o)+len(rec_c)}개")
    L.append(f"- 제외·보류: {dist['C']+dist['D']}건 (개선 제안 {dist['C']} + 보류 {dist['D']})\n")
    L.append("> 직무개발 프레임워크: 직무발굴 W/S(S3) → **직무 후보군 도출(S4·현 단계)** → 기업 후보군 선정 "
             "→ 직무 선정 평가(S5) → 최종 적합 직무 선정(S6).\n")
    L.append("---\n")
    L.append("## 2. 도출 프로세스 요약\n")
    L.append("①정제 → ②직무화 적합성 판정 → ③5대 가치 유형 분류 → ④유사 아이디어 통합 → ⑤직무 정의 → ⑥컨설턴트 추천·기업 선정 연계.\n")
    L.append(f"**직무화 적합성 판정 결과 (전수 {len(order)}건)**\n")
    L.append("| 판정 구분 | 의미 | 건수 | 처리 |")
    L.append("|------|------|------|------|")
    L.append(f"| 직무 후보 | 지속·반복 과업으로 독립 직무 성립 | {dist['A']} | 후보군 편입 |")
    L.append(f"| 과업 요소 | 단독은 약하나 통합 시 직무화 가능 | {dist['B']} | 유사 직무로 통합 후 편입 |")
    L.append(f"| 개선 제안 | 직무가 아닌 제도·근무환경 개선 요구 | {dist['C']} | 제외 → 개선 제안 보존 |")
    L.append(f"| 보류 | 권한 필요·사업 모델 등 부적합 | {dist['D']} | 보류 (사유 명시) |\n")
    L.append(f"직무 후보·과업 요소 {dist['A']+dist['B']}건을 유사 업무 속성으로 통합하여 **{len(candidates)}개 직무 후보**로 정제하였다.\n")
    L.append("**5대 가치 유형별 후보 수**: " + " · ".join(f"{k} {vt_counts[k]}개" for k in "①②③④⑤") + "\n")
    L.append("---\n")
    L.append("## 3. 5대 가치 유형별 후보군\n")
    for k in "①②③④⑤":
        L.append(f"### {VTNAME[k]}\n")
        if not by_vt[k]:
            L.append("(해당 없음)\n"); continue
        for c in by_vt[k]:
            mk = f" **{c['mark']}**" if c["mark"] else ""
            L.append(f"**[{c['code']}] {c['name']}**{mk}\n")
            L.append(f"- 정의: {c['defn']}")
            L.append(f"- 통합 아이디어(원 연번): {', '.join(str(x) for x in c['nums'])}")
            L.append(f"- 직무 특성·장애 적합성: {c['fit']}")
            if c.get("flag"): L.append(f"- 검토 필요: {c['flag']}")
            L.append("")
    L.append("---\n")
    L.append("## 4. 컨설턴트 추천 후보 (5대 추천 기준 근거)\n")
    L.append("> 추천 기준: ①선도기업 특성 부합 ②AI·디지털 활용 ③ESG(친환경·안전) ④활용·확산성 ⑤기존 직무와 차별\n")
    L.append("**◎ 적극 추천**\n")
    L.append("| 직무코드 | 직무명(안) | 유형 | 추천 근거 |")
    L.append("|---|---|---|---|")
    for c in rec_o: L.append(f"| {c['code']} | {c['name']} | {c['vt']} | {c.get('basis','')} |")
    L.append("\n**○ 추천**\n")
    L.append("| 직무코드 | 직무명(안) | 유형 | 추천 근거 |")
    L.append("|---|---|---|---|")
    for c in rec_c: L.append(f"| {c['code']} | {c['name']} | {c['vt']} | {c.get('basis','')} |")
    L.append("\n---\n")
    L.append("## 5. 제외·보류 항목 및 활용방안\n")
    cd_nums = [num for num in order if grade_map[num] in ("C", "D")]
    if cd_nums:
        L.append("| 연번 | 아이디어 | 구분 | 제외 사유 | 활용 방안 |")
        L.append("|---|---|---|---|---|")
        for num in cd_nums:
            g = grade_map[num]; _, reason, plan = cd_detail[num]
            L.append(f"| {num} | {items[num]['idea']} | {GR_KO[g]} | {reason} | {plan} |")
    else:
        L.append("워크숍 아이디어가 모두 직무 후보·과업 요소로 정제되어, 별도 제외·보류 항목은 없다.")
    L.append("\n> 개선 제안 항목은 M5(활용·확산) 및 HR 제도 개선 과제로 보존·활용한다.\n")
    L.append("---\n")
    L.append("## 6. 다음 단계\n")
    L.append("1. **기업 후보군 선정**: [6. 기업 후보군 선정서]에서 컨설턴트 추천을 참고하여 평가 대상 후보군을 선정(◎ 우선 선정 / ○ 선정).")
    L.append("2. **S5 적합 직무 선정 평가**: 선정 후보군을 4대 평가 영역(수행 부담도·구조화/안전성·실행 가능성·활용/확산성)으로 평가.")
    L.append("3. **S6 최종 적합 직무 선정**: 실행가능형 1개 이상 + 전략확산형 1개 이상 확정 → M4 직무분석으로 연계.\n")
    with open(md, "w", encoding="utf-8") as f:
        f.write("\n".join(L))


# ============================ 해양환경공단 ============================
KOEM_CANDS = [
 dict(code="A-1", vt="①", name="해양시료 분류·분석실(Lab) 자료 관리 지원", nums=[17,18,19],
   defn="해양조사 시료에 QR·RFID 태그를 부착·판독하여 채취·보관·분석·결과까지의 이력을 관리하고, 분석실(Lab) 시료 분류·데이터를 정리하여 연구·분석 업무의 신뢰성과 효율을 높인다.",
   fit="태그 부착·판독·분류가 규칙적·반복적이고 체크리스트로 구조화할 수 있어, 정확성·일관성이 강점이 되는 직무로 설계 가능.",
   mark="◎", basis="선도기업 특성(해양시료·분석) · 활용·확산성 · 기존직무 차별 · AI·디지털"),
 dict(code="A-2", vt="①", name="자산·장비 QR/RFID 관리 및 점검 지원", nums=[11,16],
   defn="장비·자재에 QR을 부착해 위치·점검주기·자산현황을 관리하고, NFC 태그 등으로 점검 진위를 확인하여 자산 분실을 예방하고 점검 시간을 단축한다.",
   fit="정해진 절차의 부착·확인·기록 과업으로 디지털 도구 활용이 가능하고 꼼꼼함이 가치가 되는 직무.",
   mark="○", basis="AI·디지털 활용 · 활용·확산성"),
 dict(code="A-3", vt="①", name="데이터·통계 정리 및 시각화 지원", nums=[15,20,24,26,31],
   defn="해경 통계·인사정보·이해관계자 정보 등 각종 데이터를 표준화·통합 관리하고, 엑셀화·시각화·기초 통계분석을 수행하여 정책·사업 활용 기반을 마련한다.",
   fit="반복적 입력·정리·분석 과업으로 디지털 도구 활용이 용이하고 정확성이 핵심 가치가 되는 직무.",
   mark="◎", basis="AI·디지털 활용 · 활용·확산성 · 기존직무 차별"),
 dict(code="A-4", vt="①", name="정보 조사·자료 수집 및 동향·언론 모니터링 지원", nums=[4,29,36,40],
   defn="업무·정책에 필요한 자료와 타기관 레퍼런스·정부 보도자료를 수집·정리하고, 언론·정책 동향을 정기 모니터링하여 담당자가 후속 분석·의사결정에 집중하도록 지원한다.",
   fit="수집·정리·요약이 반복적이고 정해진 양식으로 수행 가능, 디지털 검색·정리 도구 활용 가능.",
   mark="○", basis="활용·확산성 · AI·디지털 활용"),
 dict(code="A-5", vt="①", name="문서·기록물 전산화 및 원본 관리 지원", nums=[6,10],
   defn="비전자 원본문서와 현장 모니터링 기록을 전산화·체계화하고 중요 원본을 안전하게 보관·관리하여 문서 검색성과 보존성을 높인다.",
   fit="스캔·입력·분류가 구조화된 반복 과업으로 정확성·집중이 강점이 되는 직무.",
   mark="", basis="AI·디지털 활용 · 기존직무 차별"),
 dict(code="A-6", vt="①", name="회의·간담회 운영 지원", nums=[9,13,14,50],
   defn="회의·간담회 준비, 참석자 명부·화상회의 세팅, 자료 사전배포, 회의록 작성, 현장 의견 정리 등 회의 운영을 지원한다.",
   fit="정해진 절차·양식으로 수행 가능하고 꼼꼼함이 가치가 되는 직무.",
   mark="", basis="AI·디지털 활용"),
 dict(code="A-7", vt="①", name="구매·계약·행정 사무 지원", nums=[12,45,46,47,49],
   defn="위탁기관 계약서류 준비, 지출결의·우편물 발송·출장 예매·자료 취합 등 부서 운영에 필요한 행정 사무를 지원한다.",
   fit="정형화된 행정 절차로 매뉴얼화가 용이한 직무.",
   mark="", basis="활용·확산성"),
 dict(code="A-8", vt="①", name="정보시스템 접속·계정 보안 점검 지원", nums=[34,35],
   defn="정보시스템 접속기록을 점검하고 계정 부여의 적절성을 확인하여 정보보안을 강화하고 이상 접속을 사전에 탐지한다.",
   fit="규칙 기반 점검·기록 과업으로 정확성이 핵심 가치가 되는 직무.",
   mark="○", basis="기존직무 차별 · AI·디지털 활용"),
 dict(code="B-1", vt="②", name="민원·고객 응대 및 분석 지원", nums=[1,2,3],
   defn="외부 민원·방제분담금 관련 전화를 응대하고, 민원 유형을 자동 분류·분석하며 선박 관련 정보를 사전 조사해 고객 응대 시간을 단축한다.",
   fit="정형 응대 절차와 AI 분류 도구를 활용할 수 있는 직무로, 성실성·정확성이 강점이 됨.",
   mark="○", basis="AI·디지털 활용 · 고객 서비스"),
 dict(code="B-2", vt="②", name="홍보 콘텐츠·간행물 제작 지원(AI 활용)", nums=[22,23,28,30,33,38],
   defn="해양보호 인식증진 콘텐츠·숏폼·간행물·포스터를 AI 도구로 제작하고 홍보물을 분류·관리하여 대국민 홍보와 정책 확산을 지원한다.",
   fit="표준 프롬프트·템플릿으로 제작 가능하고 창의성·꼼꼼함이 가치가 되는 디지털 친화 직무.",
   mark="◎", basis="AI·디지털 활용 · 활용·확산성 · 기존직무 차별"),
 dict(code="C-1", vt="③", name="해양폐기물 수거·자원순환 데이터 취합·공개 지원", nums=[8,67],
   defn="AI를 활용해 해양쓰레기를 분류하고, 선박 폐유 등 수거 자료를 매월 취합·시각화하여 홈페이지에 공개함으로써 ESG(환경) 공공데이터 가치를 창출한다.",
   fit="분류·집계·게시가 반복적이고 구조화 가능하며, 환경 가치와 직접 연결되는 직무.",
   mark="◎", basis="선도기업 특성(해양환경·ESG) · ESG(친환경) · AI·디지털 활용 · 활용·확산성"),
 dict(code="C-2", vt="③", name="사회공헌·환경봉사 활동 운영 지원", nums=[60,61,62],
   defn="임직원 참여 환경봉사·사회공헌 활동과 해양폐기물 재활용 연계 나눔활동을 기획·운영 지원하여 ESG·지역상생 가치를 실현한다.",
   fit="활동 준비·기록·운영 보조가 반복적이고 명확한 절차로 수행 가능한 직무.",
   mark="○", basis="ESG(사회적 가치) · 활용·확산성"),
 dict(code="C-3", vt="③", name="안전·환경 점검 기록 관리 지원", nums=[63],
   defn="전 지사 안전점검 결과를 기록·정리하고 점검 이력을 관리하여 안전사고를 예방한다.",
   fit="체크리스트 기반 점검·기록 과업으로 정확성이 강점이 되는 직무.",
   mark="", basis="ESG(안전)", flag="현장 이동 범위·안전 확인 필요"),
 dict(code="D-1", vt="④", name="사내 그린환경(식물·공간 정돈) 관리", nums=[42,44,56,57],
   defn="사무실 생화·화분 등 식물과 공용 공간의 정돈·청결 상태를 관리하여 쾌적한 근무환경을 유지한다.",
   fit="일정 주기의 반복 관리 과업으로 절차화가 용이한 직무.",
   mark="", basis=""),
 dict(code="D-2", vt="④", name="임직원 복지·기념일 운영 지원", nums=[48,52,58],
   defn="입사 N년차 축하 우편·복리후생(피복 등)·간식 제공 등 임직원 대상 복지 서비스를 운영 지원한다.",
   fit="정해진 일정·목록으로 수행 가능하고 세심함이 가치가 되는 직무.",
   mark="", basis=""),
 dict(code="E-1", vt="⑤", name="해양환경 모니터링 데이터 분석 지원(드론·위성영상 AI)", nums=[25,66],
   defn="드론·위성영상 등 광역 모니터링 데이터를 AI로 분석해 오염지역을 조기 발견하고, 해양환경 현장 탐색·조사 효율을 높이는 신규 응용 분야를 지원한다.",
   fit="정형화된 영상 판독·라벨링·기록 과업으로 설계 가능하며 디지털 친화적 직무.",
   mark="◎", basis="선도기업 특성(해양 모니터링) · AI·디지털 활용 · 활용·확산성", flag="영상분석 도구 교육·고난도 검토"),
]
KOEM_GRADE = {17:"B",18:"A",19:"A",11:"B",16:"A",15:"A",20:"B",24:"B",26:"B",31:"A",
 4:"B",29:"B",36:"B",40:"A",6:"A",10:"A",9:"B",13:"A",14:"B",50:"B",12:"A",45:"B",46:"B",47:"B",49:"B",
 34:"A",35:"B",1:"A",2:"B",3:"B",22:"B",23:"B",28:"A",30:"A",33:"A",38:"B",8:"A",67:"A",60:"A",61:"B",62:"A",
 63:"A",42:"B",44:"B",56:"B",57:"B",48:"B",52:"B",58:"B",25:"A",66:"B",
 5:"C",7:"C",21:"C",32:"C",41:"C",43:"C",51:"C",53:"C",54:"C",55:"C",59:"C",
 27:"D",37:"D",39:"D",64:"D",65:"D",68:"D"}
KOEM_CD = {
 5:("개선 제안","직무가 아닌 복지·조직문화 행사","임직원 복지 제도 검토(M5/HR)"),
 7:("개선 제안","근태·건강관리 시스템 도입 제안(직무 아님)","HR·정보화 제도 검토"),
 21:("개선 제안","관리자 역할의 조직문화 활동(직무화 곤란)","조직문화 개선 과제"),
 32:("개선 제안","안전 인프라(지능형 CCTV) 도입 제안","안전·시설 인프라 검토"),
 41:("개선 제안","감성 근무환경 요소(직무 아님)","사무환경 개선(M5 연계)"),
 43:("개선 제안","감성 근무환경 요소(퇴근송·디퓨저)","사무환경 개선(M5 연계)"),
 51:("개선 제안","감성 근무환경 요소(향)","사무환경 개선(M5 연계)"),
 53:("개선 제안","부서 간 네트워킹·조직문화 활동(직무화 곤란)","조직문화 개선 과제"),
 54:("개선 제안","감성 근무환경 요소(음악·식물·반려)","사무환경·복지 검토"),
 55:("개선 제안","출장·지역상생 제도 제안(직무 아님)","출장 제도·지역상생 사업 검토"),
 59:("개선 제안","보상·인사 제도 요구(표창·성과금)","보상·인정 제도 검토"),
 27:("보류","외부 인력 기반 지역사업 모델로 장애인 직무화 곤란","지역상생 사업 별도 검토"),
 37:("보류","대국민 참여 플랫폼 구축이 선행되어야 함","신규 플랫폼 사업 검토"),
 39:("보류","연구기관 설립 등 대규모 신사업으로 직무 범위 초과","중장기 ESG 신사업 검토"),
 64:("보류","내용 불명확·사업 성격(직무화 곤란)","구체화 후 재검토"),
 65:("보류","주관·심사 권한이 필요하여 신규 직무 부적합","행정 보조 범위로 재검토"),
 68:("보류","외부 인력운영(아웃소싱) 모델로 장애인 직무 아님","인력운영 정책 검토"),
}

# ============================ KF ============================
KF_CANDS = [
 dict(code="A-1", vt="①", name="자료 조사·큐레이션 및 뉴스·동향 정리 지원", nums=[28,29,30,31],
   defn="부서 요청 기초자료·국내외 동향을 조사·요약하고, 분야별 자료를 큐레이션·분류·PDF화하며, 주요 뉴스를 매일 스크랩해 직원에게 공유하여 정보 탐색 시간을 줄인다.",
   fit="수집·분류·요약이 반복적이고 정해진 양식으로 수행 가능, AI·디지털 도구 활용이 용이한 직무. 일부 재택 가능.",
   mark="◎", basis="AI·디지털 활용 · 활용·확산성 · 기존직무 차별"),
 dict(code="A-2", vt="①", name="수혜자·동문·고객 정보 DB 관리 지원", nums=[15,21,27],
   defn="장학생·펠로 등 사업 수혜자, KF 동문, 협력기관·초청 인사 정보를 DB로 구축·관리하고 정기 소통을 지원한다.",
   fit="입력·갱신·분류가 반복적이고 정확성이 핵심 가치가 되는 직무.",
   mark="○", basis="활용·확산성 · 기존직무 차별"),
 dict(code="A-3", vt="①", name="회의록·기록물·데이터 관리 지원", nums=[18,38,39,40,41,44],
   defn="회의·간담회 회의록을 작성·정리하고, 사업 실적·결산자료를 전자화하며, 기록물·자산을 체계적으로 관리하고 차량 2부제 운행을 점검·기록한다.",
   fit="기록·입력·점검이 구조화된 반복 과업으로 정확성·집중이 강점이 되는 직무.",
   mark="○", basis="기존직무 차별 · AI·디지털 활용"),
 dict(code="A-4", vt="①", name="물품 발송·재고 관리 지원", nums=[23,34,35,37],
   defn="기념품·도서·발간물의 포장·발송을 지원하고, 서고 발간물과 책자 재고·사무용품을 정리·관리·주문한다.",
   fit="포장·정리·재고 점검이 반복적이고 절차화가 용이한 직무.",
   mark="", basis="활용·확산성"),
 dict(code="A-5", vt="①", name="우편·구독·부서 메일함 관리 지원", nums=[24,36,42],
   defn="외부 우편물을 수령·분류·배분하고, 발간물·이메일 구독 리스트와 부서 공용 메일함을 관리하여 중요 정보 누락을 방지한다.",
   fit="분류·전달·갱신이 반복적이고 꼼꼼함이 가치가 되는 직무.",
   mark="", basis=""),
 dict(code="A-6", vt="①", name="사진·시각자료 정리·아카이빙 지원", nums=[1],
   defn="이사장 관련 사진·자료 등 재단 시각자료를 정리·분류·업로드하고 아카이빙하여 활용도를 높인다. (AI 활용 가능)",
   fit="분류·태깅·업로드가 규칙적인 디지털 과업으로 꼼꼼함이 가치가 되는 직무.",
   mark="", basis="AI·디지털 활용"),
 dict(code="B-1", vt="②", name="홍보 콘텐츠 제작 지원(시각물·SNS·보도자료, AI 활용)", nums=[2,3,4],
   defn="행사·홍보용 포스터·X배너, 인스타툰·카드뉴스 등 SNS 콘텐츠와 보도자료 초안을 AI 도구로 기획·제작하여 재단 성과를 효과적으로 알린다.",
   fit="표준 템플릿·프롬프트로 제작 가능하고 창의성·꼼꼼함이 가치가 되는 디지털 친화 직무. 일부 재택 가능.",
   mark="◎", basis="AI·디지털 활용 · 활용·확산성 · 기존직무 차별"),
 dict(code="B-2", vt="②", name="번역 검토·웹진/홈페이지 검수 지원", nums=[5,6],
   defn="영↔한 번역 결과물의 일관성·오류를 검수하고, Koreana 웹진·홈페이지의 오류와 민감 표현을 체크리스트로 점검하여 콘텐츠 품질을 높인다.",
   fit="체크리스트 기반 검수가 반복적이고 정확성·집중이 핵심 가치가 되는 직무.",
   mark="◎", basis="선도기업 특성(국제교류·번역) · 기존직무 차별 · AI·디지털 활용"),
 dict(code="B-3", vt="②", name="대외 관계 메시지 작성 지원", nums=[12,25,26],
   defn="행사 환영 인사, 협력기관 기념일·개소일 축하, 맞춤형 손편지·자필편지 등 상황별 관계외교 메시지를 작성·발송 지원한다.",
   fit="정형 템플릿을 바탕으로 정성과 세심함이 가치가 되는 직무.",
   mark="○", basis="선도기업 특성(관계외교) · 기존직무 차별"),
 dict(code="B-4", vt="②", name="행사·이벤트 운영 및 의전 보조", nums=[11,14],
   defn="공연·국제포럼·문화행사 등 이벤트 기획·운영을 보조하고, 방한 외교관·초청 인사의 일정 인솔과 의전을 지원한다.",
   fit="정해진 진행 절차에 따른 준비·안내 과업으로 성실성이 강점이 되는 직무.",
   mark="", basis="고객 서비스", flag="인솔·의전의 이동·대면 요건 확인"),
 dict(code="B-5", vt="②", name="교육 프로그램 운영·수료 행정 지원", nums=[13,16,43],
   defn="수료증 제작·발급번호 관리·해외 발송, 외국인 전화 한국어 교육 운영 지원, 외부기관 교육 협조문 접수·안내 등 교육 프로그램 운영·행정을 지원한다.",
   fit="발급·연락·안내가 정형화된 절차로 수행 가능한 직무.",
   mark="○", basis="선도기업 특성(국제교류 교육) · 활용·확산성"),
 dict(code="B-6", vt="②", name="재단 소식·구독자 소통 지원", nums=[32,33],
   defn="재단 사업 리스트를 월간 이메일로 정리·발송하고, Koreana 독자 피드백을 수집·분류·정리하며 정기 구독을 관리한다.",
   fit="정리·발송·분류가 반복적이고 꼼꼼함이 가치가 되는 직무.",
   mark="", basis="고객 서비스"),
 dict(code="B-7", vt="②", name="기념품·제작물 기획·관리 지원", nums=[7,8,22],
   defn="한국을 알리는 K-기념품과 부서 명패, 사내 메신저 이모티콘 등 재단 제작물을 디자인·제작하고 재고를 관리한다.",
   fit="디자인·제작·재고 관리가 절차화 가능하고 창의성·꼼꼼함이 가치가 되는 직무.",
   mark="○", basis="선도기업 특성(한국 알리기) · 활용·확산성"),
 dict(code="E-1", vt="⑤", name="디지털 전시·키오스크 콘텐츠 관리", nums=[9,10],
   defn="KF글로벌센터 메타버스·VR 전시 콘텐츠와 로비 키오스크의 뉴스·성과·전시 사진을 업로드·관리하여 디지털 공공외교 콘텐츠를 최신 상태로 유지한다.",
   fit="업로드·갱신·점검이 규칙적인 디지털 과업으로 디지털 친화성이 강점이 되는 직무.",
   mark="◎", basis="선도기업 특성(디지털 공공외교) · 신규 가치 · AI·디지털 활용"),
 dict(code="C-1", vt="③", name="ESG·사회공헌 활동 운영 지원", nums=[17,19,20],
   defn="수어 지원 배리어프리 한국학 e-스쿨 운영, 장애인·이주민 연계 봉사활동, ESG 플로깅·반려해변 등 환경 공공외교 활동을 기획·운영 지원한다.",
   fit="활동 준비·기록·운영 보조가 반복적이고 명확한 절차로 수행 가능한 직무.",
   mark="◎", basis="선도기업 특성(공공외교) · ESG(사회·환경) · 활용·확산성"),
 dict(code="C-2", vt="③", name="청사 에너지 절감 관리 지원", nums=[47],
   defn="청사 에너지 사용 현황을 모니터링하고 절감 활동을 기록·보고하여 ESG(환경) 가치를 실현한다.",
   fit="검침·기록·보고가 정형화된 반복 과업.",
   mark="", basis="ESG(친환경)"),
 dict(code="D-1", vt="④", name="사내 문화·소통 프로그램 운영 지원", nums=[53,54,56],
   defn="출근 시간 사내 라디오와 퇴근 음악방송을 운영하고, 아침 명언·영감 이메일을 발송하여 활기 있는 근무 분위기를 조성한다.",
   fit="정해진 일정의 콘텐츠 선정·송출이 반복적인 운영 과업.",
   mark="", basis=""),
 dict(code="D-2", vt="④", name="사무환경·웰니스 운영 지원", nums=[45,46,48,49,55],
   defn="직원 공간·회의실 전시 그림 큐레이팅과 KF 브랜드 색 세팅, 방향·식물 관리, 기념일 사무실 장식, 스트레칭·웰니스 프로그램을 운영 지원한다.",
   fit="설치·관리·교체가 주기적이고 절차화가 용이한 직무.",
   mark="", basis=""),
 dict(code="D-3", vt="④", name="임직원 복지(간식·온보딩) 운영 지원", nums=[50,52],
   defn="신입직원 온보딩 키트를 구성·전달하고, 직원 간식·식음료 수요를 파악·운영 지원하여 조직 적응과 복지를 돕는다.",
   fit="준비·전달·재고 관리가 정형화된 과업.",
   mark="", basis=""),
]
KF_GRADE = {28:"A",29:"B",30:"A",31:"B",15:"A",21:"B",27:"A",18:"A",38:"A",39:"A",40:"B",41:"B",44:"B",
 23:"A",34:"B",35:"A",37:"B",24:"B",36:"A",42:"B",1:"A",2:"A",3:"A",4:"A",5:"A",6:"A",12:"B",25:"A",26:"B",
 11:"A",14:"B",13:"A",16:"A",43:"B",32:"B",33:"A",7:"B",8:"B",22:"A",9:"A",10:"B",17:"A",19:"A",20:"A",47:"A",
 53:"A",54:"B",56:"B",45:"A",46:"B",48:"B",49:"A",55:"B",50:"A",52:"B",51:"C"}
KF_CD = {51:("개선 제안","직원 간 긍정 문화 조성 활동으로 직무 분량이 약함(조직문화 성격)","사내 소통·문화 프로그램(M5/HR 연계)")}

BASE = "/sessions/sweet-intelligent-newton/mnt/P2026-003_2026년_선도기업_직무디자인_연구용역/outputs/기업별"
UP = "/sessions/sweet-intelligent-newton/mnt/uploads"

if __name__ == "__main__":
    r1 = build("해양환경공단", f"{UP}/[대우인력] KEAD_아이디어Pool_해양환경공단_0610.xlsx",
               KOEM_CANDS, KOEM_GRADE, KOEM_CD,
               f"{BASE}/08_해양환경공단/M2_직무발굴/S4_후보군도출/",
               "W/S 일시·참여자: (기재) — 본 결과는 직무발굴 W/S 아이디어 Pool 기반")
    print("[해양환경공단]", r1[2], "후보", r1[4], "제외보류", r1[5])
    r2 = build("한국국제교류재단", f"{UP}/[대우인력] KEAD_아이디어Pool_한국국제교류재단(KF)_0610.xlsx",
               KF_CANDS, KF_GRADE, KF_CD,
               f"{BASE}/07_한국국제교류재단/M2_직무발굴/S4_후보군도출/",
               "W/S 일시·참여자: (기재) — 본 결과는 직무발굴 W/S 아이디어 Pool 기반")
    print("[KF]", r2[2], "후보", r2[4], "제외보류", r2[5])
