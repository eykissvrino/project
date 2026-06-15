"""v12 진짜 마지막 - 04/05/06 완전 재빌드 + 01 빈 행 정리"""
import os, shutil, subprocess, glob, random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

random.seed(2026)

COLOR = {'header_dark':'1F4E78','header_mid':'2E74B5','header_light':'B4C7E7',
    'input_yellow':'FFF2CC','auto_green':'E2EFDA','info_gray':'F2F2F2',
    's_grade':'C6E0B4','a_grade':'D9E1F2','c_grade':'FFEB9C','d_grade':'FFC7CE',
    'mute':'D9D9D9'}
def fill(c): return PatternFill('solid', fgColor=c)
def font(size=11, bold=False, color='000000'):
    return Font(name='맑은 고딕', size=size, bold=bold, color=color)
def align(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
THIN = Side(style='thin', color='BFBFBF')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def sc(ws, coord, value=None, fill_color=None, bold=False, size=11, color='000000',
       h='left', v='center', wrap=True, border=BORDER_ALL, number_fmt=None):
    c = ws[coord]
    if isinstance(c, MergedCell): return
    if value is not None: c.value = value
    c.font = font(size=size, bold=bold, color=color)
    c.alignment = align(h=h, v=v, wrap=wrap)
    if fill_color: c.fill = fill(fill_color)
    if border: c.border = border
    if number_fmt: c.number_format = number_fmt

def g2s(cr):
    return f'IFERROR(CHOOSE(FIND({cr},"DCBAS"),1,2,3,4,5),0)'
def s2g(cr):
    return f'IF({cr}="","",IF({cr}>=4.5,"S",IF({cr}>=3.5,"A",IF({cr}>=2.5,"B",IF({cr}>=1.5,"C","D")))))'

DATA_START = 6
EMP_MAX = 50
DATA_END = DATA_START + EMP_MAX - 1  # 55
W1 = "'02_의사결정·종합결과'!$B$67"
W2 = "'02_의사결정·종합결과'!$B$68"

SRC = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가/베이시스소프트_HR통합관리시트.xlsx"
SRC_BACKUP = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가/베이시스소프트_HR_임금평가_관리 TOOL/베이시스소프트_HR통합관리시트_최종_v6.xlsx"
TMP = "/tmp/master_v12.xlsx"
# 원본 파일 손상 여부 확인 후 백업으로 폴백
import zipfile
try:
    zipfile.ZipFile(SRC).close()
    shutil.copy(SRC, TMP)
    print(f"원본 사용: {SRC}")
except Exception:
    shutil.copy(SRC_BACKUP, TMP)
    print(f"백업 사용(원본 손상): {SRC_BACKUP}")
wb = load_workbook(TMP)

# ============ 1. 01_인적정보 빈 행 정리 (0 → None) ============
ws = wb['01_인적정보']
# 머지 해제
for mr in list(ws.merged_cells.ranges):
    if mr.min_row >= 40 and mr.max_row <= 60:
        ws.unmerge_cells(str(mr))

for r in range(40, 56):  # 빈 행 영역
    # A: No
    sc(ws, f'A{r}', r - DATA_START + 1, h='center')
    # B~G, I, K, N, O, P: 입력 셀 - 빈 값(None)
    for col in [2, 3, 4, 5, 6, 7, 9, 11, 14, 15, 16]:
        cell = ws.cell(r, col)
        if not isinstance(cell, MergedCell):
            cell.value = None
            cell.fill = fill(COLOR['input_yellow'])
            cell.border = BORDER_ALL
            cell.alignment = align()
    # H: 역할등급
    cell = ws.cell(r, 8)
    if not isinstance(cell, MergedCell):
        cell.value = None
        cell.fill = fill(COLOR['input_yellow'])
        cell.border = BORDER_ALL
        cell.font = font(bold=True)
        cell.alignment = align(h='center')
    # J: 근속연수 (자동)
    cell = ws.cell(r, 10)
    if not isinstance(cell, MergedCell):
        cell.value = f'=IFERROR(IF(I{r}="","",ROUND((TODAY()-DATEVALUE(I{r}))/365,1)),"")'
        cell.fill = fill(COLOR['auto_green'])
        cell.border = BORDER_ALL
        cell.alignment = align(h='center')
    # L: 주요 책무 수 (자동)
    cell = ws.cell(r, 12)
    if not isinstance(cell, MergedCell):
        cell.value = f'=IF(B{r}="","",IFERROR(INDEX(\'15_직원별책무매핑\'!$E:$E,MATCH(B{r},\'15_직원별책무매핑\'!$B:$B,0)),0))'
        cell.fill = fill(COLOR['auto_green'])
        cell.border = BORDER_ALL
        cell.alignment = align(h='center')
    # M: 조회키
    cell = ws.cell(r, 13)
    if not isinstance(cell, MergedCell):
        cell.value = f'=IF(B{r}="","",E{r}&"_"&H{r})'
        cell.fill = fill(COLOR['auto_green'])
        cell.border = BORDER_ALL
        cell.alignment = align()
print("01_인적정보 빈 행 정리 완료 (0 → None)")

# ============ 2. 04/05/06 시트 완전 클리어 + 재빌드 ============
def rebuild_eval_sheet(ws, title, subtitle, src1, src2, has_main=False):
    # 모든 머지 해제
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))
    # 모든 셀 완전 클리어 (행 1~130)
    for r in range(1, 130):
        for c in range(1, 22):
            cell = ws.cell(r, c)
            if not isinstance(cell, MergedCell):
                cell.value = None
                cell.fill = PatternFill()
                cell.font = font()
                cell.alignment = align()
                cell.border = Border()
        # 행 높이 초기화
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None
    # 데이터 유효성 검사 클리어
    ws.data_validations.dataValidation = []
    # 조건부 서식도 클리어
    ws.conditional_formatting._cf_rules.clear()

    # 타이틀
    if has_main:
        end_col_letter = 'J'
    else:
        end_col_letter = 'I'
    ws.merge_cells(f'A1:{end_col_letter}1')
    sc(ws, 'A1', title, fill_color=COLOR['header_dark'],
       color='FFFFFF', bold=True, size=14, h='center')
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f'A2:{end_col_letter}2')
    sc(ws, 'A2', subtitle, fill_color=COLOR['c_grade'], size=10, bold=True, h='center')
    ws.row_dimensions[2].height = 24

    # 컬럼 정의
    if has_main:
        c_grade1, c_grade2, c_score, c_final = 'G', 'H', 'I', 'J'
        headers = ['No.','사번','성명','직무','역할등급','메인 책무 수','1차','2차','가중평균(점수)','확정등급']
    else:
        c_grade1, c_grade2, c_score, c_final = 'F', 'G', 'H', 'I'
        headers = ['No.','사번','성명','직무','역할등급','1차','2차','가중평균(점수)','확정등급']

    # 상반기 / 하반기 두 블록
    for term_idx, (term, sheet07) in enumerate([('상반기', "'07a_상반기입력'"), ('하반기', "'07b_하반기입력'")]):
        block_start = 4 + term_idx * 58  # 상반기 4, 하반기 62

        # 블록 헤더
        ws.merge_cells(f'A{block_start}:{end_col_letter}{block_start}')
        sc(ws, f'A{block_start}', f"■ {term} {title.split('(')[0].strip()} (50명까지)",
           fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=12, h='center')
        ws.row_dimensions[block_start].height = 26

        # 컬럼 헤더
        for i, h in enumerate(headers, 1):
            sc(ws, f'{get_column_letter(i)}{block_start+1}', h,
               fill_color=COLOR['header_dark'], bold=True, color='FFFFFF',
               h='center', size=10)
        ws.row_dimensions[block_start+1].height = 28

        # 데이터 행 50명
        for i in range(EMP_MAX):
            r = block_start + 2 + i
            ref = DATA_START + i

            # 인적정보 자동 참조 (사번 비면 모두 빈 값)
            sc(ws, f'A{r}', f"=IF('01_인적정보'!B{ref}=\"\",\"\",'01_인적정보'!A{ref})",
               h='center', fill_color=COLOR['info_gray'])
            sc(ws, f'B{r}', f"=IF('01_인적정보'!B{ref}=\"\",\"\",'01_인적정보'!B{ref})",
               h='center', fill_color=COLOR['info_gray'])
            sc(ws, f'C{r}', f"=IF('01_인적정보'!B{ref}=\"\",\"\",'01_인적정보'!C{ref})",
               fill_color=COLOR['info_gray'])
            sc(ws, f'D{r}', f"=IF('01_인적정보'!B{ref}=\"\",\"\",'01_인적정보'!E{ref})",
               fill_color=COLOR['info_gray'])
            sc(ws, f'E{r}', f"=IF('01_인적정보'!B{ref}=\"\",\"\",'01_인적정보'!H{ref})",
               h='center', fill_color=COLOR['info_gray'], bold=True)

            if has_main:
                sc(ws, f'F{r}', f"=IF('01_인적정보'!B{ref}=\"\",\"\",'01_인적정보'!L{ref})",
                   h='center', fill_color=COLOR['auto_green'], bold=True)

            # 1차/2차 등급
            sc(ws, f'{c_grade1}{r}',
               f"=IF('01_인적정보'!B{ref}=\"\",\"\",{sheet07}!{src1}{ref})",
               h='center', fill_color=COLOR['auto_green'])
            sc(ws, f'{c_grade2}{r}',
               f"=IF('01_인적정보'!B{ref}=\"\",\"\",{sheet07}!{src2}{ref})",
               h='center', fill_color=COLOR['auto_green'])

            # 가중평균 점수 (1차/2차 둘 다 있어야 계산)
            sc(ws, f'{c_score}{r}',
               f'=IFERROR(IF(AND({c_grade1}{r}<>"",{c_grade2}{r}<>""),{W1}*{g2s(c_grade1+str(r))}+{W2}*{g2s(c_grade2+str(r))},IF({c_grade1}{r}<>"",{g2s(c_grade1+str(r))},IF({c_grade2}{r}<>"",{g2s(c_grade2+str(r))},""))),"")',
               h='center', fill_color=COLOR['auto_green'], number_fmt='0.00')

            # 확정 등급
            sc(ws, f'{c_final}{r}', f'={s2g(c_score+str(r))}',
               h='center', fill_color=COLOR['auto_green'], bold=True, size=12)
            ws.row_dimensions[r].height = 22

        # 조건부 서식
        last_row = block_start + 1 + EMP_MAX
        for g, c in [('S',COLOR['s_grade']),('A',COLOR['a_grade']),
                     ('C',COLOR['c_grade']),('D',COLOR['d_grade'])]:
            ws.conditional_formatting.add(
                f'{c_grade1}{block_start+2}:{c_final}{last_row}',
                CellIsRule(operator='equal', formula=[f'"{g}"'], fill=fill(c)))

    # 컬럼 너비
    if has_main:
        widths = [5,10,9,14,9,10,11,11,13,11]
    else:
        widths = [5,10,9,14,9,11,11,13,11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'D6'

# 04/05/06 깨끗 재빌드
rebuild_eval_sheet(wb['04_역할평가'],
    "역할평가 (자동 산출)",
    "07a/07b의 1·2차 평가자 종합 등급 → 02_의사결정의 평가자 가중치 자동 적용 → 확정 등급",
    'G', 'H', has_main=False)
print("04_역할평가 완전 재빌드")

rebuild_eval_sheet(wb['05_공통역량평가'],
    "공통역량평가 (자동 산출)",
    "07a/07b의 1·2차 공통역량 종합 등급 → 평가자 가중치 적용 → 확정 등급",
    'I', 'J', has_main=False)
print("05_공통역량평가 완전 재빌드")

rebuild_eval_sheet(wb['06_직무역량평가'],
    "직무역량평가 (자동 산출)",
    "07a/07b의 1·2차 직무역량 종합 등급 → 평가자 가중치 적용 → 확정 등급",
    'K', 'L', has_main=True)
print("06_직무역량평가 완전 재빌드")

# ============ 3. 08_종합평가 수식 정리 (CHOOSE/FIND) ============
ws = wb['08_종합평가']
W1H = "'02_의사결정·종합결과'!$B$7"
W2H = "'02_의사결정·종합결과'!$B$8"
for i in range(EMP_MAX):
    r = DATA_START + i
    role_h = r + 58
    sc(ws, f'G{r}', f"=IF('01_인적정보'!B{r}=\"\",\"\",'04_역할평가'!I{r})", h='center', fill_color=COLOR['auto_green'])
    sc(ws, f'H{r}', f"=IF('01_인적정보'!B{r}=\"\",\"\",'04_역할평가'!I{role_h})", h='center', fill_color=COLOR['auto_green'])
    sc(ws, f'I{r}',
       f'=IFERROR(IF(AND(G{r}<>"",H{r}<>""),{g2s(f"G{r}")}*{W1H}+{g2s(f"H{r}")}*{W2H},""),"")',
       h='center', fill_color=COLOR['auto_green'], number_fmt='0.00')
    sc(ws, f'J{r}', f"=IF('01_인적정보'!B{r}=\"\",\"\",'05_공통역량평가'!I{r})", h='center', fill_color=COLOR['auto_green'])
    sc(ws, f'K{r}', f"=IF('01_인적정보'!B{r}=\"\",\"\",'05_공통역량평가'!I{role_h})", h='center', fill_color=COLOR['auto_green'])
    sc(ws, f'L{r}',
       f'=IFERROR(IF(AND(J{r}<>"",K{r}<>""),{g2s(f"J{r}")}*{W1H}+{g2s(f"K{r}")}*{W2H},""),"")',
       h='center', fill_color=COLOR['auto_green'], number_fmt='0.00')
    sc(ws, f'M{r}', f"=IF('01_인적정보'!B{r}=\"\",\"\",'06_직무역량평가'!J{r})", h='center', fill_color=COLOR['auto_green'])
    sc(ws, f'N{r}', f"=IF('01_인적정보'!B{r}=\"\",\"\",'06_직무역량평가'!J{role_h})", h='center', fill_color=COLOR['auto_green'])
    sc(ws, f'O{r}',
       f'=IFERROR(IF(AND(M{r}<>"",N{r}<>""),{g2s(f"M{r}")}*{W1H}+{g2s(f"N{r}")}*{W2H},""),"")',
       h='center', fill_color=COLOR['auto_green'], number_fmt='0.00')
    sc(ws, f'S{r}',
       f'=IF(OR(I{r}="",L{r}="",O{r}=""),"",I{r}*P{r}+L{r}*Q{r}+O{r}*R{r})',
       h='center', fill_color=COLOR['auto_green'], number_fmt='0.00', bold=True)
    sc(ws, f'T{r}', f'={s2g(f"S{r}")}', h='center', fill_color=COLOR['auto_green'], bold=True, size=12)
    sr = f"$S${DATA_START}:$S${DATA_END}"
    pct = f"IFERROR(RANK(S{r},{sr},0)/COUNT({sr}),0)"
    rel = (f'IF(S{r}="","",IF({pct}<=\'02_의사결정·종합결과\'!$B$11,"S",'
        f'IF({pct}<=\'02_의사결정·종합결과\'!$B$11+\'02_의사결정·종합결과\'!$C$11,"A",'
        f'IF({pct}<=\'02_의사결정·종합결과\'!$B$11+\'02_의사결정·종합결과\'!$C$11+\'02_의사결정·종합결과\'!$D$11,"B",'
        f'IF({pct}<=\'02_의사결정·종합결과\'!$B$11+\'02_의사결정·종합결과\'!$C$11+\'02_의사결정·종합결과\'!$D$11+\'02_의사결정·종합결과\'!$E$11,"C","D")))))')
    sc(ws, f'U{r}',
       f'=IF(S{r}="","",IF(\'02_의사결정·종합결과\'!$B$6="절대평가",T{r},{rel}))',
       h='center', fill_color=COLOR['auto_green'], bold=True, size=14)
print("08_종합평가 수식 정리")

wb.save(TMP)
print(f"\n저장: {TMP}")

# 통합시트 덮어쓰기
try:
    shutil.copy(TMP, SRC)
    print(f"통합시트 저장: {SRC}")
except PermissionError:
    SRC = SRC.replace('.xlsx','_v12.xlsx')
    shutil.copy(TMP, SRC)
    print(f"통합시트 저장 (잠금): {SRC}")

# ============ 4. 시뮬레이션 결과 재생성 ============
SIM = "/tmp/sim_v12.xlsx"
shutil.copy(TMP, SIM)
wb_s = load_workbook(SIM)

# 01 성명 마스킹
ws = wb_s['01_인적정보']
for r in range(DATA_START, DATA_END+1):
    bss = ws.cell(r, 2).value
    if bss and str(bss).startswith('BSS'):
        try: ws.cell(r, 3).value = "OOO"
        except: pass

# 15 매핑 마스킹
ws = wb_s['15_직원별책무매핑']
for r in range(4, 80):
    cell = ws.cell(r, 3)
    if isinstance(cell, MergedCell): continue
    v = cell.value
    if v and isinstance(v, str) and 'BSS' not in v:
        if '(' in v:
            grade_part = v[v.index('('):]
            cell.value = f"OOO {grade_part}"
        elif v not in ['No.','사번','성명','부서','메인 책무 수'] and len(v) <= 5:
            if any('가' <= ch <= '힣' for ch in v):
                cell.value = "OOO"

# 15 매핑 O 표시
for r in range(4, 80):
    bss = ws.cell(r, 2).value
    if bss and str(bss).startswith('BSS'):
        slots = [c for c in range(6, 30) if ws.cell(r, c).value == 'X']
        if slots:
            n_main = max(int(len(slots)*0.6), 3)
            for col in random.sample(slots, n_main):
                ws.cell(r, col).value = 'O'

# 07a/07b 가상 등급
def gen():
    r = random.random()
    return 'S' if r<0.10 else 'A' if r<0.30 else 'B' if r<0.80 else 'C' if r<0.95 else 'D'

for sn in ['07a_상반기입력','07b_하반기입력']:
    ws = wb_s[sn]
    for r in range(DATA_START, DATA_START + 34):
        for c in range(7, 13):
            ws.cell(r, c).value = gen()

ws = wb_s['03_전사업적평가']
ws['K6'].value = 0.12; ws['K7'].value = 0.07; ws['K8'].value = 0.92
ws = wb_s['02_의사결정·종합결과']
ws['B67'].value = 0.3
ws['B6'].value = '상대평가'
wb_s.save(SIM)

# LO 변환
out_dir = "/tmp/sim_v12_check"
shutil.rmtree(out_dir, ignore_errors=True)
os.makedirs(out_dir)
subprocess.run(['libreoffice','--headless','--calc','--convert-to','xlsx',
                '--outdir',out_dir,SIM], capture_output=True, timeout=120)
conv = glob.glob(out_dir + "/*.xlsx")

OUT_SIM = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가/베이시스소프트_HR_시뮬레이션결과.xlsx"
try:
    shutil.copy(conv[0], OUT_SIM)
    print(f"시뮬레이션 저장: {OUT_SIM}")
except PermissionError:
    OUT_SIM = OUT_SIM.replace('.xlsx','_v12.xlsx')
    shutil.copy(conv[0], OUT_SIM)
    print(f"시뮬레이션 저장 (잠금): {OUT_SIM}")

# 검증
wb_c = load_workbook(conv[0], data_only=True)
print("\n=== 04 검증 (빈 행 처리) ===")
ws = wb_c['04_역할평가']
for r in [6, 35, 40, 55]:
    print(f"  행{r}: 사번={ws.cell(r,2).value} 직무={ws.cell(r,4).value} 등급={ws.cell(r,5).value} 1차={ws.cell(r,6).value} 2차={ws.cell(r,7).value} 점수={ws.cell(r,8).value} 확정={ws.cell(r,9).value}")

print("\nv12 완료")
