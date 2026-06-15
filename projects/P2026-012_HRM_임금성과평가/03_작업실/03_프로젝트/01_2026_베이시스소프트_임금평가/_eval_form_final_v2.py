"""평가지 양식 7종 완전 재정비 (final v2)
1. 백업(./평가지_양식/)에서 ./평가지_양식_최종/으로 복구
2. 00_작성안내 시트 완전 재빌드 (STEP 1~7 형식, 문의 섹션 없음, STEP 7 평가자 입장)
3. 모든 시트 열 너비 통일
4. 직무역량 섹션 헤더 행을 두 줄로 만들어 BARS 등급 기준 추가
"""
import os, glob, shutil
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가"
BACKUP_DIR = os.path.join(BASE, "평가지_양식")
FINAL_DIR = os.path.join(BASE, "평가지_양식_최종")

THIN = Side(style='thin', color='BFBFBF')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLOR = {'header_dark':'1F4E78','header_mid':'2E74B5','header_light':'B4C7E7',
    'input_yellow':'FFF2CC','auto_green':'E2EFDA','info_gray':'F2F2F2',
    'callout':'FFEB9C', 'mark':'C00000'}

STD_WIDTH = {'A':7, 'B':14, 'C':24, 'D':38, 'E':38, 'F':8, 'G':8, 'H':11,
             'I':11, 'J':11, 'K':11, 'L':11}

# 직무명 - 파일명에서 추출용 매핑
JOB_NAME = {
    'BIM컨설팅': 'BIM컨설팅',
    'SW판매영업': 'SW판매영업',
    '경영전략기획': '경영전략기획',
    '교육': '교육',
    '기술지원': '기술지원',
    '연구개발RD': '연구개발(R&D)',
    '프로젝트영업': '프로젝트영업',
}

# 등급 매핑 (시트명 키 → (역할명))
GRADE_INFO = [
    ('G6',   '사업·조직 관리'),
    ('G5-M', '직무 관리(부서장)'),
    ('G5-S', '실무 스페셜리스트'),
    ('G4',   '실무 책임(PM)'),
    ('G3',   '핵심 실무'),
    ('G2',   '실무'),
    ('G1',   '실무 보조'),
]

NEW_JOB_HEADER = (
    "▶ ③ 직무역량평가 - 메인 책무(D열에 O 표시된 책무만) 1·2차 평가자 등급 입력\n"
    "★ S/A/B/C/D 등급 기준 (모든 책무 공통 적용):   "
    "S(탁월) 책무 기대 현저 초과   ┃   "
    "A(우수) 기대 상회   ┃   "
    "B(양호) 기대 부합   ┃   "
    "C(미흡) 일부 미달   ┃   "
    "D(부족) 전반 미흡"
)

def sc(ws, coord, value=None, fill_color=None, bold=False, size=11, color='000000',
       h='left', v='center', wrap=True, border=BORDER_ALL, indent=0):
    c = ws[coord]
    if isinstance(c, MergedCell): return
    if value is not None: c.value = value
    c.font = Font(name='맑은 고딕', size=size, bold=bold, color=color)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)
    if fill_color: c.fill = PatternFill('solid', fgColor=fill_color)
    if border: c.border = border

def clear_sheet(ws, max_row=80, max_col=12):
    """시트 완전 클리어 (머지/스타일 모두)"""
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))
    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell): continue
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font(name='맑은 고딕', size=11)
            cell.alignment = Alignment()
            cell.border = Border()
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None

def build_guide_sheet(ws, job_name, file_job_key):
    """00_작성안내 시트 새로 빌드 (STEP 1~7 + 부가 정보, 문의 섹션 없음)"""
    clear_sheet(ws)

    # 타이틀
    ws.merge_cells('A1:C1')
    sc(ws, 'A1', f'성과평가지 양식 - {job_name}',
       fill_color=COLOR['header_dark'], color='FFFFFF', bold=True, size=14, h='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:C2')
    sc(ws, 'A2', '베이시스소프트 1년 2회(상반기·하반기) 정기 성과평가 표준 양식',
       fill_color=COLOR['info_gray'], size=10, h='center')
    ws.row_dimensions[2].height = 20

    # 1. 평가 목적
    ws.merge_cells('A4:C4')
    sc(ws, 'A4', '■ 1. 평가 목적',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=12)
    ws.row_dimensions[4].height = 26

    for i, txt in enumerate([
        '• 객관적이고 공정한 평가를 통해 구성원의 역량과 성과를 체계적으로 파악',
        '• 성과 중심의 인사관리 기반 마련 및 승진·보상·인재 육성의 기준 제공',
        '• 구성원의 자기 인식·역량 개발 촉진과 성과 기반 보상 체계 확립',
    ]):
        r = 5 + i
        ws.merge_cells(f'A{r}:C{r}')
        sc(ws, f'A{r}', txt, size=10, indent=1)
        ws.row_dimensions[r].height = 20

    # 2. 평가 진행 절차 (STEP 1~7)
    ws.merge_cells('A9:C9')
    sc(ws, 'A9', '■ 2. 평가 진행 절차 (평가자가 해야 할 일)',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=12)
    ws.row_dimensions[9].height = 26

    steps = [
        ('STEP 1', '양식 복사',
         '평가 대상자의 역할등급에 맞는 양식 시트를 우클릭 → [시트 이동/복사 → 복사본 만들기] → '
         '시트명을 [평가_사번_성명]으로 변경 (예: 평가_BSS-001_홍길동)'),
        ('STEP 2', '정보 입력',
         '복사한 시트 상단의 [평가 대상자 정보]와 [평가자 정보] 주황색 셀에 '
         '사번·성명·직책·이메일·평가차수(상반기/하반기) 입력'),
        ('STEP 3', '역할평가',
         '5개 BARS 항목(R-1~R-5)에 1·2차 평가자가 각자 등급(S/A/B/C/D) 입력. '
         '화면 우측 [S/A/B/C/D 등급 기준] 참고하여 객관적 평가'),
        ('STEP 4', '공통역량평가',
         '11개 공통역량(C-1~C-11)에 1·2차 평가자가 각자 등급 입력. '
         '각 역량의 BARS 5단계 기준을 참고'),
        ('STEP 5', '직무역량평가',
         "직무 책무 목록의 D열(메인)에 평가 대상자의 메인 업무는 'O', 비메인은 'X' 표시 "
         "→ O 표시된 책무만 1·2차 등급 입력 (모든 책무 공통 BARS 기준은 직무역량 헤더 참조)"),
        ('STEP 6', '종합 소견',
         '강점·개선·육성 방향을 구체 사례 중심으로 작성'),
        ('STEP 7', 'HR 제출',
         '작성 완료된 평가지 파일을 HR 담당자에게 이메일로 송부 → 평가 절차 완료 '
         '(이후 처리는 HR 담당자가 통합관리시트에서 자동 산출)'),
    ]
    for i, (st, label, desc) in enumerate(steps):
        r = 10 + i
        sc(ws, f'A{r}', st, fill_color=COLOR['header_light'], bold=True, h='center', size=10)
        sc(ws, f'B{r}', label, fill_color=COLOR['callout'], bold=True, h='center', size=10)
        sc(ws, f'C{r}', desc, size=10)
        ws.row_dimensions[r].height = 36

    # 3. 평가 등급 체계
    ws.merge_cells('A18:C18')
    sc(ws, 'A18', '■ 3. 평가 등급 체계 (S/A/B/C/D)',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=12)
    ws.row_dimensions[18].height = 26

    grades = [
        ('S (탁월)', '기대 수준을 현저히 초과. 모범 사례 / 조직 전체 파급 효과 창출'),
        ('A (우수)', '기대 수준 상회. 자기주도적·선제적 우수 성과'),
        ('B (양호)', '기대 수준 부합. 책임과 과제를 안정적으로 기한 내 완수'),
        ('C (미흡)', '기대 대비 부분 미흡. 추가 지도·지원 필요'),
        ('D (부족)', '전반 미흡. 근본 개선 계획 필요'),
    ]
    for i, (g, desc) in enumerate(grades):
        r = 19 + i
        sc(ws, f'A{r}', g, fill_color=COLOR['header_light'], bold=True, h='center', size=10)
        ws.merge_cells(f'B{r}:C{r}')
        sc(ws, f'B{r}', desc, size=10)
        ws.row_dimensions[r].height = 22

    # 4. 평가 시 유의사항
    ws.merge_cells('A25:C25')
    sc(ws, 'A25', '■ 4. 평가 시 유의사항',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=12)
    ws.row_dimensions[25].height = 26

    cautions = [
        '• 객관적 사실에 근거하여 평가하세요. 평가 기간(6개월) 동안 관찰한 구체 행동·산출물을 기반으로',
        '• BARS(행동기준 등급) 5단계 기준을 반드시 참고. 각 등급의 행동 예시를 확인 후 판단',
        '• 개인 감정·관계가 아닌 업무 성과·역량 자체를 평가',
        '• B(양호)가 기준점. 기대 수준 부합 시 B 부여. 초과·미달 정도에 따라 S/A 또는 C/D',
        '• 1·2차 평가자는 독립적으로 평가. 자동 가중평균으로 종합 산출',
        '• 직무역량은 메인(O) 책무만 평가. 비메인(X)은 평가 제외 (점수에 영향 없음)',
        '• 평가 결과는 인사위원회·인사담당자 외에는 비공개. 평가 진행 중 평가지 보안에 주의',
    ]
    for i, txt in enumerate(cautions):
        r = 26 + i
        ws.merge_cells(f'A{r}:C{r}')
        sc(ws, f'A{r}', txt, size=10, indent=1)
        ws.row_dimensions[r].height = 20

    # 5. 평가 결과 활용
    ws.merge_cells('A34:C34')
    sc(ws, 'A34', '■ 5. 평가 결과 활용',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=12)
    ws.row_dimensions[34].height = 26

    uses = [
        '• 차년도 기본연봉 인상률 결정 (S 6% / A 5% / B 4% / C 3% / D 2%)',
        '• 성과급 지급률 결정 (S 20% / A 15% / B 10% / C 5% / D 0% · 월급여 기준)',
        '• 승진 심사 자료 (승진 연한 + 평가 등급 기준 충족 시 승진 대상자 상정)',
        '• 역량 개발·교육 훈련 계획 수립 자료',
    ]
    for i, txt in enumerate(uses):
        r = 35 + i
        ws.merge_cells(f'A{r}:C{r}')
        sc(ws, f'A{r}', txt, size=10, indent=1)
        ws.row_dimensions[r].height = 20

    # 6. 본 파일에 포함된 양식
    ws.merge_cells('A40:C40')
    sc(ws, 'A40', '■ 6. 본 파일에 포함된 양식 (역할등급별 7개 시트)',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=12)
    ws.row_dimensions[40].height = 26

    sc(ws, 'A41', '역할등급', fill_color=COLOR['header_light'], bold=True, h='center', size=10)
    sc(ws, 'B41', '역할명',   fill_color=COLOR['header_light'], bold=True, h='center', size=10)
    sc(ws, 'C41', '시트명 (해당 역할등급에 맞는 양식을 복사해 사용)',
       fill_color=COLOR['header_light'], bold=True, h='center', size=10)
    ws.row_dimensions[41].height = 22

    for i, (grade, role) in enumerate(GRADE_INFO):
        r = 42 + i
        sc(ws, f'A{r}', grade, h='center', bold=True, size=10)
        sc(ws, f'B{r}', role, size=10)
        sc(ws, f'C{r}', f'양식_{file_job_key}_{grade}', size=10)
        ws.row_dimensions[r].height = 22

    # 열 너비
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 76


def patch_job_competency_header(ws):
    """직무역량 섹션 헤더(③)를 두 줄로 만들어 BARS 등급 기준 추가"""
    target = None
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, 1)
        if isinstance(cell, MergedCell): continue
        v = cell.value
        if v and isinstance(v, str) and '③' in v and '직무역량' in v:
            target = r
            break
    if target is None: return False

    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == target and mr.max_row == target:
            ws.unmerge_cells(str(mr))

    ws.merge_cells(start_row=target, end_row=target,
                   start_column=1, end_column=8)
    cell = ws.cell(target, 1)
    cell.value = NEW_JOB_HEADER
    cell.fill = PatternFill('solid', fgColor=COLOR['header_mid'])
    cell.font = Font(name='맑은 고딕', size=10.5, bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='left', vertical='center',
                               wrap_text=True, indent=1)
    cell.border = BORDER_ALL
    ws.row_dimensions[target].height = 56
    return True


# ============ 1. 백업에서 복구 ============
backup_files = sorted(glob.glob(os.path.join(BACKUP_DIR, "평가지양식_*.xlsx")))
print(f"백업 파일: {len(backup_files)}개")
for src in backup_files:
    name = os.path.basename(src)
    dst = os.path.join(FINAL_DIR, name)
    try:
        if os.path.exists(dst): os.remove(dst)
        shutil.copy(src, dst)
        print(f"  복구: {name}")
    except Exception as e:
        print(f"  !! 복구 실패 {name}: {e}")

# ============ 2~4. 작성안내 재빌드 + 열너비 + 직무역량 헤더 ============
files = sorted(glob.glob(os.path.join(FINAL_DIR, "평가지양식_*.xlsx")))
print(f"\n폴리시 대상: {len(files)}개\n")

for fp in files:
    name = os.path.basename(fp)
    # 파일명에서 직무 키 추출 (평가지양식_BIM컨설팅.xlsx → BIM컨설팅)
    file_job_key = name.replace('평가지양식_', '').replace('.xlsx', '')
    job_name = JOB_NAME.get(file_job_key, file_job_key)

    wb = load_workbook(fp)

    # 00_작성안내 완전 재빌드
    if '00_작성안내' in wb.sheetnames:
        ws_guide = wb['00_작성안내']
        build_guide_sheet(ws_guide, job_name, file_job_key)

    # 평가지 시트들: 열너비 + 직무역량 헤더
    for sn in wb.sheetnames:
        if sn == '00_작성안내': continue
        ws = wb[sn]
        for col, w in STD_WIDTH.items():
            ws.column_dimensions[col].width = w
        patch_job_competency_header(ws)

    wb.save(fp)
    print(f"  ✓ {name}")

# ============ 검증 ============
print("\n=== 검증: BIM컨설팅 ===")
wb = load_workbook(files[0])

ws = wb['00_작성안내']
print("\n[00_작성안내 STEP 1~7 + 등급 체계 + 문의 제거 확인]:")
for r in [10, 11, 14, 16, 19, 23, 41, 42, 48, 50, 55]:
    a = ws.cell(r,1).value
    b = ws.cell(r,2).value
    c = ws.cell(r,3).value
    print(f"  행{r}: A={str(a)[:15]!r:18} B={str(b)[:18]!r:22} C={str(c)[:80] if c else '비어있음'}")

ws_g6 = wb['양식_BIM컨설팅_G6']
print(f"\n[직무역량 헤더 행 40 (두 줄 + BARS)]:")
print(f"  {ws_g6['A40'].value}")
print(f"  행 높이: {ws_g6.row_dimensions[40].height}")

print(f"\n[책무·종합·통합입력 무손실 확인]:")
for r in [41, 42, 43, 58, 59, 62, 63, 64]:
    row_data = [str(ws_g6.cell(r,c).value)[:25] if ws_g6.cell(r,c).value else '' for c in range(1, 9)]
    print(f"  행{r}: {row_data}")

print("\n=== 완료 ===")
