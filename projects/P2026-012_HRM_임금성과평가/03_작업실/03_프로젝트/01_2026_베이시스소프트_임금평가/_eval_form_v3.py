"""평가지 양식 7종 완전 재정비 v3
1. 백업 복구
2. 00_작성안내 STEP 1~7 빌드
3. 역할/공통 종합 등급 수식 D 버그 수정 (점수<0.5 시 빈 셀)
4. 직무역량 영역 컬럼 구조를 역할평가와 동일하게 (D열에 BARS 등급 기준)
   - 각 책무 행에 S/A/B/C/D 등급 기준 표시
   - 종합 등급도 동일 구조 + D 버그 수정
5. 직무역량 영역 이후 통합 입력 + 종합 소견 동적 재배치
6. 열 너비 통일
"""
import os, glob, shutil
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가"
BACKUP_DIR = os.path.join(BASE, "평가지_양식")
FINAL_DIR = os.path.join(BASE, "평가지_양식_최종")

THIN = Side(style='thin', color='BFBFBF')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLOR = {
    'header_dark':'1F4E78', 'header_mid':'2E74B5', 'header_light':'B4C7E7',
    'input_yellow':'FFF2CC', 'auto_green':'E2EFDA', 'info_gray':'F2F2F2',
    'bars_yellow':'FFF9E5', 'callout':'FFEB9C', 'mark':'C00000',
}

STD_WIDTH = {'A':7, 'B':14, 'C':24, 'D':38, 'E':9, 'F':10, 'G':10, 'H':13,
             'I':11, 'J':11, 'K':11, 'L':11}

JOB_NAME = {
    'BIM컨설팅': 'BIM컨설팅', 'SW판매영업': 'SW판매영업',
    '경영전략기획': '경영전략기획', '교육': '교육', '기술지원': '기술지원',
    '연구개발RD': '연구개발(R&D)', '프로젝트영업': '프로젝트영업',
}
GRADE_INFO = [
    ('G6','사업·조직 관리'), ('G5-M','직무 관리(부서장)'),
    ('G5-S','실무 스페셜리스트'), ('G4','실무 책임(PM)'),
    ('G3','핵심 실무'), ('G2','실무'), ('G1','실무 보조'),
]

BARS_JOB_TEXT = (
    "S  기대 현저히 초과\n"
    "A  기대 상회\n"
    "B  기대 부합\n"
    "C  부분 미흡\n"
    "D  전반 미흡"
)

W1 = "'02_의사결정·종합결과'!$B$67"  # 1차 평가자 가중치
W2 = "'02_의사결정·종합결과'!$B$68"  # 2차 평가자 가중치


def g2s(c):
    """등급 셀 → 점수 변환 (IFERROR 안전, 빈 셀은 0)"""
    return (f'IFERROR(IF({c}="S",5,IF({c}="A",4,IF({c}="B",3,'
            f'IF({c}="C",2,IF({c}="D",1,0))))),0)')


def grade_from_score(score_cell):
    """점수 → 등급 (D 버그 수정: 점수<0.5 일 때 빈 셀)"""
    return (f'=IF(OR({score_cell}="",{score_cell}<0.5),"",'
            f'IF({score_cell}>=4.5,"S",'
            f'IF({score_cell}>=3.5,"A",'
            f'IF({score_cell}>=2.5,"B",'
            f'IF({score_cell}>=1.5,"C","D")))))')


def sc(ws, coord, value=None, fill_color=None, bold=False, size=11, color='000000',
       h='left', v='center', wrap=True, border=BORDER_ALL, indent=0, number_fmt=None):
    c = ws[coord]
    if isinstance(c, MergedCell): return
    if value is not None: c.value = value
    c.font = Font(name='맑은 고딕', size=size, bold=bold, color=color)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)
    if fill_color: c.fill = PatternFill('solid', fgColor=fill_color)
    if border: c.border = border
    if number_fmt: c.number_format = number_fmt


def clear_rows(ws, start, end, max_col=13):
    for mr in list(ws.merged_cells.ranges):
        if start <= mr.min_row <= end or start <= mr.max_row <= end:
            ws.unmerge_cells(str(mr))
    for r in range(start, end + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell): continue
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font(name='맑은 고딕', size=11)
            cell.alignment = Alignment()
            cell.border = Border()
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None


def extract_duties(ws):
    """직무역량 영역의 책무 데이터 추출 [(no, area, duty), ...]"""
    duties = []
    in_duty_area = False
    for r in range(38, 80):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if a and isinstance(a, str) and a.startswith('J-'):
            duties.append((a, b or '', c or ''))
        elif duties and (not a or '평가자' in str(a) or '직무 종합' in str(a)):
            break
    return duties


def build_guide(ws, job_name, file_job_key):
    """00_작성안내 STEP 1~7 빌드"""
    clear_rows(ws, 1, 80)
    ws.merge_cells('A1:C1')
    sc(ws, 'A1', f'성과평가지 양식 - {job_name}',
       fill_color=COLOR['header_dark'], color='FFFFFF', bold=True, size=14, h='center')
    ws.row_dimensions[1].height = 30
    ws.merge_cells('A2:C2')
    sc(ws, 'A2', '베이시스소프트 1년 2회(상반기·하반기) 정기 성과평가 표준 양식',
       fill_color=COLOR['info_gray'], size=10, h='center')
    ws.row_dimensions[2].height = 20

    ws.merge_cells('A4:C4')
    sc(ws, 'A4', '■ 1. 평가 목적',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=12)
    ws.row_dimensions[4].height = 26
    for i, txt in enumerate([
        '• 객관적이고 공정한 평가를 통해 구성원의 역량과 성과를 체계적으로 파악',
        '• 성과 중심의 인사관리 기반 마련 및 승진·보상·인재 육성의 기준 제공',
        '• 구성원의 자기 인식·역량 개발 촉진과 성과 기반 보상 체계 확립',
    ]):
        r = 5 + i
        ws.merge_cells(f'A{r}:C{r}')
        sc(ws, f'A{r}', txt, size=10, indent=1)
        ws.row_dimensions[r].height = 20

    ws.merge_cells('A9:C9')
    sc(ws, 'A9', '■ 2. 평가 진행 절차 (평가자가 해야 할 일)',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=12)
    ws.row_dimensions[9].height = 26

    steps = [
        ('STEP 1', '양식 복사',
         '평가 대상자의 역할등급에 맞는 양식 시트를 우클릭 → [시트 이동/복사 → 복사본 만들기] → '
         '시트명을 [평가_사번_성명]으로 변경'),
        ('STEP 2', '정보 입력',
         '복사한 시트 상단의 [평가 대상자 정보]와 [평가자 정보] 주황색 셀에 사번·성명·직책·이메일·평가차수 입력'),
        ('STEP 3', '역할평가',
         '5개 BARS 항목(R-1~R-5)에 1·2차 평가자가 각자 등급(S/A/B/C/D) 입력. 각 항목 D열의 등급 기준 참고'),
        ('STEP 4', '공통역량평가',
         '11개 공통역량(C-1~C-11)에 1·2차 평가자가 각자 등급 입력. 각 역량의 BARS 5단계 기준 참고'),
        ('STEP 5', '직무역량평가',
         "직무 책무 목록의 E열(메인)에 메인 업무는 'O', 비메인은 'X' 표시 → O 표시된 책무만 1·2차 등급 입력. "
         "각 책무 D열의 S/A/B/C/D 등급 기준 참고"),
        ('STEP 6', '종합 소견', '강점·개선·육성 방향을 구체 사례 중심으로 작성'),
        ('STEP 7', 'HR 제출',
         '작성 완료된 평가지 파일을 HR 담당자에게 이메일로 송부 → 평가 절차 완료 '
         '(이후 처리는 HR 담당자가 통합관리시트에서 자동 산출)'),
    ]
    for i, (st, lbl, desc) in enumerate(steps):
        r = 10 + i
        sc(ws, f'A{r}', st, fill_color=COLOR['header_light'], bold=True, h='center', size=10)
        sc(ws, f'B{r}', lbl, fill_color=COLOR['callout'], bold=True, h='center', size=10)
        sc(ws, f'C{r}', desc, size=10)
        ws.row_dimensions[r].height = 36

    ws.merge_cells('A18:C18')
    sc(ws, 'A18', '■ 3. 평가 등급 체계 (S/A/B/C/D)',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=12)
    ws.row_dimensions[18].height = 26
    grades = [
        ('S (탁월)', '기대 수준을 현저히 초과. 모범 사례 / 조직 전체 파급 효과 창출'),
        ('A (우수)', '기대 수준 상회. 자기주도적·선제적 우수 성과'),
        ('B (양호)', '기대 수준 부합. 책임과 과제를 안정적으로 기한 내 완수'),
        ('C (미흡)', '기대 대비 부분 미흡. 추가 지도·지원 필요'),
        ('D (부족)', '전반 미흡. 근본 개선 계획 필요'),
    ]
    for i, (g, desc) in enumerate(grades):
        r = 19 + i
        sc(ws, f'A{r}', g, fill_color=COLOR['header_light'], bold=True, h='center', size=10)
        ws.merge_cells(f'B{r}:C{r}')
        sc(ws, f'B{r}', desc, size=10)
        ws.row_dimensions[r].height = 22

    ws.merge_cells('A25:C25')
    sc(ws, 'A25', '■ 4. 평가 시 유의사항',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=12)
    ws.row_dimensions[25].height = 26
    for i, txt in enumerate([
        '• 객관적 사실에 근거하여 평가하세요. 평가 기간(6개월) 동안 관찰한 구체 행동·산출물 기반',
        '• BARS(행동기준 등급) 5단계 기준을 반드시 참고. 각 등급의 행동 예시를 확인 후 판단',
        '• 개인 감정·관계가 아닌 업무 성과·역량 자체를 평가',
        '• B(양호)가 기준점. 기대 수준 부합 시 B 부여. 초과·미달 정도에 따라 S/A 또는 C/D',
        '• 1·2차 평가자는 독립적으로 평가. 통합관리시트에서 자동 가중평균 산출',
        '• 직무역량은 메인(O) 책무만 평가. 비메인(X)은 평가 제외',
        '• 평가 결과는 인사위원회·인사담당자 외에는 비공개. 보안 주의',
    ]):
        r = 26 + i
        ws.merge_cells(f'A{r}:C{r}')
        sc(ws, f'A{r}', txt, size=10, indent=1)
        ws.row_dimensions[r].height = 20

    ws.merge_cells('A34:C34')
    sc(ws, 'A34', '■ 5. 평가 결과 활용',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=12)
    ws.row_dimensions[34].height = 26
    for i, txt in enumerate([
        '• 차년도 기본연봉 인상률 (S 6% / A 5% / B 4% / C 3% / D 2%)',
        '• 성과급 지급률 (S 20% / A 15% / B 10% / C 5% / D 0%)',
        '• 승진 심사 자료 (승진 연한 + 등급 기준)',
        '• 역량 개발·교육 훈련 계획 수립 자료',
    ]):
        r = 35 + i
        ws.merge_cells(f'A{r}:C{r}')
        sc(ws, f'A{r}', txt, size=10, indent=1)
        ws.row_dimensions[r].height = 20

    ws.merge_cells('A40:C40')
    sc(ws, 'A40', '■ 6. 본 파일에 포함된 양식 (역할등급별 7개 시트)',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=12)
    ws.row_dimensions[40].height = 26
    sc(ws, 'A41', '역할등급', fill_color=COLOR['header_light'], bold=True, h='center', size=10)
    sc(ws, 'B41', '역할명', fill_color=COLOR['header_light'], bold=True, h='center', size=10)
    sc(ws, 'C41', '시트명', fill_color=COLOR['header_light'], bold=True, h='center', size=10)
    ws.row_dimensions[41].height = 22
    for i, (g, role) in enumerate(GRADE_INFO):
        r = 42 + i
        sc(ws, f'A{r}', g, h='center', bold=True, size=10)
        sc(ws, f'B{r}', role, size=10)
        sc(ws, f'C{r}', f'양식_{file_job_key}_{g}', size=10)
        ws.row_dimensions[r].height = 22

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 76


def fix_role_common_d_bug(ws):
    """역할/공통 종합 등급 수식 D 버그 수정 (점수<0.5 시 빈 셀)"""
    # 역할 종합 (행 21, 22 G열)
    for r in [21, 22]:
        cell = ws.cell(r, 7)
        if isinstance(cell, MergedCell): continue
        cell.value = grade_from_score(f'E{r}')
        cell.font = Font(name='맑은 고딕', size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill('solid', fgColor=COLOR['auto_green'])
        cell.border = BORDER_ALL
    # 공통 종합 (행 37, 38 H열, F열에 점수)
    for r in [37, 38]:
        cell = ws.cell(r, 8)
        if isinstance(cell, MergedCell): continue
        cell.value = grade_from_score(f'F{r}')
        cell.font = Font(name='맑은 고딕', size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill('solid', fgColor=COLOR['auto_green'])
        cell.border = BORDER_ALL


def build_job_area(ws, duties):
    """직무역량 영역 + 통합 입력 + 종합 소견 완전 재빌드.
       역할평가와 동일 컬럼 구조 (D열에 BARS 등급 기준).
    """
    n = len(duties)
    DUTY_START = 42
    duty_end = DUTY_START + n - 1
    SUM1_ROW = duty_end + 2
    SUM2_ROW = duty_end + 3
    HUB_HDR_ROW = SUM2_ROW + 2
    HUB_USAGE_ROW = HUB_HDR_ROW + 1
    HUB_COL_ROW = HUB_USAGE_ROW + 1
    HUB_DATA_ROW = HUB_COL_ROW + 1
    OPINION_HDR_ROW = HUB_DATA_ROW + 2
    OPINION_DATA_START = OPINION_HDR_ROW + 1

    # 클리어 (행 40 ~ OPINION_DATA_START+3)
    clear_rows(ws, 40, OPINION_DATA_START + 5)

    # 행 40: 섹션 헤더
    ws.merge_cells('A40:H40')
    sc(ws, 'A40',
       '▶ ③ 직무역량평가 — 메인 책무(E열 O 표시)만 1·2차 평가자 등급 입력',
       fill_color=COLOR['header_dark'], color='FFFFFF', bold=True, size=12,
       indent=1)
    ws.row_dimensions[40].height = 28

    # 행 41: 컬럼 헤더 (역할평가와 동일)
    headers = ['No.', '전문영역', '핵심 책무', 'S/A/B/C/D 등급 기준',
               '메인', '1차', '2차', '평균(참고)']
    for i, h in enumerate(headers):
        sc(ws, f'{chr(ord("A")+i)}41', h,
           fill_color=COLOR['header_dark'], color='FFFFFF', bold=True,
           size=10, h='center', wrap=True)
    ws.row_dimensions[41].height = 32

    # 책무 행
    for i, (no, area, duty) in enumerate(duties):
        r = DUTY_START + i
        sc(ws, f'A{r}', no, fill_color=COLOR['info_gray'], bold=True, h='center', size=10)
        sc(ws, f'B{r}', area, fill_color=COLOR['info_gray'], size=10, indent=1)
        sc(ws, f'C{r}', duty, fill_color=COLOR['info_gray'], size=10, indent=1)
        sc(ws, f'D{r}', BARS_JOB_TEXT, fill_color=COLOR['bars_yellow'],
           size=9, indent=1, h='left')
        sc(ws, f'E{r}', None, fill_color=COLOR['input_yellow'], bold=True, h='center', size=11)
        sc(ws, f'F{r}', None, fill_color=COLOR['input_yellow'], bold=True, h='center', size=11)
        sc(ws, f'G{r}', None, fill_color=COLOR['input_yellow'], bold=True, h='center', size=11)
        avg = (
            f'=IF(E{r}="O",IFERROR('
            f'IF(AND(F{r}<>"",G{r}<>""),'
            f'{W1}*{g2s(f"F{r}")}+{W2}*{g2s(f"G{r}")},'
            f'IF(F{r}<>"",{g2s(f"F{r}")},IF(G{r}<>"",{g2s(f"G{r}")},""))),"")'
            f',"")'
        )
        sc(ws, f'H{r}', avg, fill_color=COLOR['auto_green'],
           h='center', size=10, number_fmt='0.00')
        ws.row_dimensions[r].height = 85

    # 종합 행 1차
    r = SUM1_ROW
    ws.merge_cells(f'A{r}:D{r}')
    sc(ws, f'A{r}', '1차 평가자 직무 종합 (메인 책무 평균)',
       fill_color=COLOR['header_light'], bold=True, size=10, indent=1)
    # E: 점수
    score1 = (
        f'=IFERROR(SUMPRODUCT((E{DUTY_START}:E{duty_end}="O")*'
        f'IFERROR(IF(F{DUTY_START}:F{duty_end}="S",5,'
        f'IF(F{DUTY_START}:F{duty_end}="A",4,'
        f'IF(F{DUTY_START}:F{duty_end}="B",3,'
        f'IF(F{DUTY_START}:F{duty_end}="C",2,'
        f'IF(F{DUTY_START}:F{duty_end}="D",1,0))))),0))/'
        f'MAX(SUMPRODUCT((E{DUTY_START}:E{duty_end}="O")*(F{DUTY_START}:F{duty_end}<>"")),1),0)'
    )
    sc(ws, f'E{r}', score1, fill_color=COLOR['auto_green'],
       bold=True, h='center', size=11, number_fmt='0.00')
    sc(ws, f'F{r}', '→ 등급', fill_color=COLOR['header_light'],
       bold=True, h='center', size=10)
    sc(ws, f'G{r}', grade_from_score(f'E{r}'), fill_color=COLOR['auto_green'],
       bold=True, h='center', size=14)
    sc(ws, f'H{r}', None, fill_color=COLOR['info_gray'])
    ws.row_dimensions[r].height = 28

    # 종합 행 2차
    r = SUM2_ROW
    ws.merge_cells(f'A{r}:D{r}')
    sc(ws, f'A{r}', '2차 평가자 직무 종합 (메인 책무 평균)',
       fill_color=COLOR['header_light'], bold=True, size=10, indent=1)
    score2 = (
        f'=IFERROR(SUMPRODUCT((E{DUTY_START}:E{duty_end}="O")*'
        f'IFERROR(IF(G{DUTY_START}:G{duty_end}="S",5,'
        f'IF(G{DUTY_START}:G{duty_end}="A",4,'
        f'IF(G{DUTY_START}:G{duty_end}="B",3,'
        f'IF(G{DUTY_START}:G{duty_end}="C",2,'
        f'IF(G{DUTY_START}:G{duty_end}="D",1,0))))),0))/'
        f'MAX(SUMPRODUCT((E{DUTY_START}:E{duty_end}="O")*(G{DUTY_START}:G{duty_end}<>"")),1),0)'
    )
    sc(ws, f'E{r}', score2, fill_color=COLOR['auto_green'],
       bold=True, h='center', size=11, number_fmt='0.00')
    sc(ws, f'F{r}', '→ 등급', fill_color=COLOR['header_light'],
       bold=True, h='center', size=10)
    sc(ws, f'G{r}', grade_from_score(f'E{r}'), fill_color=COLOR['auto_green'],
       bold=True, h='center', size=14)
    sc(ws, f'H{r}', None, fill_color=COLOR['info_gray'])
    ws.row_dimensions[r].height = 28

    # 통합 입력 영역
    r = HUB_HDR_ROW
    ws.merge_cells(f'A{r}:H{r}')
    sc(ws, f'A{r}',
       '▶ 통합관리시트 입력 행 (아래 한 줄 A~G 7개 셀을 통째 복사 → 07a/07b의 G~M에 [값 붙여넣기])',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=11, indent=1)
    ws.row_dimensions[r].height = 24

    r = HUB_USAGE_ROW
    ws.merge_cells(f'A{r}:H{r}')
    sc(ws, f'A{r}',
       '사용법: 아래 한 줄 A~G 선택 → Ctrl+C → 통합관리시트 07a 또는 07b의 해당 사원 행 G열에 [선택하여 붙여넣기 → 값] 실행',
       fill_color=COLOR['info_gray'], size=9, indent=1)
    ws.row_dimensions[r].height = 22

    r = HUB_COL_ROW
    cols = ['역할(1차)', '역할(2차)', '공통(1차)', '공통(2차)', '직무(1차)', '직무(2차)', '소견']
    for i, h in enumerate(cols):
        sc(ws, f'{chr(ord("A")+i)}{r}', h,
           fill_color=COLOR['header_light'], bold=True, h='center', size=10)
    ws.row_dimensions[r].height = 24

    r = HUB_DATA_ROW
    refs = ['=G21', '=G22', '=H37', '=H38', f'=G{SUM1_ROW}', f'=G{SUM2_ROW}', '']
    for i, ref in enumerate(refs):
        col = chr(ord('A') + i)
        if ref:
            sc(ws, f'{col}{r}', ref, fill_color=COLOR['auto_green'],
               bold=True, h='center', size=12)
        else:
            sc(ws, f'{col}{r}', None, fill_color=COLOR['input_yellow'], size=10)
    ws.row_dimensions[r].height = 30

    # 종합 소견
    r = OPINION_HDR_ROW
    ws.merge_cells(f'A{r}:H{r}')
    sc(ws, f'A{r}', '▶ 종합 소견 (강점·개선·육성 방향)',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=11, indent=1)
    ws.row_dimensions[r].height = 24

    for i, label in enumerate(['강점', '개선', '육성 방향']):
        r = OPINION_DATA_START + i
        sc(ws, f'A{r}', label, fill_color=COLOR['header_light'],
           bold=True, h='center', size=10)
        ws.merge_cells(f'B{r}:H{r}')
        sc(ws, f'B{r}', None, fill_color=COLOR['input_yellow'], size=10, indent=1)
        ws.row_dimensions[r].height = 50


# ============ 1. 백업에서 복구 ============
backup_files = sorted(glob.glob(os.path.join(BACKUP_DIR, "평가지양식_*.xlsx")))
print(f"백업 파일: {len(backup_files)}개")
for src in backup_files:
    name = os.path.basename(src)
    dst = os.path.join(FINAL_DIR, name)
    try:
        if os.path.exists(dst): os.remove(dst)
        shutil.copy(src, dst)
        print(f"  복구: {name}")
    except Exception as e:
        print(f"  !! 복구 실패 {name}: {e}")

# ============ 2~5. 정비 ============
files = sorted(glob.glob(os.path.join(FINAL_DIR, "평가지양식_*.xlsx")))
print(f"\n정비 대상: {len(files)}개\n")

for fp in files:
    name = os.path.basename(fp)
    file_job_key = name.replace('평가지양식_', '').replace('.xlsx', '')
    job_name = JOB_NAME.get(file_job_key, file_job_key)
    wb = load_workbook(fp)

    # 00_작성안내
    if '00_작성안내' in wb.sheetnames:
        build_guide(wb['00_작성안내'], job_name, file_job_key)

    # 평가지 시트들
    for sn in wb.sheetnames:
        if sn == '00_작성안내': continue
        ws = wb[sn]

        # 책무 데이터 추출
        duties = extract_duties(ws)

        # D 버그 수정 (역할 + 공통)
        fix_role_common_d_bug(ws)

        # 직무역량 영역 + 통합 입력 + 종합 소견 재빌드
        build_job_area(ws, duties)

        # 열 너비
        for col, w in STD_WIDTH.items():
            ws.column_dimensions[col].width = w

    wb.save(fp)
    print(f"  ✓ {name} (책무 수 보존)")

# ============ 검증 ============
print("\n=== 검증 ===")
for fp in files[:2]:
    name = os.path.basename(fp)
    wb = load_workbook(fp)
    ws = wb[wb.sheetnames[1]]  # G6
    print(f"\n[{name} / {wb.sheetnames[1]}]:")

    # 역할 종합 (D 버그 수정 확인)
    print(f"  역할 종합 G21: {ws['G21'].value[:80] if ws['G21'].value else 'None'}")
    print(f"  공통 종합 H37: {ws['H37'].value[:80] if ws['H37'].value else 'None'}")

    # 직무역량 영역
    print(f"  직무역량 헤더 A40: {ws['A40'].value[:60] if ws['A40'].value else 'None'}")
    print(f"  컬럼 헤더 행41 A:H: {[ws.cell(41,c).value for c in range(1,9)]}")
    print(f"  첫 책무 행42 A:H: {[str(ws.cell(42,c).value)[:25] if ws.cell(42,c).value else '' for c in range(1,9)]}")

    # 종합 행 위치
    duties = extract_duties(ws)
    sum1 = 42 + len(duties) + 1
    print(f"  종합 1차 행 {sum1}: A={ws.cell(sum1,1).value} E={str(ws.cell(sum1,5).value)[:50]} G(등급)={str(ws.cell(sum1,7).value)[:60]}")
    print(f"  통합 입력 데이터: A={ws.cell(sum1+4,1).value} E={ws.cell(sum1+4,5).value} F={ws.cell(sum1+4,6).value}")

print("\n=== 완료 ===")
