"""v13 진짜 최종본 - 02_의사결정 절대/상대평가 드롭다운 + 강조"""
import os, shutil, subprocess, glob, random
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

random.seed(2026)

COLOR = {'header_dark':'1F4E78','header_mid':'2E74B5','header_light':'B4C7E7',
    'input_yellow':'FFF2CC','auto_green':'E2EFDA','info_gray':'F2F2F2',
    's_grade':'C6E0B4','a_grade':'D9E1F2','c_grade':'FFEB9C','d_grade':'FFC7CE',
    'toggle_on':'FFD966','toggle_strong':'C00000','mute':'D9D9D9'}
def fill(c): return PatternFill('solid', fgColor=c)
def font(size=11, bold=False, color='000000'):
    return Font(name='맑은 고딕', size=size, bold=bold, color=color)
def align(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
THIN = Side(style='thin', color='BFBFBF')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
THICK = Side(style='medium', color='C00000')
BORDER_TOGGLE = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

def sc(ws, coord, value=None, fill_color=None, bold=False, size=11, color='000000',
       h='left', v='center', wrap=True, border=BORDER_ALL, number_fmt=None):
    c = ws[coord]
    if isinstance(c, MergedCell): return
    if value is not None: c.value = value
    c.font = font(size=size, bold=bold, color=color)
    c.alignment = align(h=h, v=v, wrap=wrap)
    if fill_color: c.fill = fill(fill_color)
    if border: c.border = border
    if number_fmt: c.number_format = number_fmt

SRC = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가/베이시스소프트_HR통합관리시트_v12.xlsx"
TMP = "/tmp/master_v13.xlsx"
shutil.copy(SRC, TMP)
wb = load_workbook(TMP)

# ============ 02_의사결정·종합결과 시트 절대/상대평가 토글 강화 ============
ws = wb['02_의사결정·종합결과']

# 기존 머지 정리 (행 4~12 영역)
for mr in list(ws.merged_cells.ranges):
    if 4 <= mr.min_row <= 13:
        ws.unmerge_cells(str(mr))

# ① 평가 방식·반기 가중치 박스 재설계 - 강조 디자인
# 행 4: 섹션 헤더
ws.merge_cells('A4:G4')
sc(ws, 'A4', '■ ① 평가 방식 · 반기 가중치 (먼저 선택)',
   fill_color=COLOR['header_dark'], color='FFFFFF', bold=True, size=13, h='left')
ws.row_dimensions[4].height = 28

# 행 5: 평가 척도
sc(ws, 'A5', '평가 척도', fill_color=COLOR['header_light'], bold=True, h='center')
sc(ws, 'B5', '5점', fill_color=COLOR['input_yellow'], bold=True, h='center', size=12)
ws.merge_cells('C5:G5')
sc(ws, 'C5', '5점 / 100점 중 선택 (현재 시스템은 5점 BARS 척도 기반)',
   fill_color=COLOR['info_gray'], h='left', size=10)
ws.row_dimensions[5].height = 28

# 행 6: ★ 평가 방식 (절대/상대) - 가장 강조 ★
sc(ws, 'A6', '★ 평가 방식', fill_color=COLOR['toggle_strong'], color='FFFFFF',
   bold=True, h='center', size=12)
sc(ws, 'B6', '절대평가', fill_color=COLOR['toggle_on'], bold=True,
   h='center', size=14, color='C00000', border=BORDER_TOGGLE)
ws.merge_cells('C6:G6')
sc(ws, 'C6',
   '▼ 셀 클릭 → 드롭다운에서 [절대평가 / 상대평가] 선택 (08_종합평가 최종 등급에 즉시 반영)',
   fill_color=COLOR['c_grade'], bold=True, h='left', size=10)
ws.row_dimensions[6].height = 34

# 행 7: 상반기 가중치
sc(ws, 'A7', '상반기 가중치', fill_color=COLOR['header_light'], bold=True, h='center')
sc(ws, 'B7', 0.5, fill_color=COLOR['input_yellow'], h='center', number_fmt='0%', bold=True)
ws.merge_cells('C7:G7')
sc(ws, 'C7', '상반기 평가의 연간 가중치 (0~1). 1차/2차 평가자 가중치는 ⑥번 항목 참고.',
   fill_color=COLOR['info_gray'], h='left', size=10)
ws.row_dimensions[7].height = 24

# 행 8: 하반기 가중치 (자동)
sc(ws, 'A8', '하반기 가중치', fill_color=COLOR['header_light'], bold=True, h='center')
sc(ws, 'B8', '=1-B7', fill_color=COLOR['auto_green'], h='center', number_fmt='0%', bold=True)
ws.merge_cells('C8:G8')
sc(ws, 'C8', '자동 계산 (1 - 상반기). 합계 100% 강제.',
   fill_color=COLOR['info_gray'], h='left', size=10)
ws.row_dimensions[8].height = 24

# 행 9: 안내
ws.merge_cells('A9:G9')
sc(ws, 'A9',
   '💡 절대평가: 평가 점수에 따라 등급 산출 (S≥4.5, A≥3.5, B≥2.5, C≥1.5, 그 외 D)  ┃  '
   '상대평가: 행 11의 강제분포(S/A/B/C/D 비중)에 맞춰 PERCENTRANK 기반 등급 부여',
   fill_color=COLOR['header_light'], h='left', size=10, wrap=True)
ws.row_dimensions[9].height = 36

# 행 10: 강제 분포 헤더
sc(ws, 'A10', '강제분포 비중 (상대평가 시)', fill_color=COLOR['header_dark'],
   color='FFFFFF', bold=True, h='center')
for i, g in enumerate(['S','A','B','C','D']):
    sc(ws, f'{get_column_letter(2+i)}10', g, fill_color=COLOR['header_mid'],
       color='FFFFFF', bold=True, h='center')
sc(ws, 'G10', '합계', fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, h='center')
ws.row_dimensions[10].height = 24

# 행 11: 강제 분포 값
sc(ws, 'A11', '비중', fill_color=COLOR['header_light'], bold=True, h='center')
for i, v in enumerate([0.10, 0.15, 0.50, 0.15, 0.10]):
    col = get_column_letter(2+i)
    sc(ws, f'{col}11', v, fill_color=COLOR['input_yellow'], h='center',
       number_fmt='0%', bold=True)
sc(ws, 'G11', '=SUM(B11:F11)', fill_color=COLOR['auto_green'], h='center',
   number_fmt='0%', bold=True)
ws.row_dimensions[11].height = 24

# 행 12: 강제분포 검증
ws.merge_cells('A12:G12')
sc(ws, 'A12',
   '=IF(G11=1,"✓ 강제분포 합계 100% OK","✗ 강제분포 합계 ≠ 100% — 수정 필요")',
   fill_color=COLOR['info_gray'], h='center', bold=True, size=11)
ws.row_dimensions[12].height = 22

# ============ 데이터 유효성 검사(드롭다운) 설치 ============
# 기존 DV 클리어 (중복 방지)
ws.data_validations.dataValidation = []

# B6: 절대평가/상대평가 드롭다운
dv_method = DataValidation(type="list", formula1='"절대평가,상대평가"', allow_blank=False)
dv_method.error = '절대평가 또는 상대평가만 선택 가능합니다'
dv_method.errorTitle = '평가 방식 선택'
dv_method.prompt = '드롭다운에서 평가 방식을 선택하세요'
dv_method.promptTitle = '★ 평가 방식'
dv_method.showErrorMessage = True
dv_method.showInputMessage = True
ws.add_data_validation(dv_method)
dv_method.add('B6')

# B5: 평가 척도 드롭다운
dv_scale = DataValidation(type="list", formula1='"5점,100점"', allow_blank=False)
dv_scale.prompt = '5점 또는 100점 척도 선택'
dv_scale.promptTitle = '평가 척도'
dv_scale.showInputMessage = True
ws.add_data_validation(dv_scale)
dv_scale.add('B5')

# B7: 0~1 사이 비율
dv_w = DataValidation(type="decimal", operator="between",
                     formula1=0, formula2=1, allow_blank=False)
dv_w.prompt = '0~1 사이의 값을 입력 (예: 0.5)'
dv_w.promptTitle = '상반기 가중치'
dv_w.showInputMessage = True
ws.add_data_validation(dv_w)
dv_w.add('B7')

# B11:F11: 강제분포 0~1
dv_dist = DataValidation(type="decimal", operator="between",
                        formula1=0, formula2=1, allow_blank=False)
dv_dist.prompt = '0~1 사이의 비율. 5개 셀 합이 1이 되도록 입력'
dv_dist.promptTitle = '강제분포'
dv_dist.showInputMessage = True
ws.add_data_validation(dv_dist)
dv_dist.add('B11:F11')

# ============ 조건부 서식: 토글 상태 시각화 ============
ws.conditional_formatting._cf_rules.clear()  # 02 시트 한정 클리어

# B6가 "상대평가"면 토글 셀 색상이 진해짐
ws.conditional_formatting.add('B6',
    FormulaRule(formula=['$B$6="상대평가"'],
                fill=fill('70AD47'), font=Font(name='맑은 고딕', bold=True, size=14, color='FFFFFF')))
ws.conditional_formatting.add('B6',
    FormulaRule(formula=['$B$6="절대평가"'],
                fill=fill(COLOR['toggle_on']), font=Font(name='맑은 고딕', bold=True, size=14, color='C00000')))

# 강제분포 셀: 상대평가 선택 시만 활성화 (절대평가 시 회색 처리)
ws.conditional_formatting.add('B11:F11',
    FormulaRule(formula=['$B$6="절대평가"'],
                fill=fill(COLOR['mute']), font=Font(name='맑은 고딕', color='999999')))

# 컬럼 너비
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 18

print("02_의사결정 - 절대/상대평가 드롭다운 + 시각 강조 완료")

# ============ 00_운영가이드에도 안내 추가 ============
try:
    ws_g = wb['00_운영가이드']
    # 안내 박스를 찾기 위해 행 1~30 스캔, 비어있는 영역에 추가
    # 가장 안전한 방법: 마지막 행 다음에 추가
    last_r = 1
    for r in range(1, 100):
        for c in range(1, 10):
            v = ws_g.cell(r, c).value
            if v is not None:
                last_r = r
                break

    next_r = last_r + 2
    ws_g.merge_cells(f'A{next_r}:H{next_r}')
    sc(ws_g, f'A{next_r}',
       '★ 평가 방식 선택 (절대평가 / 상대평가) - 02_의사결정 시트 B6 셀에서 드롭다운으로 선택',
       fill_color=COLOR['toggle_strong'], color='FFFFFF', bold=True, size=12, h='center')
    ws_g.row_dimensions[next_r].height = 28

    next_r += 1
    ws_g.merge_cells(f'A{next_r}:H{next_r}')
    sc(ws_g, f'A{next_r}',
       '• 절대평가: 점수→등급 자동 환산 (S≥4.5 / A≥3.5 / B≥2.5 / C≥1.5 / D)\n'
       '• 상대평가: 02_의사결정 행 11의 강제분포(S 10%·A 15%·B 50%·C 15%·D 10% 등)에 맞춰 PERCENTRANK 기반 부여\n'
       '• 전환 효과: 08_종합평가 U열(최종 등급) → 09_보상연계 → 10_성과급관리 → 12_시뮬레이션 까지 즉시 반영',
       fill_color=COLOR['c_grade'], h='left', size=10, wrap=True)
    ws_g.row_dimensions[next_r].height = 64
    print("00_운영가이드 - 절대/상대평가 안내 추가")
except KeyError:
    print("00_운영가이드 없음 - 스킵")

# 저장
wb.save(TMP)
print(f"\n저장: {TMP}")

OUT = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가/베이시스소프트_HR통합관리시트_FINAL.xlsx"
try:
    if os.path.exists(OUT):
        os.remove(OUT)
    shutil.copy(TMP, OUT)
    print(f"통합시트 저장: {OUT}")
except PermissionError:
    OUT = OUT.replace('.xlsx','_v13.xlsx')
    shutil.copy(TMP, OUT)
    print(f"통합시트 저장(잠금): {OUT}")

# ============ 시뮬레이션 결과 재생성 (상대평가 데모) ============
SIM = "/tmp/sim_v13.xlsx"
shutil.copy(TMP, SIM)
wb_s = load_workbook(SIM)

DATA_START = 6
DATA_END = 55

# 이름 마스킹
ws = wb_s['01_인적정보']
for r in range(DATA_START, DATA_END+1):
    bss = ws.cell(r, 2).value
    if bss and str(bss).startswith('BSS'):
        try: ws.cell(r, 3).value = "OOO"
        except: pass

ws = wb_s['15_직원별책무매핑']
for r in range(4, 80):
    cell = ws.cell(r, 3)
    if isinstance(cell, MergedCell): continue
    v = cell.value
    if v and isinstance(v, str) and 'BSS' not in v:
        if '(' in v:
            grade_part = v[v.index('('):]
            cell.value = f"OOO {grade_part}"
        elif v not in ['No.','사번','성명','부서','메인 책무 수'] and len(v) <= 5:
            if any('가' <= ch <= '힣' for ch in v):
                cell.value = "OOO"

# 가상 등급 (34명)
def gen():
    r = random.random()
    return 'S' if r<0.10 else 'A' if r<0.30 else 'B' if r<0.80 else 'C' if r<0.95 else 'D'

for sn in ['07a_상반기입력','07b_하반기입력']:
    ws = wb_s[sn]
    for r in range(DATA_START, DATA_START + 34):
        for c in range(7, 13):
            ws.cell(r, c).value = gen()

# 02 의사결정 - 상대평가 + 평가자 가중치 0.3/0.7
ws = wb_s['02_의사결정·종합결과']
ws['B67'].value = 0.3
ws['B6'].value = '상대평가'  # 데모는 상대평가로

wb_s.save(SIM)

# LO 변환
out_dir = "/tmp/sim_v13_check"
shutil.rmtree(out_dir, ignore_errors=True)
os.makedirs(out_dir)
subprocess.run(['libreoffice','--headless','--calc','--convert-to','xlsx',
                '--outdir',out_dir,SIM], capture_output=True, timeout=120)
conv = glob.glob(out_dir + "/*.xlsx")

OUT_SIM = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가/베이시스소프트_HR_시뮬레이션결과.xlsx"
try:
    shutil.copy(conv[0], OUT_SIM)
    print(f"시뮬레이션 저장: {OUT_SIM}")
except PermissionError:
    OUT_SIM = OUT_SIM.replace('.xlsx','_v13.xlsx')
    shutil.copy(conv[0], OUT_SIM)
    print(f"시뮬레이션 저장(잠금): {OUT_SIM}")

# 검증
wb_c = load_workbook(conv[0], data_only=True)
ws = wb_c['02_의사결정·종합결과']
print("\n=== 02_의사결정 토글 검증 ===")
print(f"  B5 (평가 척도): {ws['B5'].value}")
print(f"  B6 (★ 평가 방식): {ws['B6'].value}")
print(f"  B7 (상반기 W): {ws['B7'].value}")
print(f"  B8 (하반기 W): {ws['B8'].value}")
print(f"  B11~F11 (강제분포): {[ws.cell(11,c).value for c in range(2,7)]}")
print(f"  G11 (분포 합): {ws['G11'].value}")
print(f"  B67 (1차 평가자 W): {ws['B67'].value}")

ws = wb_c['08_종합평가']
print("\n=== 08_종합평가 절대 vs 상대 비교 (행 6~10) ===")
for r in range(6, 11):
    b=ws.cell(r,2).value; s=ws.cell(r,19).value; t=ws.cell(r,20).value; u=ws.cell(r,21).value
    print(f"  {b}: 점수={s}  절대(T)={t}  최종(U)={u}")

# 등급 분포 검증
print("\n=== 최종등급 분포 (상대평가 모드) ===")
from collections import Counter
grades = []
for r in range(DATA_START, DATA_END+1):
    u = ws.cell(r, 21).value
    if u: grades.append(u)
print(f"  분포: {Counter(grades)}")

print("\n[v13 진짜 최종본 완료]")
