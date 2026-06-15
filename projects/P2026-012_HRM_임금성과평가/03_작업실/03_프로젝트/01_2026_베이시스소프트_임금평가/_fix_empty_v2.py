"""빈 셀 참조 0 평가 버그 수정 v2 (LEN 체크 기반)"""
import os, shutil, subprocess, glob, zipfile
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

BASE = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가"
SRC = os.path.join(BASE, "베이시스소프트_HR통합관리시트_FINAL.xlsx")
TMP = "/tmp/master_fix_v2.xlsx"
try:
    zipfile.ZipFile(SRC).close()
    shutil.copy(SRC, TMP)
except Exception:
    shutil.copy("/sessions/compassionate-bold-ride/mnt/uploads/베이시스소프트_HR통합관리시트_FINAL.xlsx", TMP)

wb = load_workbook(TMP)

DATA_START = 6
EMP_MAX = 50
DATA_END = DATA_START + EMP_MAX - 1
HALF_OFFSET = 58

W1 = "'02_의사결정·종합결과'!$B$67"
W2 = "'02_의사결정·종합결과'!$B$68"
W1H = "'02_의사결정·종합결과'!$B$7"
W2H = "'02_의사결정·종합결과'!$B$8"

def g2s(c):
    return f'IFERROR(CHOOSE(FIND({c},"DCBAS"),1,2,3,4,5),0)'

def score_formula(c1, c2, w1, w2):
    return (f'=IFERROR('
            f'IF(AND(LEN({c1})>0,LEN({c2})>0),{w1}*{g2s(c1)}+{w2}*{g2s(c2)},'
            f'IF(LEN({c1})>0,{g2s(c1)},'
            f'IF(LEN({c2})>0,{g2s(c2)},""))),"")')

def grade_formula(sc):
    return (f'=IF(OR({sc}="",NOT(ISNUMBER({sc})),{sc}<0.5),"",'
            f'IF({sc}>=4.5,"S",IF({sc}>=3.5,"A",IF({sc}>=2.5,"B",'
            f'IF({sc}>=1.5,"C","D")))))')

def ref_cell(src_sheet, src_col, ref):
    return (f'=IF(OR(\'01_인적정보\'!B{ref}="",{src_sheet}!{src_col}{ref}=""),"",'
            f'{src_sheet}!{src_col}{ref})')


def patch_eval(ws, cg1, cg2, cs, cf, s1, s2):
    fixed = 0
    for half_idx, sh in enumerate(["'07a_상반기입력'", "'07b_하반기입력'"]):
        for i in range(EMP_MAX):
            ref = DATA_START + i
            r = ref + (half_idx * HALF_OFFSET)
            for col, src in [(cg1, s1), (cg2, s2)]:
                cell = ws[f'{col}{r}']
                if not isinstance(cell, MergedCell):
                    cell.value = ref_cell(sh, src, ref)
                    fixed += 1
            cell = ws[f'{cs}{r}']
            if not isinstance(cell, MergedCell):
                cell.value = score_formula(f'{cg1}{r}', f'{cg2}{r}', W1, W2)
                fixed += 1
            cell = ws[f'{cf}{r}']
            if not isinstance(cell, MergedCell):
                cell.value = grade_formula(f'{cs}{r}')
                fixed += 1
    return fixed

print(f"04 수정: {patch_eval(wb['04_역할평가'], 'F','G','H','I', 'G','H')}")
print(f"05 수정: {patch_eval(wb['05_공통역량평가'], 'F','G','H','I', 'I','J')}")
print(f"06 수정: {patch_eval(wb['06_직무역량평가'], 'G','H','I','J', 'K','L')}")

# 08_종합평가
ws = wb['08_종합평가']
for i in range(EMP_MAX):
    r = DATA_START + i
    rh = r + HALF_OFFSET
    ws[f'G{r}'].value = ref_cell("'04_역할평가'", 'I', r)
    ws[f'H{r}'].value = (f'=IF(OR(\'01_인적정보\'!B{r}="",\'04_역할평가\'!I{rh}=""),"",'
                        f'\'04_역할평가\'!I{rh})')
    ws[f'I{r}'].value = score_formula(f'G{r}', f'H{r}', W1H, W2H)
    ws[f'J{r}'].value = ref_cell("'05_공통역량평가'", 'I', r)
    ws[f'K{r}'].value = (f'=IF(OR(\'01_인적정보\'!B{r}="",\'05_공통역량평가\'!I{rh}=""),"",'
                        f'\'05_공통역량평가\'!I{rh})')
    ws[f'L{r}'].value = score_formula(f'J{r}', f'K{r}', W1H, W2H)
    ws[f'M{r}'].value = ref_cell("'06_직무역량평가'", 'J', r)
    ws[f'N{r}'].value = (f'=IF(OR(\'01_인적정보\'!B{r}="",\'06_직무역량평가\'!J{rh}=""),"",'
                        f'\'06_직무역량평가\'!J{rh})')
    ws[f'O{r}'].value = score_formula(f'M{r}', f'N{r}', W1H, W2H)
    ws[f'S{r}'].value = (f'=IF(OR(NOT(ISNUMBER(I{r})),NOT(ISNUMBER(L{r})),'
                        f'NOT(ISNUMBER(O{r}))),"",I{r}*P{r}+L{r}*Q{r}+O{r}*R{r})')
    ws[f'T{r}'].value = grade_formula(f'S{r}')
    sr = f"$S${DATA_START}:$S${DATA_END}"
    pct = f'IFERROR(RANK(S{r},{sr},0)/MAX(COUNTIF({sr},">0"),1),0)'
    rel = (f'IF({pct}<=\'02_의사결정·종합결과\'!$B$11,"S",'
           f'IF({pct}<=\'02_의사결정·종합결과\'!$B$11+\'02_의사결정·종합결과\'!$C$11,"A",'
           f'IF({pct}<=\'02_의사결정·종합결과\'!$B$11+\'02_의사결정·종합결과\'!$C$11+\'02_의사결정·종합결과\'!$D$11,"B",'
           f'IF({pct}<=\'02_의사결정·종합결과\'!$B$11+\'02_의사결정·종합결과\'!$C$11+\'02_의사결정·종합결과\'!$D$11+\'02_의사결정·종합결과\'!$E$11,"C","D"))))')
    ws[f'U{r}'].value = (f'=IF(OR(S{r}="",NOT(ISNUMBER(S{r})),S{r}<0.5),"",'
                        f'IF(\'02_의사결정·종합결과\'!$B$6="절대평가",T{r},{rel}))')
print("08 수정 완료")

wb.save(TMP)
OUT = SRC
try:
    if os.path.exists(OUT): os.remove(OUT)
    shutil.copy(TMP, OUT)
    print(f"저장: {OUT}")
except PermissionError:
    OUT = OUT.replace('.xlsx', '_FIX.xlsx')
    shutil.copy(TMP, OUT)
    print(f"잠금 - 저장: {OUT}")

# LO 변환 검증
od = '/tmp/master_fix_v2_check'
shutil.rmtree(od, ignore_errors=True); os.makedirs(od)
subprocess.run(['libreoffice','--headless','--calc','--convert-to','xlsx',
                '--outdir',od,TMP], capture_output=True, timeout=120)
conv = glob.glob(od + "/*.xlsx")[0]
wbc = load_workbook(conv, data_only=True)
ws = wbc['08_종합평가']

print("\n=== 빈 입력 검증 (5명 샘플) ===")
print(f"{'사번':10} {'역점수':8} {'공점수':8} {'직점수':8} {'종합':8} {'절대':6} {'최종':6}")
for r in range(6, 11):
    b = ws.cell(r,2).value
    if not b: continue
    vals = [str(ws.cell(r,c).value)[:7] for c in [9,12,15,19,20,21]]
    print(f"{str(b):10} {vals[0]:8} {vals[1]:8} {vals[2]:8} {vals[3]:8} {vals[4]:6} {vals[5]:6}")
