"""평가지 양식 7종 안전 재정비
1. 백업(./평가지_양식/)에서 ./평가지_양식_최종/으로 복구
2. 00_작성안내: 문의 제거 + STEP 7 평가자 입장으로 변경
3. 모든 시트 열 너비 통일
4. 직무역량 섹션 헤더 행(행 40)을 두 줄로 만들어 BARS 등급 기준 추가
   (insert_rows 미사용 - 행 구조 보존, 데이터 무손실)
"""
import os, glob, shutil
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가"
BACKUP_DIR = os.path.join(BASE, "평가지_양식")
FINAL_DIR = os.path.join(BASE, "평가지_양식_최종")

THIN = Side(style='thin', color='BFBFBF')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 표준 열 너비
STD_WIDTH = {'A':7, 'B':14, 'C':24, 'D':38, 'E':38, 'F':8, 'G':8, 'H':11,
             'I':11, 'J':11, 'K':11, 'L':11}

NEW_STEP7_DESC = (
    "작성 완료된 평가지 파일을 HR 담당자에게 이메일로 송부 → "
    "평가 절차 완료 (이후 처리는 HR 담당자가 통합관리시트에서 자동 산출)"
)

NEW_JOB_HEADER = (
    "▶ ③ 직무역량평가 - 메인 책무(D열에 O 표시된 책무만) 1·2차 평가자 등급 입력\n"
    "★ S/A/B/C/D 등급 기준 (모든 책무 공통 적용):   "
    "S(탁월) 책무 기대 현저 초과   ┃   "
    "A(우수) 기대 상회   ┃   "
    "B(양호) 기대 부합   ┃   "
    "C(미흡) 일부 미달   ┃   "
    "D(부족) 전반 미흡"
)

# ============ 1. 백업에서 복구 ============
backup_files = sorted(glob.glob(os.path.join(BACKUP_DIR, "평가지양식_*.xlsx")))
print(f"백업 파일: {len(backup_files)}개")
for src in backup_files:
    name = os.path.basename(src)
    dst = os.path.join(FINAL_DIR, name)
    try:
        if os.path.exists(dst): os.remove(dst)
        shutil.copy(src, dst)
        print(f"  복구: {name}")
    except Exception as e:
        print(f"  !! 복구 실패 {name}: {e}")

# ============ 2~4. 폴리시 + 헤더 BARS ============
files = sorted(glob.glob(os.path.join(FINAL_DIR, "평가지양식_*.xlsx")))
print(f"\n폴리시 대상: {len(files)}개\n")

for fp in files:
    name = os.path.basename(fp)
    wb = load_workbook(fp)

    # --- 00_작성안내: STEP 7 변경 + 문의 제거 ---
    if '00_작성안내' in wb.sheetnames:
        ws = wb['00_작성안내']
        # STEP 7 찾기
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if a and isinstance(a, str) and 'STEP 7' in a:
                ws.cell(r, 3).value = NEW_STEP7_DESC
                break

        # 문의 섹션 행 삭제(셀 값 None)
        rows_to_clear = []
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if a and isinstance(a, str) and '문의' in a and \
               ('■' in a or '담당자' in a or '컨설턴트' in a):
                rows_to_clear.append(r)
        for r in rows_to_clear:
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row <= r <= mr.max_row:
                    ws.unmerge_cells(str(mr))
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if not isinstance(cell, MergedCell):
                    cell.value = None
                    cell.fill = PatternFill()
                    cell.border = Border()
                    cell.font = Font(name='맑은 고딕', size=11)

    # --- 열 너비 통일 + 직무역량 헤더 두 줄로 ---
    for sn in wb.sheetnames:
        ws = wb[sn]
        if sn == '00_작성안내':
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 80
            for col in 'DEFGHIJKL':
                ws.column_dimensions[col].width = 12
            continue

        # 평가지 시트 열 너비
        for col, w in STD_WIDTH.items():
            ws.column_dimensions[col].width = w

        # 직무역량 섹션 헤더 (③) 찾아 두 줄로 변경
        target = None
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(r, 1)
            if isinstance(cell, MergedCell): continue
            v = cell.value
            if v and isinstance(v, str) and '③' in v and '직무역량' in v:
                target = r
                break
        if target is None:
            continue

        # 기존 머지 해제
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row == target and mr.max_row == target:
                ws.unmerge_cells(str(mr))

        # A:H 머지로 재설정
        ws.merge_cells(start_row=target, end_row=target,
                       start_column=1, end_column=8)
        cell = ws.cell(target, 1)
        cell.value = NEW_JOB_HEADER
        cell.fill = PatternFill('solid', fgColor='2E74B5')
        cell.font = Font(name='맑은 고딕', size=10.5, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='left', vertical='center',
                                   wrap_text=True, indent=1)
        cell.border = BORDER_ALL
        ws.row_dimensions[target].height = 56

    wb.save(fp)
    print(f"  ✓ {name}")

# ============ 검증 ============
print("\n=== 검증: BIM컨설팅 G6 시트 ===")
wb = load_workbook(files[0])
ws = wb['양식_BIM컨설팅_G6']
print(f"\n[직무역량 헤더 행 40 내용]:")
print(f"  {ws['A40'].value}")
print(f"\n[행 41~45 컬럼 헤더 + 책무 데이터 - 무손실 확인]:")
for r in [41, 42, 43, 44]:
    row_data = [str(ws.cell(r,c).value)[:25] if ws.cell(r,c).value else '' for c in range(1, 10)]
    print(f"  행{r}: {row_data}")

print(f"\n[종합 + 통합 입력 - 무손실 확인]:")
for r in [58, 59, 62, 63, 64]:
    row_data = [str(ws.cell(r,c).value)[:25] if ws.cell(r,c).value else '' for c in range(1, 10)]
    print(f"  행{r}: {row_data}")

# 작성안내 STEP 7 + 문의 제거 확인
ws2 = wb['00_작성안내']
print(f"\n[00_작성안내 검증]:")
for r in [16, 50, 51]:
    a = ws2.cell(r,1).value
    b = ws2.cell(r,2).value
    c = ws2.cell(r,3).value
    print(f"  행{r}: A={a!r}  B={b!r}  C={str(c)[:60] if c else '비어있음'}")

print("\n=== 완료 ===")
