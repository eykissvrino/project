"""최종본 기반 시뮬레이션
- 입력: 사용자가 편집 완료한 최종 통합관리시트 (상대평가, 강제분포 10/15/50/15/10)
- 가상 평가: B 중심 정규분포, 1·2차 변동, 상·하반기 트렌드, 영역별 다양성
- 이름 OOO 마스킹
- LO 변환으로 계산값 확정
"""
import os, shutil, subprocess, glob, random
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

random.seed(2026)

SRC = "/sessions/compassionate-bold-ride/mnt/uploads/베이시스소프트_HR통합관리시트_FINAL-5e0b9760.xlsx"
TMP = "/tmp/sim_final_2026.xlsx"
shutil.copy(SRC, TMP)

wb = load_workbook(TMP)

# ============ 1. 이름 마스킹 ============
ws = wb['01_인적정보']
masked_count = 0
for r in range(6, 56):
    bss = ws.cell(r, 2).value
    if bss and str(bss).startswith('BSS'):
        try:
            ws.cell(r, 3).value = "OOO"
            masked_count += 1
        except: pass
print(f"01_인적정보 이름 마스킹: {masked_count}명")

# 15_직원별책무매핑 이름 마스킹
if '15_직원별책무매핑' in wb.sheetnames:
    ws = wb['15_직원별책무매핑']
    for r in range(4, 80):
        cell = ws.cell(r, 3)
        if isinstance(cell, MergedCell): continue
        v = cell.value
        if v and isinstance(v, str) and 'BSS' not in v:
            if '(' in v:
                grade_part = v[v.index('('):]
                cell.value = f"OOO {grade_part}"
            elif v not in ['No.','사번','성명','부서','메인 책무 수'] and len(v) <= 5:
                if any('가' <= ch <= '힣' for ch in v):
                    cell.value = "OOO"
    print("15_직원별책무매핑 이름 마스킹")

# ============ 2. 가상 평가 등급 (중소기업 현실적 분포) ============
N = 34

# 직원별 "잠재 평균 등급" (B=3 중심 정규분포)
# 중소기업 현실: B 중심, S/D 극단 적음, 일부 핵심 인재 + 일부 저성과자
employee_base = []
for i in range(N):
    # 정규분포: 평균 3.0 (B), 표준편차 0.75
    base = random.gauss(3.0, 0.75)
    # 약간의 "스타급" 또는 "저성과자" 추가 (꼬리 강조)
    rand_tail = random.random()
    if rand_tail < 0.05:        # 5% 스타급
        base = random.uniform(4.3, 4.8)
    elif rand_tail < 0.10:      # 5% 저성과자
        base = random.uniform(1.2, 1.8)
    employee_base.append(max(1.0, min(5.0, base)))

def score_to_grade(s):
    if s >= 4.5: return 'S'
    if s >= 3.5: return 'A'
    if s >= 2.5: return 'B'
    if s >= 1.5: return 'C'
    return 'D'

def gen_grade(base, noise=0.3):
    s = base + random.gauss(0, noise)
    return score_to_grade(max(1.0, min(5.0, s)))

# 상·하반기 트렌드 (개선/유지/하락)
employee_trend = []
for _ in range(N):
    r = random.random()
    if r < 0.25:      employee_trend.append(0.3)   # 25% 개선
    elif r < 0.70:    employee_trend.append(0.0)   # 45% 유지
    elif r < 0.90:    employee_trend.append(-0.2)  # 20% 약간 하락
    else:             employee_trend.append(-0.5)  # 10% 큰 하락

# 영역별 편차 (역할/공통/직무 약간 다른 베이스)
employee_area_var = []
for _ in range(N):
    employee_area_var.append({
        'role':   random.gauss(0, 0.25),
        'common': random.gauss(0, 0.25),
        'job':    random.gauss(0, 0.25),
    })

# 1차/2차 평가자별 노이즈 (조정자는 약간 다르게 평가)
def gen_pair(base, eval_noise=0.25):
    """1차/2차 평가자 등급 쌍 - 약간의 차이"""
    g1 = gen_grade(base, eval_noise)
    g2 = gen_grade(base + random.gauss(0, 0.2), eval_noise)
    return g1, g2

# 07a 상반기, 07b 하반기 채우기
# 07a/07b 컬럼: G(역할 1차), H(역할 2차), I(공통 1차), J(공통 2차), K(직무 1차), L(직무 2차)
for half_idx, sn in enumerate(['07a_상반기입력', '07b_하반기입력']):
    ws = wb[sn]
    for i in range(N):
        r = 6 + i
        base = employee_base[i]
        trend = employee_trend[i] if half_idx == 1 else 0  # 하반기에만 트렌드 적용
        av = employee_area_var[i]

        # 역할평가 1·2차
        role_g1, role_g2 = gen_pair(base + trend + av['role'])
        ws.cell(r, 7).value  = role_g1
        ws.cell(r, 8).value  = role_g2

        # 공통역량 1·2차
        com_g1, com_g2 = gen_pair(base + trend + av['common'])
        ws.cell(r, 9).value  = com_g1
        ws.cell(r, 10).value = com_g2

        # 직무역량 1·2차
        job_g1, job_g2 = gen_pair(base + trend + av['job'])
        ws.cell(r, 11).value = job_g1
        ws.cell(r, 12).value = job_g2

print(f"07a/07b 가상 평가 등급 입력: {N}명 × 6개 셀 × 2반기 = {N*12}개")

# ============ 3. LibreOffice 변환 (계산값 확정) ============
wb.save(TMP)

out_dir = '/tmp/sim_final_out'
shutil.rmtree(out_dir, ignore_errors=True)
os.makedirs(out_dir)
subprocess.run(['libreoffice', '--headless', '--calc', '--convert-to', 'xlsx',
                '--outdir', out_dir, TMP], capture_output=True, timeout=180)
conv = glob.glob(out_dir + "/*.xlsx")[0]

# ============ 4. 시뮬레이션 결과 저장 ============
OUT = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가/베이시스소프트_HR_시뮬레이션결과.xlsx"
try:
    if os.path.exists(OUT): os.remove(OUT)
    shutil.copy(conv, OUT)
    print(f"\n저장: {OUT}")
except PermissionError:
    OUT = OUT.replace('.xlsx', '_v2.xlsx')
    shutil.copy(conv, OUT)
    print(f"\n잠금 - 저장: {OUT}")

# ============ 5. 검증 (분포 + 샘플) ============
wbc = load_workbook(conv, data_only=True)
from collections import Counter

# 08_종합평가 최종 등급 분포
ws = wbc['08_종합평가']
final_grades = []
abs_grades = []
emps = []
for r in range(6, 56):
    bss = ws.cell(r, 2).value
    if not bss: continue
    s = ws.cell(r, 19).value
    t = ws.cell(r, 20).value
    u = ws.cell(r, 21).value
    if u: final_grades.append(u)
    if t: abs_grades.append(t)
    emps.append((bss, s, t, u))

print(f"\n=== 시뮬레이션 결과 검증 ===")
print(f"\n[최종 등급 분포 (상대평가, 강제분포 10/15/50/15/10)]")
c = Counter(final_grades)
for g in ['S','A','B','C','D']:
    cnt = c.get(g, 0)
    pct = cnt / len(final_grades) * 100 if final_grades else 0
    bar = '█' * int(pct/2)
    print(f"  {g}: {cnt:2}명 ({pct:5.1f}%) {bar}")

print(f"\n[참고: 절대등급 분포]")
c2 = Counter(abs_grades)
for g in ['S','A','B','C','D']:
    cnt = c2.get(g, 0)
    pct = cnt / len(abs_grades) * 100 if abs_grades else 0
    print(f"  {g}: {cnt:2}명 ({pct:5.1f}%)")

# 09_보상연계 인상액 합계
ws9 = wbc['09_보상연계']
total_increase = 0
salary_total_before = 0
salary_total_after = 0
for r in range(6, 56):
    bss = ws9.cell(r, 2).value
    if not bss: continue
    base = ws9.cell(r, 7).value
    rate = ws9.cell(r, 9).value
    if isinstance(base, (int, float)) and isinstance(rate, (int, float)):
        increase = base * rate
        total_increase += increase
        salary_total_before += base
        salary_total_after += base + increase

print(f"\n[보상연계]")
print(f"  현재 연봉 총액: {salary_total_before:,.0f}원")
print(f"  인상 후 연봉 총액: {salary_total_after:,.0f}원")
print(f"  총 인상액: {total_increase:,.0f}원")
print(f"  평균 인상률: {total_increase/salary_total_before*100:.2f}%")

# 12_시뮬레이션 대시보드
ws12 = wbc['12_시뮬레이션']
print(f"\n[12_시뮬레이션 대시보드]")
for r in range(4, 12):
    a = ws12.cell(r, 1).value
    b = ws12.cell(r, 2).value
    if a:
        print(f"  {a}: {b}")
