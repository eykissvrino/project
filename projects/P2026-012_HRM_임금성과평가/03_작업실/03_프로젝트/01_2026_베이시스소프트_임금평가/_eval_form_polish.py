"""평가지 양식 일괄 정비:
1. 00_작성안내 "■ 문의" 섹션 제거
2. STEP 7 문구 변경 (HR 담당자 할 일 → 평가자 입장으로)
3. 모든 시트 열 너비 통일
"""
import os, glob, shutil
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

THIN = Side(style='thin', color='BFBFBF')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 표준 통일 열 너비 (전 시트 동일 적용)
STD_WIDTH = {
    'A': 7,    # No
    'B': 14,   # 영역/군
    'C': 24,   # 항목/역량/책무
    'D': 38,   # BARS 기준 / 메인 표시
    'E': 38,   # S/A/B/C/D 등급기준
    'F': 8,    # 1차
    'G': 8,    # 2차
    'H': 11,   # 평균(참고) / 등급
    'I': 11,   # 보조
    'J': 11,
    'K': 11,
    'L': 11,
}

# 새 STEP 7 문구 (평가자 입장)
NEW_STEP7_DESC = (
    "작성 완료된 평가지 파일을 HR 담당자에게 이메일로 송부 → "
    "평가 절차 완료 (이후 처리는 HR 담당자가 통합관리시트에서 자동 산출)"
)

FORM_DIR = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가/평가지_양식_최종"

files = sorted(glob.glob(os.path.join(FORM_DIR, "평가지양식_*.xlsx")))
print(f"대상 파일: {len(files)}개")
for f in files:
    print(f"  - {os.path.basename(f)}")

for fp in files:
    name = os.path.basename(fp)
    print(f"\n=== {name} ===")
    wb = load_workbook(fp)

    # ---------- 1) 00_작성안내: 문의 제거 + STEP 7 변경 ----------
    if '00_작성안내' in wb.sheetnames:
        ws = wb['00_작성안내']

        # STEP 7 문구 변경 (행 16 C열)
        step7_cell = ws.cell(16, 3)
        if step7_cell.value and 'STEP' not in str(step7_cell.value) and \
           ('HR' in str(step7_cell.value) or '통합관리' in str(step7_cell.value)):
            step7_cell.value = NEW_STEP7_DESC
            print(f"  ✓ STEP 7 문구 변경 (행 16 C열)")
        else:
            # 더 안전하게: STEP 7가 있는 행 찾아 C열 갱신
            for r in range(1, ws.max_row+1):
                a = ws.cell(r, 1).value
                if a and 'STEP 7' in str(a):
                    ws.cell(r, 3).value = NEW_STEP7_DESC
                    print(f"  ✓ STEP 7 문구 변경 (행 {r} C열)")
                    break

        # "■ 문의" 섹션 행 삭제
        # 보통 행 50: "■ 문의", 행 51: 문의 본문
        # 행 단위로 검색해서 삭제
        rows_to_clear = []
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if a and isinstance(a, str) and ('문의' in a and ('■' in a or '담당자' in a or '컨설턴트' in a)):
                rows_to_clear.append(r)

        # 같은 영역에 추가로 비어있는 행 포함, 단순히 셀 값 None 처리
        for r in rows_to_clear:
            # 머지 해제
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row <= r <= mr.max_row:
                    ws.unmerge_cells(str(mr))
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if not isinstance(cell, MergedCell):
                    cell.value = None
                    cell.fill = PatternFill()
                    cell.border = Border()
                    cell.font = Font(name='맑은 고딕', size=11)
        if rows_to_clear:
            print(f"  ✓ 문의 섹션 제거: 행 {rows_to_clear}")

    # ---------- 2) 모든 시트 열 너비 통일 ----------
    for sn in wb.sheetnames:
        ws = wb[sn]
        if sn == '00_작성안내':
            # 작성안내는 텍스트 위주라 별도 설정
            ws.column_dimensions['A'].width = 12   # STEP / 라벨
            ws.column_dimensions['B'].width = 18   # 부제
            ws.column_dimensions['C'].width = 80   # 설명문 (긴 텍스트)
            for col in 'DEFGHIJKL':
                ws.column_dimensions[col].width = 12
        else:
            # 평가지 시트 모두 동일 너비
            for col, w in STD_WIDTH.items():
                ws.column_dimensions[col].width = w

    # ---------- 3) 저장 ----------
    try:
        wb.save(fp)
        print(f"  ✓ 저장 완료: {fp}")
    except PermissionError:
        alt = fp.replace('.xlsx', '_polished.xlsx')
        wb.save(alt)
        print(f"  ! 잠금 - 대체 저장: {alt}")

print("\n=== 전체 정비 완료 ===")

# 검증
print("\n=== 검증: BIM컨설팅 STEP 7 + 열너비 ===")
wb = load_workbook(files[0])
ws = wb['00_작성안내']
for r in range(14, 18):
    a = ws.cell(r,1).value; b = ws.cell(r,2).value; c = ws.cell(r,3).value
    if a or b or c:
        print(f"  행{r}: A={a!r}  B={b!r}  C={str(c)[:80]!r}")

# 문의 영역 확인
print("\n  [행 48~55 - 문의 제거 확인]:")
for r in range(48, 56):
    a = ws.cell(r,1).value
    if a:
        print(f"  행{r}: {a!r}")
    else:
        print(f"  행{r}: (비어있음) ✓")

# 열 너비
print("\n  [열 너비 - G6 시트]:")
ws_g6 = wb[wb.sheetnames[1]]  # 첫 번째 평가지 시트
for col in 'ABCDEFGH':
    w = ws_g6.column_dimensions[col].width
    print(f"    {col}: {w}")
