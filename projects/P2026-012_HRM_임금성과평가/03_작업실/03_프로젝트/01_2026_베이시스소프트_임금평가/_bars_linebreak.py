"""역할평가 D열, 공통역량 E열의 S/A/B/C/D 등급 기준을 5줄 줄바꿈으로 통일
- "S: xxx / A: xxx / B: xxx / C: xxx / D: xxx" → 각 등급별 줄바꿈
- 행 높이도 5줄 들어가게 조정
"""
import os, glob, re
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font

FORM_DIR = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가/평가지_양식_최종"


def beautify_bars(text):
    """슬래시 구분 BARS 텍스트를 5줄 줄바꿈으로 변환.
       지원 패턴: 'S: xxx / A: xxx / B: xxx / C: xxx / D: xxx'
                'S xxx / A xxx / B xxx / C xxx / D xxx'
       2줄 형태("S ... \n A ...")는 이미 처리된 것으로 간주, 그대로 반환.
    """
    if not isinstance(text, str) or not text.strip():
        return text, False

    # 이미 줄바꿈된 형태면 변환 안 함
    if '\n' in text and text.count('\n') >= 4:
        return text, False

    # 슬래시 분할 (등급 라벨로 시작하는 5개 부분이 있는지 확인)
    # 패턴: 슬래시로 5등분, 각 부분이 S/A/B/C/D로 시작
    parts = re.split(r'\s*/\s*', text.strip())
    if len(parts) != 5:
        return text, False

    # 각 부분이 S/A/B/C/D로 시작하는지 확인
    expected = ['S', 'A', 'B', 'C', 'D']
    ok = True
    for i, p in enumerate(parts):
        # 'S:', 'S ', 'S xxx' 등 시작 패턴
        cleaned = p.strip()
        if not cleaned or cleaned[0] != expected[i]:
            ok = False
            break

    if not ok:
        return text, False

    # 각 줄 정리: "S: xxx" → "S  xxx" (콜론 제거 + 일관성)
    cleaned_lines = []
    for p in parts:
        p = p.strip()
        # 'S:' 또는 'S' 뒤의 공백/콜론 정리
        m = re.match(r'^([SABCD])[:\s]+(.+)$', p)
        if m:
            grade, desc = m.group(1), m.group(2).strip()
            cleaned_lines.append(f'{grade}  {desc}')
        else:
            cleaned_lines.append(p)

    return '\n'.join(cleaned_lines), True


def patch_sheet(ws):
    """역할평가 D열, 공통역량 E열의 BARS 줄바꿈 처리"""
    fixed = 0
    # 역할평가 R-1~R-5: 행 16~20, D열 (4)
    for r in range(16, 21):
        a = ws.cell(r, 1).value
        if a and isinstance(a, str) and a.startswith('R-'):
            cell = ws.cell(r, 4)
            if isinstance(cell, MergedCell): continue
            new_val, changed = beautify_bars(cell.value)
            if changed:
                cell.value = new_val
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                           wrap_text=True, indent=1)
                cell.font = Font(name='맑은 고딕', size=9)
                ws.row_dimensions[r].height = 78
                fixed += 1
            else:
                # 이미 줄바꿈된 경우도 정렬과 행 높이 보장
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                           wrap_text=True, indent=1)
                if (ws.row_dimensions[r].height or 0) < 70:
                    ws.row_dimensions[r].height = 78

    # 공통역량 C-1~C-11: 행 26~36, E열 (5)
    for r in range(25, 38):
        a = ws.cell(r, 1).value
        if a and isinstance(a, str) and a.startswith('C-'):
            cell = ws.cell(r, 5)
            if isinstance(cell, MergedCell): continue
            new_val, changed = beautify_bars(cell.value)
            if changed:
                cell.value = new_val
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                           wrap_text=True, indent=1)
                cell.font = Font(name='맑은 고딕', size=9)
                ws.row_dimensions[r].height = 78
                fixed += 1
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                           wrap_text=True, indent=1)
                if (ws.row_dimensions[r].height or 0) < 70:
                    ws.row_dimensions[r].height = 78

    return fixed


files = sorted(glob.glob(os.path.join(FORM_DIR, "평가지양식_*.xlsx")))
print(f"대상 파일: {len(files)}개\n")

for fp in files:
    name = os.path.basename(fp)
    wb = load_workbook(fp)
    total = 0
    for sn in wb.sheetnames:
        if sn == '00_작성안내': continue
        ws = wb[sn]
        total += patch_sheet(ws)
    wb.save(fp)
    print(f"  ✓ {name}: BARS {total}개 셀 줄바꿈 변환")

# 검증
print("\n=== 검증: 경영전략기획 G5-M ===")
wb = load_workbook([f for f in files if '경영전략기획' in f][0])
ws = wb['양식_경영전략기획_G5-M']

print("\n[① 역할평가 R-1 D열 - 5줄 줄바꿈]:")
d16 = ws['D16'].value
if d16:
    for line in str(d16).split('\n'):
        print(f"  {line}")
print(f"  행 16 높이: {ws.row_dimensions[16].height}")

print("\n[② 공통역량 C-1 E열 - 5줄 줄바꿈]:")
e26 = ws['E26'].value
if e26:
    for line in str(e26).split('\n'):
        print(f"  {line}")
print(f"  행 26 높이: {ws.row_dimensions[26].height}")

print("\n[③ 직무역량 J-1 D열 - 5줄 줄바꿈 (이미 처리됨)]:")
d42 = ws['D42'].value
if d42:
    for line in str(d42).split('\n'):
        print(f"  {line}")
