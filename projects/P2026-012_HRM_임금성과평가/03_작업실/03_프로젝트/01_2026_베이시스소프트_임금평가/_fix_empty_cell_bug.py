"""빈 셀 참조 0 평가 버그 수정
원인: =IF('01'!B="","", '07a'!G) 형태에서 07a!G가 빈 셀이면 0으로 평가
      → 점수 수식의 F<>"" 조건을 우회 → 점수 0 → 등급 D 산출

해결:
- 04/05/06의 1·2차 등급 셀: 7a/7b 참조 셀이 빈 셀이면 "" 반환
- 04/05/06의 점수 수식: F/G가 0 또는 ""이면 빈 셀
- 04/05/06의 등급 수식: 점수가 0 또는 빈 셀이면 빈 셀
- 08의 G/H/J/K/M/N: 04/05/06 등급 참조 시 빈 셀 처리
- 08의 I/L/O 가중평균 수식: 점수 0 또는 빈 셀이면 빈 셀
- 08의 U(최종 등급): 종합 점수 0 또는 빈 셀이면 빈 셀
"""
import os, shutil
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

BASE = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가"
SRC = os.path.join(BASE, "베이시스소프트_HR통합관리시트_FINAL.xlsx")
TMP = "/tmp/master_fix.xlsx"

# 원본 파일 손상 여부 점검 후 복사
import zipfile
try:
    zipfile.ZipFile(SRC).close()
    shutil.copy(SRC, TMP)
except Exception:
    # 업로드 파일 사용
    UP = "/sessions/compassionate-bold-ride/mnt/uploads/베이시스소프트_HR통합관리시트_FINAL.xlsx"
    shutil.copy(UP, TMP)
    print(f"원본 손상 - 업로드 파일 사용")

wb = load_workbook(TMP)

DATA_START = 6
EMP_MAX = 50
DATA_END = DATA_START + EMP_MAX - 1   # 55
HALF_OFFSET = 58                       # 상반기 행 기준 +58 = 하반기 행

W1 = "'02_의사결정·종합결과'!$B$67"
W2 = "'02_의사결정·종합결과'!$B$68"
W1H = "'02_의사결정·종합결과'!$B$7"
W2H = "'02_의사결정·종합결과'!$B$8"

def g2s(c):
    return f'IFERROR(CHOOSE(FIND({c},"DCBAS"),1,2,3,4,5),0)'

def score_formula(c1, c2, w1, w2):
    """1차/2차 등급 → 가중평균 점수. LEN>0 체크로 빈 텍스트("") 우회."""
    return (
        f'=IFERROR('
        f'IF(AND(LEN({c1})>0,LEN({c2})>0),{w1}*{g2s(c1)}+{w2}*{g2s(c2)},'
        f'IF(LEN({c1})>0,{g2s(c1)},'
        f'IF(LEN({c2})>0,{g2s(c2)},""))),"")'
    )

def grade_formula(score_cell):
    """점수 → 등급. 점수가 빈 셀 또는 0.5 미만이면 ""."""
    return (
        f'=IF(OR({score_cell}="",NOT(ISNUMBER({score_cell})),{score_cell}<0.5),"",'
        f'IF({score_cell}>=4.5,"S",'
        f'IF({score_cell}>=3.5,"A",'
        f'IF({score_cell}>=2.5,"B",'
        f'IF({score_cell}>=1.5,"C","D")))))'
    )

def ref_cell(src_sheet, src_col, ref):
    """다른 시트 셀 참조 - 빈 셀이면 "" 반환"""
    return (
        f'=IF(OR(\'01_인적정보\'!B{ref}="",{src_sheet}!{src_col}{ref}=""),"",'
        f'{src_sheet}!{src_col}{ref})'
    )


# ============ 04/05/06 1·2차 등급 셀 + 점수 + 확정등급 수식 갱신 ============
def patch_eval_sheet(ws, c_grade1, c_grade2, c_score, c_final,
                     src1, src2, src_sheet_half1, src_sheet_half2):
    """평가 시트의 1·2차 + 점수 + 등급 수식 갱신 (상/하반기 모두)"""
    fixed = 0
    for half_idx, src_sheet in enumerate([src_sheet_half1, src_sheet_half2]):
        for i in range(EMP_MAX):
            ref = DATA_START + i           # 07a/07b 참조 행 (항상 6~55)
            r = DATA_START + i + (half_idx * HALF_OFFSET)  # 상/하반기 행

            # 1차/2차 셀: 07a/07b 참조 시 빈 셀이면 ""
            cell = ws[f'{c_grade1}{r}']
            if not isinstance(cell, MergedCell):
                cell.value = ref_cell(src_sheet, src1, ref)
                fixed += 1
            cell = ws[f'{c_grade2}{r}']
            if not isinstance(cell, MergedCell):
                cell.value = ref_cell(src_sheet, src2, ref)
                fixed += 1

            # 점수 셀
            cell = ws[f'{c_score}{r}']
            if not isinstance(cell, MergedCell):
                cell.value = score_formula(f'{c_grade1}{r}', f'{c_grade2}{r}', W1, W2)
                fixed += 1

            # 확정 등급 셀
            cell = ws[f'{c_final}{r}']
            if not isinstance(cell, MergedCell):
                cell.value = grade_formula(f'{c_score}{r}')
                fixed += 1
    return fixed

# 04_역할평가: F=1차, G=2차, H=점수, I=등급
n = patch_eval_sheet(wb['04_역할평가'],
                     'F', 'G', 'H', 'I',
                     'G', 'H',  # 07a/07b의 G(역할 1차), H(역할 2차)
                     "'07a_상반기입력'", "'07b_하반기입력'")
print(f"04_역할평가: {n}개 셀 수정")

# 05_공통역량평가: F=1차, G=2차, H=점수, I=등급
n = patch_eval_sheet(wb['05_공통역량평가'],
                     'F', 'G', 'H', 'I',
                     'I', 'J',  # 07a/07b의 I(공통 1차), J(공통 2차)
                     "'07a_상반기입력'", "'07b_하반기입력'")
print(f"05_공통역량평가: {n}개 셀 수정")

# 06_직무역량평가: G=1차, H=2차, I=점수, J=등급
n = patch_eval_sheet(wb['06_직무역량평가'],
                     'G', 'H', 'I', 'J',
                     'K', 'L',  # 07a/07b의 K(직무 1차), L(직무 2차)
                     "'07a_상반기입력'", "'07b_하반기입력'")
print(f"06_직무역량평가: {n}개 셀 수정")

# ============ 08_종합평가 수식 갱신 ============
ws = wb['08_종합평가']
fixed8 = 0
for i in range(EMP_MAX):
    r = DATA_START + i
    role_h = r + HALF_OFFSET   # 하반기 행 (04/05/06의 하반기 위치)

    # G: 역할 상반기 (04!I{r})
    ws[f'G{r}'].value = ref_cell("'04_역할평가'", 'I', r)
    # H: 역할 하반기 (04!I{role_h})
    ws[f'H{r}'].value = (
        f'=IF(OR(\'01_인적정보\'!B{r}="",\'04_역할평가\'!I{role_h}=""),"",'
        f'\'04_역할평가\'!I{role_h})'
    )
    # I: 역할 가중평균 점수
    ws[f'I{r}'].value = score_formula(f'G{r}', f'H{r}', W1H, W2H)

    # J/K: 공통 상/하반기
    ws[f'J{r}'].value = ref_cell("'05_공통역량평가'", 'I', r)
    ws[f'K{r}'].value = (
        f'=IF(OR(\'01_인적정보\'!B{r}="",\'05_공통역량평가\'!I{role_h}=""),"",'
        f'\'05_공통역량평가\'!I{role_h})'
    )
    ws[f'L{r}'].value = score_formula(f'J{r}', f'K{r}', W1H, W2H)

    # M/N: 직무 상/하반기 (06의 등급 셀은 J열)
    ws[f'M{r}'].value = ref_cell("'06_직무역량평가'", 'J', r)
    ws[f'N{r}'].value = (
        f'=IF(OR(\'01_인적정보\'!B{r}="",\'06_직무역량평가\'!J{role_h}=""),"",'
        f'\'06_직무역량평가\'!J{role_h})'
    )
    ws[f'O{r}'].value = score_formula(f'M{r}', f'N{r}', W1H, W2H)

    # S: 종합 점수 (3축 가중평균) - 빈 셀 또는 0 이면 ""
    ws[f'S{r}'].value = (
        f'=IF(OR(NOT(ISNUMBER(I{r})),NOT(ISNUMBER(L{r})),NOT(ISNUMBER(O{r}))),"",'
        f'I{r}*P{r}+L{r}*Q{r}+O{r}*R{r})'
    )

    # T: 절대등급
    ws[f'T{r}'].value = grade_formula(f'S{r}')

    # U: 최종등급 (절대 / 상대)
    sr = f"$S${DATA_START}:$S${DATA_END}"
    # 상대평가: 점수가 0보다 큰 행만 카운트 + RANK로 백분위
    pct = (
        f'IFERROR(RANK(S{r},{sr},0)/'
        f'MAX(COUNTIF({sr},">0"),1),0)'
    )
    rel = (
        f'IF({pct}<=\'02_의사결정·종합결과\'!$B$11,"S",'
        f'IF({pct}<=\'02_의사결정·종합결과\'!$B$11+\'02_의사결정·종합결과\'!$C$11,"A",'
        f'IF({pct}<=\'02_의사결정·종합결과\'!$B$11+\'02_의사결정·종합결과\'!$C$11+\'02_의사결정·종합결과\'!$D$11,"B",'
        f'IF({pct}<=\'02_의사결정·종합결과\'!$B$11+\'02_의사결정·종합결과\'!$C$11+\'02_의사결정·종합결과\'!$D$11+\'02_의사결정·종합결과\'!$E$11,"C","D"))))'
    )
    ws[f'U{r}'].value = (
        f'=IF(OR(S{r}="",NOT(ISNUMBER(S{r})),S{r}<0.5),"",'
        f'IF(\'02_의사결정·종합결과\'!$B$6="절대평가",T{r},{rel}))'
    )
    fixed8 += 1

print(f"08_종합평가: {fixed8}명 종합 수식 수정")

# 저장
wb.save(TMP)
OUT = os.path.join(BASE, "베이시스소프트_HR통합관리시트_FINAL.xlsx")
try:
    if os.path.exists(OUT): os.remove(OUT)
    shutil.copy(TMP, OUT)
    print(f"\n저장: {OUT}")
except PermissionError:
    OUT = OUT.replace('.xlsx', '_FIX.xlsx')
    shutil.copy(TMP, OUT)
    print(f"\n잠금 - 저장: {OUT}")

# 검증: LibreOffice로 변환해 계산값 확인
import subprocess, glob
out_dir = '/tmp/master_fix_check'
shutil.rmtree(out_dir, ignore_errors=True)
os.makedirs(out_dir)
subprocess.run(['libreoffice','--headless','--calc','--convert-to','xlsx',
                '--outdir', out_dir, TMP], capture_output=True, timeout=120)
conv = glob.glob(out_dir + "/*.xlsx")[0]
wbc = load_workbook(conv, data_only=True)

print("\n=== 검증: 빈 입력 상태 (07a/07b 모두 비어있음) ===")
ws = wbc['08_종합평가']
print(f"{'사번':10} {'역할점수':10} {'공통점수':10} {'직무점수':10} {'종합점수':10} {'절대':6} {'최종':6}")
for r in range(6, 41):
    b = ws.cell(r, 2).value
    if not b: continue
    i = ws.cell(r, 9).value
    l = ws.cell(r, 12).value
    o = ws.cell(r, 15).value
    s = ws.cell(r, 19).value
    t = ws.cell(r, 20).value
    u = ws.cell(r, 21).value
    print(f"{str(b):10} {str(i):10} {str(l):10} {str(o):10} {str(s):10} {str(t):6} {str(u):6}")
(u):6}")
