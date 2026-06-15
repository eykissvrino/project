"""평가지 외부 참조 제거
- 직무역량 책무 평균 H열 수식: 외부 참조 제거, 단순 1·2차 평균 ((F+G)/2)
- 직무역량 종합 행 수식도 점검 (외부 참조 없는지 확인)
- 평가지 평균(참고)은 단순 참고용. 통합관리시트가 평가자 가중치 적용한 가중평균 산출.
"""
import os, glob, re
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Font, Alignment

FORM_DIR = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가/평가지_양식_최종"

def g2s(c):
    return f'IFERROR(IF({c}="S",5,IF({c}="A",4,IF({c}="B",3,IF({c}="C",2,IF({c}="D",1,0))))),0)'

def new_duty_avg(r):
    """책무 평균 수식 (외부 참조 없음, 단순 1·2차 평균)"""
    return (
        f'=IF(E{r}="O",IFERROR('
        f'IF(AND(F{r}<>"",G{r}<>""),({g2s(f"F{r}")}+{g2s(f"G{r}")})/2,'
        f'IF(F{r}<>"",{g2s(f"F{r}")},IF(G{r}<>"",{g2s(f"G{r}")},""))),"")'
        f',"")'
    )

def has_external_ref(value):
    """수식에 외부 통합 문서 참조 포함 여부 ('[숫자]...! 또는 'C:\파일!' 등)"""
    if not isinstance(value, str): return False
    return bool(re.search(r"'?\[\d+\]", value)) or ".xls" in value.lower()

files = sorted(glob.glob(os.path.join(FORM_DIR, "평가지양식_*.xlsx")))
print(f"대상 파일: {len(files)}개\n")

for fp in files:
    name = os.path.basename(fp)
    wb = load_workbook(fp)
    total_fixed = 0
    other_ext = 0

    for sn in wb.sheetnames:
        if sn == '00_작성안내': continue
        ws = wb[sn]

        # 1) 책무 행 H열 평균 수식 갱신
        for r in range(42, 80):
            a = ws.cell(r, 1).value
            if a and isinstance(a, str) and a.startswith('J-'):
                # 책무 행이면 H열 평균 수식 새 버전으로
                h_cell = ws.cell(r, 8)
                if not isinstance(h_cell, MergedCell):
                    h_cell.value = new_duty_avg(r)
                    total_fixed += 1

        # 2) 다른 시트 모든 셀에서 외부 참조 잔존 확인
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell): continue
                if has_external_ref(cell.value):
                    other_ext += 1

    wb.save(fp)
    print(f"  ✓ {name}: 책무 평균 {total_fixed}건 갱신, 잔존 외부 참조 {other_ext}건")

# 검증
print("\n=== 검증: BIM G6 H42 수식 (외부 참조 없는지) ===")
wb = load_workbook(files[0])
ws = wb['양식_BIM컨설팅_G6']
print(f"  H42: {ws['H42'].value[:200]}")
print(f"  외부 참조 ([숫자]) 포함: {'[' in ws['H42'].value and ']' in ws['H42'].value.split(chr(39))[0] if isinstance(ws['H42'].value, str) else 'N/A'}")

# 전체 외부 참조 다시 점검
print("\n=== 모든 파일의 잔존 외부 참조 ===")
for fp in files:
    name = os.path.basename(fp)
    wb = load_workbook(fp)
    ext_count = 0
    sample = None
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell): continue
                v = cell.value
                if isinstance(v, str) and re.search(r"'?\[\d+\]", v):
                    ext_count += 1
                    if sample is None: sample = (sn, cell.coordinate, str(v)[:80])
    if ext_count > 0:
        print(f"  ✗ {name}: 외부 참조 {ext_count}건 잔존 - 샘플: {sample}")
    else:
        print(f"  ✓ {name}: 외부 참조 0건")

# 경영전략기획 추가 검증
print("\n=== 경영전략기획 - 모든 등급 시트 직무역량 영역 ===")
wb = load_workbook([f for f in files if '경영전략기획' in f][0])
for sn in wb.sheetnames:
    if sn == '00_작성안내': continue
    ws = wb[sn]
    duty_count = 0
    for r in range(42, 80):
        v = ws.cell(r, 1).value
        if v and isinstance(v, str) and v.startswith('J-'):
            duty_count += 1
    h42 = ws['H42'].value if ws['H42'].value else 'None'
    has_ext = '[' in str(h42) and bool(re.search(r"\[\d+\]", str(h42)))
    print(f"  {sn}: 책무 {duty_count}개, H42 외부참조={has_ext}")
