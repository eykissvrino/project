"""직무역량 평가에 S/A/B/C/D 등급 기준 박스 추가
- 컬럼 헤더 행 위에 1행 삽입 → BARS 기준 박스
- 이후 모든 수식의 행 번호를 +1 처리 (시트 참조 제외)
"""
import os, glob, re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell

THIN = Side(style='thin', color='BFBFBF')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FORM_DIR = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가/평가지_양식_최종"

BARS_TEXT = (
    '★ 직무역량 S/A/B/C/D 등급 기준  ━  모든 책무 공통 적용  |  '
    'S(탁월) 책무 기대 현저 초과  |  '
    'A(우수) 기대 상회  |  '
    'B(양호) 기대 부합  |  '
    'C(미흡) 일부 미달  |  '
    'D(부족) 전반 미흡'
)

def shift_formula(formula, threshold, shift):
    """수식 내 '같은 시트' 셀 참조의 행 번호가 threshold 이상이면 shift만큼 증가.
       다른 시트 참조('Sheet'!cell)는 placeholder로 보호.
    """
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula

    # 다른 시트 참조 (예: '02_의사결정'!$B$67) 보호
    sheet_ref_pattern = r"('[^']+'!\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)"
    placeholders = []
    def stash(m):
        placeholders.append(m.group(0))
        return f"@@PH{len(placeholders)-1}@@"
    safe = re.sub(sheet_ref_pattern, stash, formula)

    # 일반 셀 참조 행번호 +shift
    def cell_shift(m):
        col_part = m.group(1)
        row = int(m.group(2))
        if row >= threshold:
            return f'{col_part}{row + shift}'
        return m.group(0)
    safe = re.sub(r'(\$?[A-Z]{1,3}\$?)(\d+)', cell_shift, safe)

    # placeholder 복원
    for i, ph in enumerate(placeholders):
        safe = safe.replace(f"@@PH{i}@@", ph)
    return safe


def patch_sheet(ws):
    """직무역량 헤더(③) 행 다음(컬럼 헤더 행 위)에 BARS 기준 박스 행 1개 삽입.
       이후 모든 수식의 행번호 갱신.
    """
    # 직무역량 헤더 위치 자동 감지
    target = None
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, 1)
        if isinstance(cell, MergedCell): continue
        v = cell.value
        if v and isinstance(v, str) and '③' in v and '직무역량' in v:
            target = r
            break
    if target is None:
        return False, '직무역량 헤더 못 찾음'

    # 이미 BARS 기준이 들어있는지 체크 (재실행 안전)
    next_cell_val = ws.cell(target + 1, 1).value
    if isinstance(next_cell_val, str) and 'S/A/B/C/D 등급 기준' in next_cell_val and '직무역량' in next_cell_val:
        return False, '이미 BARS 박스 있음 (스킵)'

    insert_row = target + 1   # 컬럼 헤더 행 위에 삽입

    # 1행 삽입 (openpyxl이 셀/스타일/머지/DV/조건부서식 자동 이동)
    ws.insert_rows(idx=insert_row, amount=1)

    # 모든 셀 수식 행번호 +1 (insert_row 이상)
    for row in ws.iter_rows(min_row=insert_row + 1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell, MergedCell): continue
            v = cell.value
            if isinstance(v, str) and v.startswith('='):
                cell.value = shift_formula(v, threshold=insert_row, shift=1)

    # 새 BARS 박스 행 작성
    ws.merge_cells(start_row=insert_row, end_row=insert_row,
                   start_column=1, end_column=8)
    c = ws.cell(insert_row, 1)
    c.value = BARS_TEXT
    c.fill = PatternFill('solid', fgColor='FFEB9C')
    c.font = Font(name='맑은 고딕', size=10, bold=True, color='C00000')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    c.border = BORDER_ALL
    ws.row_dimensions[insert_row].height = 38

    return True, f'삽입 행 {insert_row}'


files = sorted(glob.glob(os.path.join(FORM_DIR, "평가지양식_*.xlsx")))
print(f"대상 파일: {len(files)}개\n")

for fp in files:
    name = os.path.basename(fp)
    wb = load_workbook(fp)
    ok = 0; skip = 0
    for sn in wb.sheetnames:
        if sn == '00_작성안내': continue
        ws = wb[sn]
        success, msg = patch_sheet(ws)
        if success: ok += 1
        else: skip += 1
    wb.save(fp)
    print(f"  ✓ {name}: 추가 {ok}개 / 스킵 {skip}개")

# 검증
print("\n=== 검증: BIM컨설팅 G6 ===")
wb = load_workbook(files[0])
ws = wb['양식_BIM컨설팅_G6']
# 직무역량 영역 행 40~60
for r in range(40, 62):
    a = ws.cell(r,1).value
    g = ws.cell(r,7).value
    h = ws.cell(r,8).value
    if a or g or h:
        a_str = str(a)[:70] if a else ''
        g_str = str(g)[:35] if g else ''
        h_str = str(h)[:35] if h else ''
        print(f"  행{r}: A={a_str}  G={g_str}  H={h_str}")

# 통합 입력 행 수식 검증
print("\n[통합 입력 행 - 직무역량 종합 셀 참조 확인]:")
for r in range(60, 70):
    a = ws.cell(r,1).value
    e = ws.cell(r,5).value
    f = ws.cell(r,6).value
    if a or e or f:
        print(f"  행{r}: A={str(a)[:30]}  E={str(e)[:30]}  F={str(f)[:30]}")

print("\n=== 완료 ===")
