"""최종본 기반 시뮬레이션"""
import os, shutil, subprocess, glob, random
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from collections import Counter

random.seed(2026)

SRC = "/sessions/compassionate-bold-ride/mnt/uploads/베이시스소프트_HR통합관리시트_FINAL-5e0b9760.xlsx"
TMP = "/tmp/sim_2026.xlsx"
if os.path.exists(TMP):
    os.chmod(TMP, 0o644)
shutil.copy(SRC, TMP)
os.chmod(TMP, 0o644)

wb = load_workbook(TMP)

# 이름 마스킹
ws = wb['01_인적정보']
for r in range(6, 56):
    bss = ws.cell(r, 2).value
    if bss and str(bss).startswith('BSS'):
        try: ws.cell(r, 3).value = "OOO"
        except: pass

if '15_직원별책무매핑' in wb.sheetnames:
    ws = wb['15_직원별책무매핑']
    for r in range(4, 80):
        cell = ws.cell(r, 3)
        if isinstance(cell, MergedCell): continue
        v = cell.value
        if v and isinstance(v, str) and 'BSS' not in v:
            if '(' in v:
                grade_part = v[v.index('('):]
                cell.value = f"OOO {grade_part}"
            elif v not in ['No.','사번','성명','부서','메인 책무 수'] and len(v) <= 5:
                if any('가' <= ch <= '힣' for ch in v):
                    cell.value = "OOO"

# 가상 평가
N = 34
employee_base = []
for i in range(N):
    base = random.gauss(3.0, 0.75)
    r = random.random()
    if r < 0.05:    base = random.uniform(4.3, 4.8)
    elif r < 0.10:  base = random.uniform(1.2, 1.8)
    employee_base.append(max(1.0, min(5.0, base)))

def s2g(s):
    if s >= 4.5: return 'S'
    if s >= 3.5: return 'A'
    if s >= 2.5: return 'B'
    if s >= 1.5: return 'C'
    return 'D'

def gen(base, noise=0.3):
    return s2g(max(1.0, min(5.0, base + random.gauss(0, noise))))

trends = []
for _ in range(N):
    r = random.random()
    trends.append(0.3 if r < 0.25 else 0.0 if r < 0.70 else -0.2 if r < 0.90 else -0.5)

area_var = [{'role': random.gauss(0,0.25), 'common': random.gauss(0,0.25), 'job': random.gauss(0,0.25)} for _ in range(N)]

for half_idx, sn in enumerate(['07a_상반기입력', '07b_하반기입력']):
    ws = wb[sn]
    for i in range(N):
        r = 6 + i
        base = employee_base[i]
        trend = trends[i] if half_idx == 1 else 0
        av = area_var[i]
        ws.cell(r, 7).value  = gen(base + trend + av['role'])
        ws.cell(r, 8).value  = gen(base + trend + av['role'] + random.gauss(0, 0.2))
        ws.cell(r, 9).value  = gen(base + trend + av['common'])
        ws.cell(r, 10).value = gen(base + trend + av['common'] + random.gauss(0, 0.2))
        ws.cell(r, 11).value = gen(base + trend + av['job'])
        ws.cell(r, 12).value = gen(base + trend + av['job'] + random.gauss(0, 0.2))

print(f"가상 평가 입력 완료: {N}명")

wb.save(TMP)

# LO 변환
od = '/tmp/sim_2026_out'
shutil.rmtree(od, ignore_errors=True); os.makedirs(od)
subprocess.run(['libreoffice','--headless','--calc','--convert-to','xlsx',
                '--outdir',od,TMP], capture_output=True, timeout=180)
conv = glob.glob(od + "/*.xlsx")[0]

OUT = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가/베이시스소프트_HR_시뮬레이션결과.xlsx"
try:
    if os.path.exists(OUT): os.remove(OUT)
    shutil.copy(conv, OUT)
    print(f"저장: {OUT}")
except PermissionError:
    OUT = OUT.replace('.xlsx', '_NEW.xlsx')
    shutil.copy(conv, OUT)
    print(f"잠금-저장: {OUT}")

# 검증
wbc = load_workbook(conv, data_only=True)
ws = wbc['08_종합평가']
final, abs_gr = [], []
results = []
for r in range(6, 56):
    bss = ws.cell(r, 2).value
    if not bss: continue
    s = ws.cell(r, 19).value
    t = ws.cell(r, 20).value
    u = ws.cell(r, 21).value
    if u: final.append(u)
    if t: abs_gr.append(t)
    results.append((bss, s, t, u))

print(f"\n[최종 등급 분포 - 상대평가 강제분포 10/15/50/15/10]")
c = Counter(final)
for g in ['S','A','B','C','D']:
    cnt = c.get(g, 0)
    pct = cnt/len(final)*100 if final else 0
    print(f"  {g}: {cnt:2}명 ({pct:5.1f}%)  {'#'*int(pct/2)}")

print(f"\n[참고: 절대등급 분포]")
c2 = Counter(abs_gr)
for g in ['S','A','B','C','D']:
    cnt = c2.get(g, 0)
    pct = cnt/len(abs_gr)*100 if abs_gr else 0
    print(f"  {g}: {cnt:2}명 ({pct:5.1f}%)")

# 보상연계
ws9 = wbc['09_보상연계']
total_inc = 0; total_before = 0
for r in range(6, 56):
    if not ws9.cell(r, 2).value: continue
    base = ws9.cell(r, 7).value
    rate = ws9.cell(r, 9).value
    if isinstance(base,(int,float)) and isinstance(rate,(int,float)):
        total_inc += base * rate
        total_before += base
print(f"\n[보상연계]")
print(f"  현재 연봉총액: {total_before:,.0f}원")
print(f"  인상 후 연봉총액: {total_before+total_inc:,.0f}원")
print(f"  총 인상액: {total_inc:,.0f}원")
print(f"  평균 인상률: {total_inc/total_before*100:.2f}%")

# 12_시뮬레이션
ws12 = wbc['12_시뮬레이션']
print(f"\n[12_시뮬레이션 대시보드]")
for r in range(4, 22):
    a = ws12.cell(r, 1).value
    b = ws12.cell(r, 2).value
    cc = ws12.cell(r, 3).value
    if a:
        line = f"  {a}"
        if b is not None: line += f" : {b}"
        if cc is not None: line += f" ({cc})"
        print(line)

# 샘플 5명
print(f"\n[샘플 5명]")
print(f"  {'사번':10} {'종합점수':10} {'절대':6} {'최종':6}")
for bss, s, t, u in results[:10]:
    print(f"  {bss:10} {str(s):10} {str(t):6} {str(u):6}")
