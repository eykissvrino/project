"""직무역량 컬럼 위치 통일 (1차=E, 2차=F, 평균=G, 메인=H)
- 업로드된 평가지 7개를 평가지_양식_최종/에 적용
- 직무역량 영역만 셀 이동 (책무 데이터 보존)
- 종합 행 + 통합 입력 행 셀 참조 일관성 유지
"""
import os, glob, shutil
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE = "/sessions/compassionate-bold-ride/mnt/P2026-012_HRM_임금성과평가/03_프로젝트/01_2026_베이시스소프트_임금평가"
UPLOAD_DIR = "/sessions/compassionate-bold-ride/mnt/uploads"
FINAL_DIR = os.path.join(BASE, "평가지_양식_최종")

THIN = Side(style='thin', color='BFBFBF')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLOR = {
    'header_dark':'1F4E78', 'header_mid':'2E74B5', 'header_light':'B4C7E7',
    'input_yellow':'FFF2CC', 'auto_green':'E2EFDA', 'info_gray':'F2F2F2',
    'bars_yellow':'FFF9E5',
}

BARS_JOB_TEXT = (
    "S  기대 현저히 초과\n"
    "A  기대 상회\n"
    "B  기대 부합\n"
    "C  부분 미흡\n"
    "D  전반 미흡"
)

def g2s(c):
    return f'IFERROR(IF({c}="S",5,IF({c}="A",4,IF({c}="B",3,IF({c}="C",2,IF({c}="D",1,0))))),0)'

def grade_from_score(sc):
    return (f'=IF(OR({sc}="",{sc}<0.5),"",'
            f'IF({sc}>=4.5,"S",IF({sc}>=3.5,"A",IF({sc}>=2.5,"B",'
            f'IF({sc}>=1.5,"C","D")))))')

def sc(ws, coord, value=None, fill_color=None, bold=False, size=11, color='000000',
       h='left', v='center', wrap=True, border=BORDER_ALL, indent=0, number_fmt=None):
    c = ws[coord]
    if isinstance(c, MergedCell): return
    if value is not None: c.value = value
    c.font = Font(name='맑은 고딕', size=size, bold=bold, color=color)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)
    if fill_color: c.fill = PatternFill('solid', fgColor=fill_color)
    if border: c.border = border
    if number_fmt: c.number_format = number_fmt

def clear_rows(ws, start, end, max_col=12):
    for mr in list(ws.merged_cells.ranges):
        if start <= mr.min_row <= end or start <= mr.max_row <= end:
            ws.unmerge_cells(str(mr))
    for r in range(start, end + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell): continue
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font(name='맑은 고딕', size=11)
            cell.alignment = Alignment()
            cell.border = Border()
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None

def extract_duties(ws):
    """직무역량 영역의 책무 데이터 추출 - 메인 표시도 보존"""
    duties = []
    for r in range(42, 80):
        a = ws.cell(r, 1).value
        if a and isinstance(a, str) and a.startswith('J-'):
            no = a
            area = ws.cell(r, 2).value or ''
            duty = ws.cell(r, 3).value or ''
            # 현재 구조 (D=BARS, E=메인, F=1차, G=2차, H=평균)에서 메인은 E열
            main = ws.cell(r, 5).value
            duties.append((no, area, duty, main))
        elif duties and (not a or '평가자' in str(a) or '직무 종합' in str(a)):
            break
    return duties


def rebuild_job_area(ws, duties):
    """직무역량 영역 + 통합 입력 + 종합 소견 재빌드 (새 컬럼: E=1차, F=2차, G=평균, H=메인)"""
    n = len(duties)
    DUTY_START = 42
    duty_end = DUTY_START + n - 1
    SUM1 = duty_end + 2
    SUM2 = duty_end + 3
    HUB_HDR = SUM2 + 2
    HUB_USAGE = HUB_HDR + 1
    HUB_COL = HUB_USAGE + 1
    HUB_DATA = HUB_COL + 1
    OPN_HDR = HUB_DATA + 2
    OPN_START = OPN_HDR + 1

    clear_rows(ws, 40, OPN_START + 5)

    # 행 40: 섹션 헤더
    ws.merge_cells('A40:H40')
    sc(ws, 'A40',
       '▶ ③ 직무역량평가 — 메인 책무(H열 O 표시)만 1·2차 평가자 등급 입력',
       fill_color=COLOR['header_dark'], color='FFFFFF', bold=True, size=12, indent=1)
    ws.row_dimensions[40].height = 28

    # 행 41: 컬럼 헤더 (새 구조 - 역할평가와 1차/2차 위치 통일)
    headers = ['No.', '전문영역', '핵심 책무', 'S/A/B/C/D 등급 기준',
               '1차', '2차', '평균(참고)', '메인']
    for i, h in enumerate(headers):
        sc(ws, f'{chr(ord("A")+i)}41', h,
           fill_color=COLOR['header_dark'], color='FFFFFF', bold=True,
           size=10, h='center', wrap=True)
    ws.row_dimensions[41].height = 32

    # 책무 행
    for i, (no, area, duty, main) in enumerate(duties):
        r = DUTY_START + i
        sc(ws, f'A{r}', no, fill_color=COLOR['info_gray'], bold=True, h='center', size=10)
        sc(ws, f'B{r}', area, fill_color=COLOR['info_gray'], size=10, indent=1)
        sc(ws, f'C{r}', duty, fill_color=COLOR['info_gray'], size=10, indent=1)
        sc(ws, f'D{r}', BARS_JOB_TEXT, fill_color=COLOR['bars_yellow'],
           size=9, indent=1, h='left')
        # E: 1차, F: 2차 (입력 셀)
        sc(ws, f'E{r}', None, fill_color=COLOR['input_yellow'], bold=True, h='center', size=11)
        sc(ws, f'F{r}', None, fill_color=COLOR['input_yellow'], bold=True, h='center', size=11)
        # G: 평균(참고) - 메인일 때만 산출 (외부 참조 없는 단순 평균)
        avg = (
            f'=IF(H{r}="O",IFERROR('
            f'IF(AND(LEN(E{r})>0,LEN(F{r})>0),({g2s(f"E{r}")}+{g2s(f"F{r}")})/2,'
            f'IF(LEN(E{r})>0,{g2s(f"E{r}")},IF(LEN(F{r})>0,{g2s(f"F{r}")},""))),"")'
            f',"")'
        )
        sc(ws, f'G{r}', avg, fill_color=COLOR['auto_green'],
           h='center', size=10, number_fmt='0.00')
        # H: 메인 (O/X) - 기존 값 있으면 복원
        sc(ws, f'H{r}', main, fill_color=COLOR['input_yellow'],
           bold=True, h='center', size=11)
        ws.row_dimensions[r].height = 85

    # 종합 행 1차 - A:D 머지 라벨, E=점수, F=화살표, G=등급 (역할평가와 동일 구조)
    r = SUM1
    ws.merge_cells(f'A{r}:D{r}')
    sc(ws, f'A{r}', '1차 평가자 직무 종합 (메인 책무 평균)',
       fill_color=COLOR['header_light'], bold=True, size=10, indent=1)
    # E: 점수 (메인=O인 책무의 1차 평가 평균)
    score1 = (
        f'=IFERROR(SUMPRODUCT((H{DUTY_START}:H{duty_end}="O")*'
        f'IFERROR(IF(E{DUTY_START}:E{duty_end}="S",5,'
        f'IF(E{DUTY_START}:E{duty_end}="A",4,'
        f'IF(E{DUTY_START}:E{duty_end}="B",3,'
        f'IF(E{DUTY_START}:E{duty_end}="C",2,'
        f'IF(E{DUTY_START}:E{duty_end}="D",1,0))))),0))/'
        f'MAX(SUMPRODUCT((H{DUTY_START}:H{duty_end}="O")*(LEN(E{DUTY_START}:E{duty_end})>0)),1),0)'
    )
    sc(ws, f'E{r}', score1, fill_color=COLOR['auto_green'],
       bold=True, h='center', size=11, number_fmt='0.00')
    sc(ws, f'F{r}', '→ 등급', fill_color=COLOR['header_light'],
       bold=True, h='center', size=10)
    sc(ws, f'G{r}', grade_from_score(f'E{r}'), fill_color=COLOR['auto_green'],
       bold=True, h='center', size=14)
    sc(ws, f'H{r}', None, fill_color=COLOR['info_gray'])
    ws.row_dimensions[r].height = 28

    # 종합 행 2차
    r = SUM2
    ws.merge_cells(f'A{r}:D{r}')
    sc(ws, f'A{r}', '2차 평가자 직무 종합 (메인 책무 평균)',
       fill_color=COLOR['header_light'], bold=True, size=10, indent=1)
    score2 = (
        f'=IFERROR(SUMPRODUCT((H{DUTY_START}:H{duty_end}="O")*'
        f'IFERROR(IF(F{DUTY_START}:F{duty_end}="S",5,'
        f'IF(F{DUTY_START}:F{duty_end}="A",4,'
        f'IF(F{DUTY_START}:F{duty_end}="B",3,'
        f'IF(F{DUTY_START}:F{duty_end}="C",2,'
        f'IF(F{DUTY_START}:F{duty_end}="D",1,0))))),0))/'
        f'MAX(SUMPRODUCT((H{DUTY_START}:H{duty_end}="O")*(LEN(F{DUTY_START}:F{duty_end})>0)),1),0)'
    )
    sc(ws, f'E{r}', score2, fill_color=COLOR['auto_green'],
       bold=True, h='center', size=11, number_fmt='0.00')
    sc(ws, f'F{r}', '→ 등급', fill_color=COLOR['header_light'],
       bold=True, h='center', size=10)
    sc(ws, f'G{r}', grade_from_score(f'E{r}'), fill_color=COLOR['auto_green'],
       bold=True, h='center', size=14)
    sc(ws, f'H{r}', None, fill_color=COLOR['info_gray'])
    ws.row_dimensions[r].height = 28

    # 통합 입력 영역
    r = HUB_HDR
    ws.merge_cells(f'A{r}:H{r}')
    sc(ws, f'A{r}',
       '▶ 통합관리시트 입력 행 (아래 한 줄 A~G 7개 셀을 통째 복사 → 07a/07b의 G~M에 [값 붙여넣기])',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=11, indent=1)
    ws.row_dimensions[r].height = 24

    r = HUB_USAGE
    ws.merge_cells(f'A{r}:H{r}')
    sc(ws, f'A{r}',
       '사용법: 아래 한 줄 A~G 선택 → Ctrl+C → 통합관리시트 07a 또는 07b의 해당 사원 행 G열에 [선택하여 붙여넣기 → 값] 실행',
       fill_color=COLOR['info_gray'], size=9, indent=1)
    ws.row_dimensions[r].height = 22

    r = HUB_COL
    cols = ['역할(1차)', '역할(2차)', '공통(1차)', '공통(2차)', '직무(1차)', '직무(2차)', '소견']
    for i, h in enumerate(cols):
        sc(ws, f'{chr(ord("A")+i)}{r}', h,
           fill_color=COLOR['header_light'], bold=True, h='center', size=10)
    ws.row_dimensions[r].height = 24

    r = HUB_DATA
    refs = ['=G21', '=G22', '=H37', '=H38', f'=G{SUM1}', f'=G{SUM2}', '']
    for i, ref in enumerate(refs):
        col = chr(ord('A') + i)
        if ref:
            sc(ws, f'{col}{r}', ref, fill_color=COLOR['auto_green'],
               bold=True, h='center', size=12)
        else:
            sc(ws, f'{col}{r}', None, fill_color=COLOR['input_yellow'], size=10)
    ws.row_dimensions[r].height = 30

    # 종합 소견
    r = OPN_HDR
    ws.merge_cells(f'A{r}:H{r}')
    sc(ws, f'A{r}', '▶ 종합 소견 (강점·개선·육성 방향)',
       fill_color=COLOR['header_mid'], color='FFFFFF', bold=True, size=11, indent=1)
    ws.row_dimensions[r].height = 24

    for i, label in enumerate(['강점', '개선', '육성 방향']):
        r = OPN_START + i
        sc(ws, f'A{r}', label, fill_color=COLOR['header_light'],
           bold=True, h='center', size=10)
        ws.merge_cells(f'B{r}:H{r}')
        sc(ws, f'B{r}', None, fill_color=COLOR['input_yellow'], size=10, indent=1)
        ws.row_dimensions[r].height = 50


# ============ 실행 ============
# 업로드된 7개 파일을 평가지_양식_최종/으로 복사
upload_files = sorted(glob.glob(os.path.join(UPLOAD_DIR, "평가지양식_*.xlsx")))
print(f"업로드 파일: {len(upload_files)}개")
for src in upload_files:
    name = os.path.basename(src)
    dst = os.path.join(FINAL_DIR, name)
    try:
        if os.path.exists(dst): os.chmod(dst, 0o644); os.remove(dst)
        shutil.copy(src, dst)
        os.chmod(dst, 0o644)
        print(f"  복사: {name}")
    except Exception as e:
        print(f"  !! 실패: {e}")

# 각 파일의 직무역량 영역 재빌드
files = sorted(glob.glob(os.path.join(FINAL_DIR, "평가지양식_*.xlsx")))
print(f"\n재빌드 대상: {len(files)}개\n")

for fp in files:
    name = os.path.basename(fp)
    wb = load_workbook(fp)
    for sn in wb.sheetnames:
        if sn == '00_작성안내': continue
        ws = wb[sn]
        duties = extract_duties(ws)
        rebuild_job_area(ws, duties)
    wb.save(fp)
    print(f"  ✓ {name}")

# 검증
print("\n=== 검증: BIM G6 직무역량 영역 ===")
wb = load_workbook(files[0])
ws = wb['양식_BIM컨설팅_G6']
print(f"\n[행 41 컬럼 헤더 A~H]:")
print(f"  {[ws.cell(41,c).value for c in range(1,9)]}")
print(f"\n[행 42 J-1 책무 데이터 A~H]:")
print(f"  {[str(ws.cell(42,c).value)[:25] if ws.cell(42,c).value else '' for c in range(1,9)]}")
print(f"\n[행 58 종합 1차 A~H]:")
print(f"  {[str(ws.cell(58,c).value)[:40] if ws.cell(58,c).value else '' for c in range(1,9)]}")
print(f"\n[행 64 통합 입력 데이터]:")
print(f"  {[str(ws.cell(64,c).value)[:15] if ws.cell(64,c).value else '' for c in range(1,9)]}")
