# -*- coding: utf-8 -*-
"""KTR AI 역량 스킬셋 v4 → Excel 생성 스크립트"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── 색상 정의 ──
HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="맑은 고딕", size=12, bold=True)
BODY_FONT = Font(name="맑은 고딕", size=9)
BODY_FONT_BOLD = Font(name="맑은 고딕", size=9, bold=True)

FILL_R1 = PatternFill("solid", fgColor="4472C4")  # 파랑
FILL_R2 = PatternFill("solid", fgColor="548235")  # 초록
FILL_R3 = PatternFill("solid", fgColor="BF8F00")  # 주황
FILL_R4 = PatternFill("solid", fgColor="7030A0")  # 보라

FILL_R1_LIGHT = PatternFill("solid", fgColor="D6E4F0")
FILL_R2_LIGHT = PatternFill("solid", fgColor="E2EFDA")
FILL_R3_LIGHT = PatternFill("solid", fgColor="FFF2CC")
FILL_R4_LIGHT = PatternFill("solid", fgColor="E8D5F5")

FILL_HEADER = PatternFill("solid", fgColor="2F5496")
FILL_SUBHEADER = PatternFill("solid", fgColor="D6E4F0")
FILL_LV1 = PatternFill("solid", fgColor="FFF2CC")
FILL_LV2 = PatternFill("solid", fgColor="D6E4F0")
FILL_LV3 = PatternFill("solid", fgColor="F2DCDB")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_GREY = PatternFill("solid", fgColor="F2F2F2")

FILL_BP = PatternFill("solid", fgColor="548235")  # 바이오플라스틱 (초록)
FILL_MD = PatternFill("solid", fgColor="2E75B6")  # 의료기기 (파랑)
FILL_CM = PatternFill("solid", fgColor="C55A11")  # 화장품 (주황)
FILL_FC = PatternFill("solid", fgColor="7030A0")  # 정밀화학 (보라)

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def fill_for_area(code):
    if code.startswith("R1"): return FILL_R1_LIGHT
    if code.startswith("R2"): return FILL_R2_LIGHT
    if code.startswith("R3"): return FILL_R3_LIGHT
    if code.startswith("R4"): return FILL_R4_LIGHT
    return FILL_WHITE


def fill_for_level(lv):
    if lv == "Lv1": return FILL_LV1
    if lv == "Lv2": return FILL_LV2
    if lv == "Lv3": return FILL_LV3
    return FILL_WHITE


def style_header(ws, row, cols, fill=FILL_HEADER):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_row(ws, row, cols, font=BODY_FONT, fill=None, align=WRAP):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.alignment = align
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════
# Sheet 1: 역량체계 총괄
# ══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "①역량체계 총괄"

headers1 = ["역량영역", "영역코드", "영역 정의 (실무 행동 중심)", "역량항목 코드", "역량항목", '"AI로 뭘 하는가"', "이론적 근거"]
set_col_widths(ws1, [18, 8, 55, 10, 25, 55, 40])

ws1.append(headers1)
style_header(ws1, 1, len(headers1))

items1 = [
    ["AI 업무 기초 활용", "R1", "생성형 AI 등 범용 AI 도구로 문서 작성, 정보 검색, 업무 커뮤니케이션의 생산성을 향상시키는 능력",
     "R1-1", "생성형 AI 업무 문서 작성", "보고서·기획서·이메일 초안을 AI로 작성하고 검토·보완", "OECD AI Usage + DigComp 2.2"],
    ["", "", "", "R1-2", "AI 활용 정보 검색·요약", "산업 동향·규제·기술 트렌드를 AI로 검색·요약", "WEF Digital Literacy"],
    ["", "", "", "R1-3", "AI 생성 결과물 검증·보완", "AI 출력물의 오류·편향을 식별하고 전문적으로 수정", "OECD Metacognition & Critical Thinking"],
    ["AI 데이터 실무", "R2", "업무 현장 데이터를 수집·정리하고, AI 분석 도구를 활용하여 데이터 기반 의사결정과 업무 개선을 수행하는 능력",
     "R2-1", "업무 데이터 표준화 수집·정리", "현장 데이터를 AI 분석 가능한 형태로 수집·정리", "NCS 데이터분석 + OECD Data"],
    ["", "", "", "R2-2", "AI 도구 활용 데이터 분석", "Excel·BI·AI 분석 도구로 패턴·이상·트렌드 분석", "DigComp 2.2 + OECD Data"],
    ["", "", "", "R2-3", "데이터 시각화·대시보드 구축", "분석 결과를 차트·대시보드로 시각화하여 의사결정 활용", "DigComp 2.2 + NCS"],
    ["AI 직무 도구 운용", "R3", "산업별·직무별 전문 AI 솔루션을 실제 업무에 적용하여 직무 성과를 개선하는 능력",
     "R3-1", "직무 특화 AI 솔루션 운용", "공정AI, 품질AI, 규제AI 등 전문 도구 운용", "WEF Automation Collaboration"],
    ["", "", "", "R3-2", "AI 분석 결과 해석·업무 적용", "AI 산출 결과를 해석하고 업무 판단에 반영", "OECD AI Capability Indicators"],
    ["", "", "", "R3-3", "AI 도구 선정·평가", "업무 목적에 맞는 AI 도구 비교·평가·PoC 실행", "WEF Empowering Frontlines"],
    ["AI 업무 혁신 설계", "R4", "AI를 활용하여 업무 프로세스를 자동화·재설계하고, AI 도입 프로젝트를 기획·실행하는 능력",
     "R4-1", "AI 기반 업무 자동화 구현", "반복 업무를 AI-RPA로 자동화, 워크플로우 설계", "WEF Continuous Improvement"],
    ["", "", "", "R4-2", "AI 예측·최적화 모델 요건 정의", "도메인 전문가로서 AI 모델 입출력·품질 기준 정의·검증", "PwC 제조AI 로드맵"],
    ["", "", "", "R4-3", "AI 도입 프로젝트 기획·관리", "ROI 분석, 프로젝트 계획 수립, 변화관리 실행", "DigComp 2.2 Safety + OECD"],
]

area_fills = {"R1": FILL_R1_LIGHT, "R2": FILL_R2_LIGHT, "R3": FILL_R3_LIGHT, "R4": FILL_R4_LIGHT}
current_area = "R1"
for i, row in enumerate(items1, 2):
    ws1.append(row)
    code = row[3][:2]
    if row[0]:
        current_area = row[1]
    style_row(ws1, i, len(headers1), fill=area_fills.get(current_area, FILL_WHITE))
    if row[0]:
        for c in range(1, 4):
            ws1.cell(row=i, column=c).font = BODY_FONT_BOLD

# 수준 체계 추가 (빈 행 후)
r = ws1.max_row + 2
ws1.cell(row=r, column=1, value="역량 수준 체계").font = TITLE_FONT
r += 1
level_headers = ["수준", "명칭", "산업전환 연동", "현장 기준", "행동 기준"]
for c, h in enumerate(level_headers, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, len(level_headers))

levels = [
    ["Lv1", "활용 기초", "선제대응(14.3%)", "생성형 AI로 문서 작성하고, 내 데이터를 정리하여 AI에 넣을 수 있다", "작성한다, 정리한다, 활용해 본다"],
    ["Lv2", "실무 적용", "전환(81.4%)", "직무 특화 AI 도구를 일상 업무에 적용하고, AI 결과로 업무를 개선한다", "운용한다, 분석한다, 개선한다"],
    ["Lv3", "혁신 설계", "정착(4.3%)", "AI 기반 프로세스를 설계하고, AI 도입 프로젝트를 기획·주도한다", "설계한다, 기획한다, 주도한다"],
]
for lv in levels:
    r += 1
    ws1.append(lv)
    style_row(ws1, r, len(level_headers), fill=fill_for_level(lv[0]))


# ══════════════════════════════════════════════════════════
# Sheet 2: 공통 스킬
# ══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("②공통 스킬")
headers2 = ["스킬 ID", "역량영역", "역량항목", "스킬명", "수준", "전문지식 (배경)", "행동 수준 (이걸 할 수 있다)"]
set_col_widths(ws2, [12, 20, 25, 35, 8, 55, 60])

ws2.append(headers2)
style_header(ws2, 1, len(headers2))

common_skills = [
    # R1-1
    ["R1-1-S1", "R1 AI 업무 기초 활용", "R1-1 생성형 AI 업무 문서 작성", "AI 활용 보고서·기획서 초안 작성", "Lv1",
     "프롬프트 설계 기법(역할 부여, 맥락 제공, 단계적 사고, 출력 형식 지정, Few-shot)",
     "업무 보고서·기획서의 초안을 생성형 AI로 작성하고, 전문적 검토를 거쳐 부서 제출 수준으로 완성할 수 있다"],
    ["R1-1-S2", "", "", "AI 활용 이메일·공문 작성", "Lv1",
     "비즈니스 문서 유형별 톤앤매너, 생성형 AI에서의 문체 지정 방법",
     "상황(보고·요청·회신·사과 등)에 맞는 이메일·공문을 AI로 작성하고, 적절성을 검토하여 발송할 수 있다"],
    ["R1-1-S3", "", "", "AI 활용 기술 문서·매뉴얼 초안 작성", "Lv2",
     "기술 문서 구조(SOP, 작업표준서, 시험성적서), 전문용어 정확성 검증",
     "기술 문서·매뉴얼의 구조와 초안을 AI로 생성하고, 기술적 정확성을 검증하여 최종본을 완성할 수 있다"],
    # R1-2
    ["R1-2-S1", "", "R1-2 AI 활용 정보 검색·요약", "AI 활용 산업 동향·규제 변화 조사", "Lv1",
     "검색 프롬프트 최적화, 정보 출처 교차 검증 방법, 바이오화학 주요 정보원",
     "바이오화학 산업 동향·규제 변화·기술 트렌드를 AI로 검색·요약하여 주간/월간 보고에 반영할 수 있다"],
    ["R1-2-S2", "", "", "AI 활용 경쟁사·시장 정보 수집", "Lv1",
     "경쟁 정보 수집 체계, AI 뉴스 모니터링 도구 활용",
     "경쟁사 동향·신제품·특허 정보를 AI로 수집·정리하여 전략 회의 자료를 준비할 수 있다"],
    ["R1-2-S3", "", "", "AI 활용 기술 문헌·논문 검색·요약", "Lv2",
     "학술 DB 검색 전략, AI 논문 요약 도구(Semantic Scholar, Elicit 등)",
     "연구 주제 관련 최신 논문·특허를 AI로 검색하고, 핵심 내용을 요약하여 R&D 방향 수립에 활용할 수 있다"],
    ["R1-2-S4", "", "", "AI 활용 번역·다국어 문서 작성", "Lv1",
     "산업 전문 용어, 번역 프롬프트 최적화, 번역 품질 검증",
     "기술 문서·인증 서류를 AI로 번역하고, 전문 용어의 정확성을 검토하여 최종본을 완성할 수 있다"],
    # R1-3
    ["R1-3-S1", "", "R1-3 AI 생성 결과물 검증·보완", "AI 생성물 팩트체크·오류 보정", "Lv1",
     "환각(hallucination)·편향·최신성 부족 등 생성형 AI의 한계 유형",
     "AI가 작성한 문서·분석의 사실 오류를 식별하고, 신뢰 가능한 출처로 확인하여 수정·보완할 수 있다"],
    ["R1-3-S2", "", "", "AI 출력의 업무 적합성 평가", "Lv2",
     "업무 맥락에 맞는 품질 기준, 전문 용어 정확성, 규정 준수 여부 판단",
     "AI 출력물이 업무 기준(정확성·전문성·규정 준수)에 부합하는지 평가하고, 부적합 시 개선하여 활용할 수 있다"],
    # R2-1
    ["R2-1-S1", "R2 AI 데이터 실무", "R2-1 업무 데이터 표준화 수집·정리", "업무 데이터 표준 양식 설계·수집", "Lv1",
     "데이터 명명규칙, 단위 통일, 필수항목 정의, 수집 주기 설정",
     "현장 업무 데이터(공정·품질·시험 등)를 표준화된 양식으로 수집·기록하여 AI 분석이 가능한 품질로 관리할 수 있다"],
    ["R2-1-S2", "", "", "데이터 이상치·결측치 처리", "Lv2",
     "이상치 탐지(IQR, Z-score), 결측치 처리(삭제, 대체, 보간)",
     "업무 데이터에서 이상치와 결측치를 식별하고, 적절한 방법으로 처리하여 AI 분석에 적합한 품질을 확보할 수 있다"],
    ["R2-1-S3", "", "", "AI 학습용 데이터 라벨링·가공", "Lv2",
     "라벨링 기준 설정, 피처 엔지니어링, 데이터 정규화/표준화",
     "도메인 전문가로서 AI 학습용 데이터의 라벨링 기준을 정의하고, 데이터를 가공하여 모델 학습에 적합한 형태로 준비할 수 있다"],
    # R2-2
    ["R2-2-S1", "", "R2-2 AI 도구 활용 데이터 분석", "Excel·BI 도구 활용 데이터 분석", "Lv1",
     "Excel 피벗·함수, 기초 통계(평균·분산·상관), Power BI/Tableau 기초",
     "업무 데이터를 Excel 고급 기능이나 BI 도구로 분석하여 트렌드·패턴·이상을 파악하고 보고할 수 있다"],
    ["R2-2-S2", "", "", "AI 분석 도구 활용 패턴 탐색", "Lv2",
     "AutoML 기초, 군집분석, 회귀분석, 상관분석 개념",
     "AI 분석 도구(AutoML 등)에 업무 데이터를 입력하여 숨겨진 패턴과 상관관계를 탐색하고, 업무 인사이트를 도출할 수 있다"],
    ["R2-2-S3", "", "", "다변량 데이터 AI 분석", "Lv2",
     "PCA(주성분분석), 다변량 회귀, 변수 간 상호작용 분석",
     "여러 변수가 얽힌 복잡한 업무 데이터를 AI로 다변량 분석하여, 핵심 영향 변수를 식별할 수 있다"],
    # R2-3
    ["R2-3-S1", "", "R2-3 데이터 시각화·대시보드 구축", "업무 KPI 대시보드 구축·운영", "Lv2",
     "대시보드 설계 원칙, KPI 정의, BI 도구 활용",
     "핵심 업무 지표(생산량, 불량률, 공정 상태 등)를 대시보드로 구축하여 실시간 모니터링할 수 있다"],
    ["R2-3-S2", "", "", "데이터 기반 업무 개선 제안", "Lv2",
     "데이터 스토리텔링, 근거 기반 의사결정, 시각화 자료 작성",
     "데이터 분석 결과를 시각화 자료로 정리하여, 구체적 수치와 근거를 포함한 업무 개선안을 제안할 수 있다"],
    # R3-1
    ["R3-1-S1", "R3 AI 직무 도구 운용", "R3-1 직무 특화 AI 솔루션 운용", "공정 최적화 AI 도구 운용", "Lv2",
     "공정 파라미터-품질 관계, AI 최적화(Bayesian Opt., GA 등) 입출력 구조",
     "목표 품질을 설정하고 공정 변수를 AI 도구에 입력하여 최적 운전 조건을 도출하고 현장에 적용할 수 있다"],
    ["R3-1-S2", "", "", "품질 예측 AI 시스템 운용", "Lv2",
     "예측 품질(Predictive Quality) 모델, 이상탐지 알고리즘, 알람 유형별 대응 프로토콜",
     "AI 품질 예측 시스템의 알람을 판독하고, 이상 예측 시 원인을 파악하여 사전 조치를 실행할 수 있다"],
    ["R3-1-S3", "", "", "머신비전 AI 검사 시스템 운용", "Lv2",
     "이미지 인식 기반 불량 분류(스크래치, 기포, 변색, 치수이탈), 오탐/미탐 판단",
     "머신비전 외관 검사 AI의 판정 결과를 검토하고, 오탐·미탐을 식별하여 판정 기준을 개선할 수 있다"],
    ["R3-1-S4", "", "", "규제·인증 문서 AI 분석 도구 활용", "Lv2",
     "규제 문서 자동 크롤링, NLP 기반 요구사항 추출, 변경사항 자동 추적",
     "AI 도구로 규제 문서를 자동 수집·분석하여 핵심 요구사항을 추출하고 대응 체크리스트를 생성할 수 있다"],
    ["R3-1-S5", "", "", "AI 기반 실시간 공정 모니터링 운용", "Lv2",
     "실시간 데이터 스트리밍, 통계적 공정관리+ML 이상탐지, 알람 체계",
     "AI 실시간 모니터링 시스템의 알람을 판독하고, 알람 유형(경고/위험/정지)에 따른 대응을 실행할 수 있다"],
    ["R3-1-S6", "", "", "AI 기반 수요 예측·생산계획 수립", "Lv3",
     "AI 수요 예측 모델, APS(Advanced Planning & Scheduling), 제약조건 설정",
     "AI 수요 예측 결과와 현장 제약조건을 종합하여, 가동률·재고·납기를 최적화하는 생산계획을 수립할 수 있다"],
    # R3-2
    ["R3-2-S1", "", "R3-2 AI 분석 결과 해석·업무 적용", "AI 예측 결과의 현장 적합성 판단", "Lv2",
     "AI 모델 성능 지표(정확도, 재현율, F1), 도메인 지식과의 교차 검증",
     "AI가 산출한 예측·분류 결과를 현장 경험과 대조하여 적합성을 판단하고, 부적합 시 원인을 피드백할 수 있다"],
    ["R3-2-S2", "", "", "AI 결과 기반 업무 의사결정", "Lv2",
     "데이터 기반 의사결정 프레임워크, 예측 불확실성 해석",
     "AI 분석 결과와 도메인 전문성을 결합하여 업무 판단을 내리고, 그 근거를 문서화할 수 있다"],
    # R3-3
    ["R3-3-S1", "", "R3-3 AI 도구 선정·평가", "업무 목적별 AI 도구 비교·선정", "Lv2",
     "AI 솔루션 평가 기준(정확도, 비용, 확장성, 사용성, 벤더 신뢰도), PoC 설계",
     "업무 목적에 맞는 AI 도구 후보를 비교·평가하고, PoC를 설계·실행하여 도입 타당성을 검증할 수 있다"],
    # R4-1
    ["R4-1-S1", "R4 AI 업무 혁신 설계", "R4-1 AI 기반 업무 자동화 구현", "반복 업무 AI 자동화 워크플로우 설계", "Lv3",
     "RPA+AI 연계, 워크플로우 도구(Power Automate, Zapier 등), 자동화 ROI",
     "현재 반복·수동 업무를 분석하여 AI 자동화 대상을 선정하고, 자동화 워크플로우를 설계·구현할 수 있다"],
    ["R4-1-S2", "", "", "AI-OCR 문서 처리 자동화 구축", "Lv2",
     "OCR+AI 문서 인식, 데이터 추출·입력 자동화, 양식 표준화",
     "종이 문서·PDF를 AI-OCR로 자동 인식하여 데이터를 추출하고, 시스템에 자동 입력하는 프로세스를 구축할 수 있다"],
    ["R4-1-S3", "", "", "AI 기반 보고서 자동 생성 체계 구축", "Lv3",
     "데이터 파이프라인, 자동 리포트 생성, 템플릿 설계",
     "업무 데이터가 자동으로 수집되어 AI가 정기 보고서를 생성하는 체계를 구축할 수 있다"],
    # R4-2
    ["R4-2-S1", "", "R4-2 AI 예측·최적화 모델 요건 정의", "AI 모델 요건 정의서 작성", "Lv3",
     "AI 프로젝트 라이프사이클(문제정의→데이터→모델링→배포→모니터링)",
     "도메인 전문가로서 AI 모델의 입력 데이터·출력 형태·성공 기준·제약조건을 정의한 요건 정의서를 작성할 수 있다"],
    ["R4-2-S2", "", "", "AI 모델 결과 현장 실증·피드백", "Lv3",
     "모델 성능 평가, A/B 테스트, 현장 적용 시 변수, 피드백 루프 설계",
     "AI 모델의 예측 결과를 현장에서 실증 검증하고, 정량적 개선 사항을 데이터 사이언티스트에게 피드백할 수 있다"],
    # R4-3
    ["R4-3-S1", "", "R4-3 AI 도입 프로젝트 기획·관리", "AI 도입 ROI 분석·타당성 보고서 작성", "Lv3",
     "AI 도입 비용 구조(솔루션, 데이터 준비, 인력 교육, 유지보수), ROI 산출",
     "AI 도입 시 예상 비용과 기대 효과를 산출하여 도입 타당성 보고서를 작성하고 경영진에 제안할 수 있다"],
    ["R4-3-S2", "", "", "AI 도입 변화관리 계획 수립·실행", "Lv3",
     "변화관리 방법론, 저항 요인 분석, 이해관계자 소통 전략",
     "AI 도입 시 조직 내 저항 요인을 분석하고, 부서별 교육·소통 계획을 수립하여 변화관리를 실행할 수 있다"],
    ["R4-3-S3", "", "", "AI 거버넌스 기본 체계 수립", "Lv3",
     "데이터 보안, AI 윤리(편향, 프라이버시), 산업별 AI 규제(SaMD, REACH 등)",
     "AI 활용 시 데이터 보안·윤리 기준을 정의하고, 산업별 규제 준수를 위한 기본 거버넌스 체계를 수립할 수 있다"],
]

for i, row in enumerate(common_skills, 2):
    ws2.append(row)
    fill = fill_for_area(row[0])
    style_row(ws2, i, len(headers2), fill=fill)
    # 수준 셀 색상
    ws2.cell(row=i, column=5).fill = fill_for_level(row[4])
    ws2.cell(row=i, column=5).alignment = CENTER


# ══════════════════════════════════════════════════════════
# Sheet 3: 바이오플라스틱
# ══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("③바이오플라스틱")
headers3 = ["스킬 ID", "직무", "스킬명", "수준", "전문지식 (배경)", "행동 수준 (이걸 할 수 있다)"]
set_col_widths(ws3, [12, 10, 35, 8, 55, 60])

ws3.append(headers3)
style_header(ws3, 1, len(headers3), fill=FILL_BP)

bp_skills = [
    # 생산
    ["BP-P-01", "생산", "바이오 소재 공정 파라미터 데이터 수집·관리", "Lv2",
     "PLA/PHA/PBS/Starch blends 제조 공정별 핵심 파라미터(배럴 온도, 사출 압력, 냉각 시간, 스크류 속도)와 품질 결과의 관계",
     "공정별 핵심 파라미터를 식별하고, 품질 결과와 연관시켜 AI 분석에 활용 가능한 수준으로 체계적으로 수집·관리할 수 있다"],
    ["BP-P-02", "생산", "AI 공정 모니터링 시스템 운용·대응", "Lv2",
     "실시간 공정 모니터링 AI의 이상탐지 알고리즘, 알람 유형별 대응 프로토콜",
     "실시간 AI 모니터링 시스템의 알람을 판독하고, 알람 유형(경고/위험/정지)에 따른 대응을 실행할 수 있다"],
    ["BP-P-03", "생산", "AI 기반 소재 혼합비 최적화 도구 활용", "Lv2",
     "소재 혼합(Starch blends, PLA/PHA 블렌딩) 시 물성 변수와 AI 최적화 도구의 입출력 구조",
     "목표 물성(인장강도, 생분해 속도 등)을 설정하고, AI 도구로 최적 배합비를 도출한 후 현장 적용 가능성을 판단할 수 있다"],
    ["BP-P-04", "생산", "AI 연동 스마트팩토리 생산계획 수립", "Lv3",
     "AI 수요 예측과 APS 연동 구조, 바이오플라스틱 특유의 제약조건(소재 유통기한, 습도 민감성)",
     "AI 수요 예측 결과와 바이오플라스틱 고유 제약조건을 종합하여, 가동률·재고·납기를 최적화하는 생산계획을 수립할 수 있다"],
    # 품질
    ["BP-Q-01", "품질", "생분해성·바이오매스 시험 데이터 체계 관리", "Lv2",
     "ISO 14855, ISO 17088, ASTM D6866 시험 항목·데이터 구조·인증 요건",
     "국제 표준 시험 데이터를 체계적으로 기록·관리하고, 인증 신청 및 AI 분석에 즉시 활용 가능한 상태로 유지할 수 있다"],
    ["BP-Q-02", "품질", "머신비전 AI 외관 검사 운용·판정 개선", "Lv2",
     "비전 AI의 불량 유형 분류(스크래치, 기포, 변색, 치수 이탈), 오탐/미탐 판단 기준",
     "머신비전 검사 시스템을 운용하고, AI 판정의 오탐·미탐을 식별하여 판정 기준을 현장에 맞게 개선할 수 있다"],
    ["BP-Q-03", "품질", "인증·규제 문서 AI 작성 보조 활용", "Lv2",
     "OK Compost, TÜV AUSTRIA 등 친환경 인증 문서 요건·구조",
     "인증 문서의 구조와 요건을 이해한 상태에서, 생성형 AI로 초안을 작성하고 전문 검토를 거쳐 최종 문서를 완성할 수 있다"],
    ["BP-Q-04", "품질", "공정-품질 예측 모델 구축 참여(요건 정의·검증)", "Lv3",
     "공정 파라미터-품질 결과 상관관계, 예측 품질 모델 설계 원리",
     "도메인 전문가로서 품질 예측 AI 모델의 핵심 변수를 정의하고, 모델 결과의 현장 적합성을 실증 검증할 수 있다"],
    # R&D
    ["BP-R-01", "R&D", "소재 물성 데이터 AI 다변량 분석", "Lv2",
     "소재 물성(인장강도, 생분해 속도, Tg, Tm, 결정화도) 다변량 분석(PCA, 회귀, 군집)",
     "다수 소재 물성 변수 간 관계를 AI 다변량 분석으로 파악하고, 소재 설계 방향 도출에 활용할 수 있다"],
    ["BP-R-02", "R&D", "AI 기반 배합 설계 최적화 도구 활용", "Lv2",
     "DoE+AI 하이브리드 최적화(Bayesian Optimization, GA), 변수-목적함수 설정",
     "목표 물성을 목적함수로 설정하고 AI 배합 최적화 도구로 효율적으로 최적 배합 조건을 탐색할 수 있다"],
    ["BP-R-03", "R&D", "실험 데이터 AI 학습용 전처리·증강", "Lv2",
     "소규모 실험 데이터 한계, 데이터 증강(augmentation), 전이학습 활용",
     "소규모 실험 데이터를 증강·전처리하여 AI 모델 학습 가능성을 높이고, 전이학습 적용 방안을 검토할 수 있다"],
    ["BP-R-04", "R&D", "AI 소재 물성 예측 모델 활용", "Lv3",
     "Materials Informatics, QSPR(정량적 구조-물성 관계), 역설계(inverse design)",
     "AI 물성 예측 모델에 후보 소재 조건을 입력하여 물성을 사전 예측하고, 실험 대상을 사전 선별할 수 있다"],
    ["BP-R-05", "R&D", "Closed-loop R&D 프로세스 설계", "Lv3",
     "실험→분석→AI 예측→실험 반복 자동화, 능동학습(Active Learning)",
     "AI 예측 결과를 바탕으로 다음 실험을 자동 제안하는 Closed-loop R&D 프로세스를 설계하여 개발 기간을 단축할 수 있다"],
    ["BP-R-06", "R&D", "LCA-AI 통합 친환경 소재 전략 수립", "Lv3",
     "전과정평가(LCA), AI 기반 탄소발자국 시뮬레이션, 환경 영향 최소화 설계",
     "LCA 데이터와 AI 시뮬레이션을 결합하여 탄소발자국을 최소화하는 소재 개발 전략을 수립할 수 있다"],
    # 제품기획
    ["BP-G-01", "제품기획", "친환경 소재 시장 AI 자동 모니터링", "Lv2",
     "AI 기반 시장 모니터링(뉴스 크롤링, 특허 분석, 규제 변화 추적)",
     "AI 도구로 바이오플라스틱 시장 동향·규제 변화·경쟁사 동향을 자동 모니터링하고, 기획에 반영할 인사이트를 도출할 수 있다"],
    ["BP-G-02", "제품기획", "소비자·시장 데이터 AI 분석", "Lv2",
     "소비자 니즈 분석(설문, 리뷰, SNS), 수요 예측, 가격 탄력성",
     "소비자·시장 데이터를 AI로 분석하여 수요를 예측하고, 제품 기획 의사결정에 데이터 기반 근거를 제시할 수 있다"],
    ["BP-G-03", "제품기획", "ESG·탄소규제 시나리오 AI 시뮬레이션", "Lv3",
     "PCR 의무 비율, CBAM, EU PPWR 등 규제와 AI 시나리오 시뮬레이션",
     "주요 규제 시나리오별 제품 포트폴리오의 경제성·규제 적합성을 AI로 시뮬레이션하고, 중장기 전략을 수립할 수 있다"],
]

for i, row in enumerate(bp_skills, 2):
    ws3.append(row)
    style_row(ws3, i, len(headers3))
    ws3.cell(row=i, column=4).fill = fill_for_level(row[3])
    ws3.cell(row=i, column=4).alignment = CENTER
    # 직무별 교차 색상
    if i % 2 == 0:
        for c in [1, 2, 3]:
            ws3.cell(row=i, column=c).fill = FILL_GREY


# ══════════════════════════════════════════════════════════
# Sheet 4: 의료기기
# ══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("④의료기기")
headers4 = ["스킬 ID", "직무", "스킬명", "수준", "행동 수준 (이걸 할 수 있다)"]
set_col_widths(ws4, [12, 10, 40, 8, 65])

ws4.append(headers4)
style_header(ws4, 1, len(headers4), fill=FILL_MD)

md_skills = [
    ["MD-P-01", "생산", "GMP 공정 데이터 ALCOA+ 기준 관리", "Lv2",
     "GMP 데이터 무결성 원칙에 따라 공정 데이터를 수집하여 AI 분석 가능한 품질로 관리할 수 있다"],
    ["MD-P-02", "생산", "AI 기반 클린룸 환경 모니터링 운용", "Lv2",
     "클린룸 AI 모니터링 시스템의 이상 예측 알람에 따른 사전 대응을 실행할 수 있다"],
    ["MD-Q-01", "품질", "비전 AI 외관·치수 자동 검사 운용", "Lv2",
     "의료기기 비전 AI 검사 시스템을 운용하고 판정 기준을 현장에 맞게 조정할 수 있다"],
    ["MD-Q-02", "품질", "인허가 기술문서 AI 작성 보조 활용", "Lv2",
     "FDA 510(k), CE-MDR, MFDS 인허가 서류의 초안을 AI로 작성하고 전문 검토하여 완성할 수 있다"],
    ["MD-Q-03", "품질", "CAPA AI 지원 근본원인 분석", "Lv2",
     "AI 도구로 품질 이슈의 근본원인을 체계적으로 분석하고 CAPA 문서를 효율적으로 작성할 수 있다"],
    ["MD-R-01", "R&D", "임상·비임상 데이터 AI 분석", "Lv2",
     "임상·비임상 데이터에서 AI로 유의미한 패턴을 도출하여 제품 개선 방향을 제시할 수 있다"],
    ["MD-R-02", "R&D", "AI/ML 의료기기 개발 프로세스 설계", "Lv3",
     "FDA GMLP에 맞는 AI/ML 의료기기 개발·검증 프로세스를 설계할 수 있다"],
    ["MD-G-01", "기획", "글로벌 의료기기 규제 AI 자동 추적·비교", "Lv2",
     "AI 도구로 다수 국가 의료기기 규제 변화를 자동 추적·비교하여 RA 전략에 반영할 수 있다"],
]

for i, row in enumerate(md_skills, 2):
    ws4.append(row)
    style_row(ws4, i, len(headers4))
    ws4.cell(row=i, column=4).fill = fill_for_level(row[3])
    ws4.cell(row=i, column=4).alignment = CENTER
    if i % 2 == 0:
        for c in [1, 2, 3]:
            ws4.cell(row=i, column=c).fill = FILL_GREY


# ══════════════════════════════════════════════════════════
# Sheet 5: 화장품
# ══════════════════════════════════════════════════════════
ws5 = wb.create_sheet("⑤화장품")
headers5 = ["스킬 ID", "직무", "스킬명", "수준", "행동 수준 (이걸 할 수 있다)"]
set_col_widths(ws5, [12, 10, 40, 8, 65])

ws5.append(headers5)
style_header(ws5, 1, len(headers5), fill=FILL_CM)

cm_skills = [
    ["CM-P-01", "생산", "CGMP 공정 데이터 표준 관리", "Lv2",
     "CGMP 기준에 맞게 유화·분산·충전 공정 데이터를 표준화하여 배치별 품질 추적과 AI 분석에 활용할 수 있다"],
    ["CM-P-02", "생산", "AI 기반 배합·유화 공정 최적화 도구 활용", "Lv2",
     "AI 예측 모델로 배합·유화 공정의 품질 결과를 사전 예측하고 최적 조건을 도출할 수 있다"],
    ["CM-Q-01", "품질", "AI 기반 성분 안전성 자동 스크리닝", "Lv2",
     "처방 내 전 성분에 대해 AI 스크리닝 도구로 안전성·규제 적합성을 자동 검증할 수 있다"],
    ["CM-Q-02", "품질", "클린뷰티 인증 기준 AI 자동 매칭", "Lv2",
     "AI 도구로 제품의 다수 인증 기준 충족 여부를 자동 검증하여 취득 가능한 인증을 매칭할 수 있다"],
    ["CM-Q-03", "품질", "안전성 예측 AI(QSAR) 활용", "Lv3",
     "QSAR AI 모델로 신원료·신처방의 독성을 사전 예측하여 동물실험 없이 안전성 평가 근거를 생성할 수 있다"],
    ["CM-R-01", "R&D", "AI 기반 처방 최적화 도구 활용", "Lv2",
     "목표 효능·제형 조건을 설정하고 AI 처방 최적화 도구로 최적 처방 후보를 도출할 수 있다"],
    ["CM-R-02", "R&D", "천연·바이오 원료 AI 탐색", "Lv2",
     "AI 도구로 천연·바이오 원료 DB에서 특정 효능 원료를 탐색하고 안전성·규제 적합성을 동시 확인할 수 있다"],
    ["CM-R-03", "R&D", "피부과학 데이터 AI 분석", "Lv2",
     "피부 측정 데이터·임상 사진을 AI로 분석하여 제품 효능을 정량 평가하고 시각화할 수 있다"],
    ["CM-G-01", "기획", "뷰티 트렌드·소비자 AI 소셜 리스닝", "Lv2",
     "AI 소셜 리스닝 도구로 뷰티 트렌드·소비자 니즈를 분석하여 데이터 기반 제품 기획 방향을 도출할 수 있다"],
    ["CM-G-02", "기획", "글로벌 화장품 규제 AI 자동 비교 분석", "Lv2",
     "다수 국가 화장품 규제 차이를 AI로 자동 비교·분석하여 글로벌 출시 전략에 반영할 수 있다"],
]

for i, row in enumerate(cm_skills, 2):
    ws5.append(row)
    style_row(ws5, i, len(headers5))
    ws5.cell(row=i, column=4).fill = fill_for_level(row[3])
    ws5.cell(row=i, column=4).alignment = CENTER
    if i % 2 == 0:
        for c in [1, 2, 3]:
            ws5.cell(row=i, column=c).fill = FILL_GREY


# ══════════════════════════════════════════════════════════
# Sheet 6: 정밀화학
# ══════════════════════════════════════════════════════════
ws6 = wb.create_sheet("⑥정밀화학")
headers6 = ["스킬 ID", "스킬명", "수준", "행동 수준 (이걸 할 수 있다)"]
set_col_widths(ws6, [12, 40, 8, 70])

ws6.append(headers6)
style_header(ws6, 1, len(headers6), fill=FILL_FC)

fc_skills = [
    ["FC-01", "화학 반응 조건 AI 최적화 도구 활용", "Lv2",
     "AI 도구로 반응 수율·선택성을 최적화하는 조건을 도출하고 실험실 검증을 거쳐 적용할 수 있다"],
    ["FC-02", "화학물질 AI 위험성 평가(QSAR) 실행", "Lv2",
     "AI 위험성 평가 도구로 화학물질의 유해성을 예측하고 규제 등록에 필요한 데이터를 생성할 수 있다"],
    ["FC-03", "분광 데이터 AI 자동 분석 수행", "Lv2",
     "IR, NMR, MS 등 분광 데이터를 AI로 자동 해석하여 물질 동정·정량을 수행할 수 있다"],
    ["FC-04", "AI 역합성 분석 기반 합성 경로 탐색", "Lv3",
     "역합성 AI로 목표 물질의 합성 경로를 탐색하고 경제성·안전성을 평가하여 최적 경로를 선정할 수 있다"],
]

for i, row in enumerate(fc_skills, 2):
    ws6.append(row)
    style_row(ws6, i, len(headers6))
    ws6.cell(row=i, column=3).fill = fill_for_level(row[2])
    ws6.cell(row=i, column=3).alignment = CENTER
    if i % 2 == 0:
        ws6.cell(row=i, column=1).fill = FILL_GREY
        ws6.cell(row=i, column=2).fill = FILL_GREY


# ══════════════════════════════════════════════════════════
# Sheet 7: 요약 통계
# ══════════════════════════════════════════════════════════
ws7 = wb.create_sheet("⑦요약 통계")
set_col_widths(ws7, [25, 12, 12, 12, 12, 12])

# 영역별 통계
ws7.cell(row=1, column=1, value="■ 전체 스킬 규모").font = TITLE_FONT
headers7a = ["구분", "수량"]
for c, h in enumerate(headers7a, 1):
    ws7.cell(row=2, column=c, value=h)
style_header(ws7, 2, 2)

summary_rows = [
    ["역량 영역", "4개"],
    ["역량 항목", "12개"],
    ["공통 스킬", f"{len(common_skills)}개"],
    ["바이오플라스틱 특화", f"{len(bp_skills)}개"],
    ["의료기기 특화", f"{len(md_skills)}개"],
    ["화장품 특화", f"{len(cm_skills)}개"],
    ["정밀화학 확장", f"{len(fc_skills)}개"],
    ["전체 스킬 수", f"{len(common_skills) + len(bp_skills) + len(md_skills) + len(cm_skills) + len(fc_skills)}개"],
]
for i, row in enumerate(summary_rows, 3):
    ws7.append(row)
    style_row(ws7, i, 2)
    if row[0] == "전체 스킬 수":
        ws7.cell(row=i, column=1).font = BODY_FONT_BOLD
        ws7.cell(row=i, column=2).font = BODY_FONT_BOLD

# 수준별 분포
r = ws7.max_row + 2
ws7.cell(row=r, column=1, value="■ 수준별 분포").font = TITLE_FONT
r += 1
headers7b = ["수준", "공통", "바이오플라스틱", "의료기기", "화장품", "정밀화학", "합계"]
set_col_widths(ws7, [25, 12, 16, 12, 12, 12, 10])
for c, h in enumerate(headers7b, 1):
    ws7.cell(row=r, column=c, value=h)
style_header(ws7, r, len(headers7b))

# 수준 카운팅
def count_by_level(skills, lv_col_idx):
    counts = {"Lv1": 0, "Lv2": 0, "Lv3": 0}
    for s in skills:
        lv = s[lv_col_idx]
        if lv in counts:
            counts[lv] += 1
    return counts

c_common = count_by_level(common_skills, 4)
c_bp = count_by_level(bp_skills, 3)
c_md = count_by_level(md_skills, 3)
c_cm = count_by_level(cm_skills, 3)
c_fc = count_by_level(fc_skills, 2)

for lv in ["Lv1", "Lv2", "Lv3"]:
    r += 1
    total = c_common[lv] + c_bp[lv] + c_md[lv] + c_cm[lv] + c_fc[lv]
    row_data = [lv, c_common[lv], c_bp[lv], c_md[lv], c_cm[lv], c_fc[lv], total]
    for c, v in enumerate(row_data, 1):
        ws7.cell(row=r, column=c, value=v)
    style_row(ws7, r, len(headers7b), fill=fill_for_level(lv))

r += 1
total_all = len(common_skills) + len(bp_skills) + len(md_skills) + len(cm_skills) + len(fc_skills)
row_total = ["합계", len(common_skills), len(bp_skills), len(md_skills), len(cm_skills), len(fc_skills), total_all]
for c, v in enumerate(row_total, 1):
    ws7.cell(row=r, column=c, value=v)
style_row(ws7, r, len(headers7b))
for c in range(1, len(headers7b) + 1):
    ws7.cell(row=r, column=c).font = BODY_FONT_BOLD

# 직무별 분포 (산업 특화만)
r = ws7.max_row + 2
ws7.cell(row=r, column=1, value="■ 산업특화 스킬 — 직무별 분포").font = TITLE_FONT
r += 1
headers7c = ["직무", "바이오플라스틱", "의료기기", "화장품", "합계"]
for c, h in enumerate(headers7c, 1):
    ws7.cell(row=r, column=c, value=h)
style_header(ws7, r, len(headers7c))

def count_by_job(skills, job_col):
    counts = {}
    for s in skills:
        job = s[job_col]
        counts[job] = counts.get(job, 0) + 1
    return counts

j_bp = count_by_job(bp_skills, 1)
j_md = count_by_job(md_skills, 1)
j_cm = count_by_job(cm_skills, 1)

for job in ["생산", "품질", "R&D", "제품기획", "기획"]:
    bp_v = j_bp.get(job, 0)
    md_v = j_md.get(job, 0)
    cm_v = j_cm.get(job, 0)
    if bp_v + md_v + cm_v > 0:
        r += 1
        row_data = [job, bp_v, md_v, cm_v, bp_v + md_v + cm_v]
        for c, v in enumerate(row_data, 1):
            ws7.cell(row=r, column=c, value=v)
        style_row(ws7, r, len(headers7c))


# ── 행 높이 자동 조정 (wrap text 대비) ──
for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7]:
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 45
    ws.row_dimensions[1].height = 30
    ws.sheet_properties.pageSetUpPr = None
    ws.freeze_panes = "A2"  # 헤더 고정


# ── 저장 ──
output_path = r"C:\Users\eykis\OneDrive\vrin_AI_hub\projects\P2026-008_KTR 훈련로드맵\outputs\AI_역량_스킬셋_v4.xlsx"
wb.save(output_path)
print(f"✅ 저장 완료: {output_path}")
print(f"   공통 스킬: {len(common_skills)}개")
print(f"   바이오플라스틱: {len(bp_skills)}개")
print(f"   의료기기: {len(md_skills)}개")
print(f"   화장품: {len(cm_skills)}개")
print(f"   정밀화학: {len(fc_skills)}개")
print(f"   전체: {len(common_skills) + len(bp_skills) + len(md_skills) + len(cm_skills) + len(fc_skills)}개")
